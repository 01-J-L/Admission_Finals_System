from flask import Blueprint, render_template, request, flash, redirect, session, url_for, jsonify, current_app, Response, send_file
import mysql.connector
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import datetime
from datetime import timedelta
import base64
import traceback
import secrets
from flask_mail import Message
import os
import io
import time
from collections import defaultdict
import pandas as pd
from flask import send_file
import os
from flask import send_from_directory
from werkzeug.utils import secure_filename
from collections import defaultdict
from flask import jsonify
import datetime
import calendar
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import subprocess # Required for running system commands
from flask import send_file, Response, current_app
import tempfile
import json
import shutil


# ----------------- BLUEPRINT SETUP -----------------
auth = Blueprint('auth', __name__)

auth = Blueprint('auth', __name__)

# Register the filter
@auth.app_template_filter('to_words')
def to_words_filter(amount):
    return amount_to_words(amount)
# ----------------- DATABASE CONNECTION -----------------
def get_db_connection():
    try:
        # print("--- Attempting to Read DB Environment Variables ---")
        env_db_host_from_os = os.getenv('DB_HOST')
        env_db_user_from_os = os.getenv('DB_USER')
        # For logging actual password existence, not value
        # env_db_password_exists_from_os = "Yes" if os.getenv('DB_PASSWORD') else "No"
        env_db_password_from_os = os.getenv('DB_PASSWORD') # Actual password value from env
        env_db_name_from_os = os.getenv('DB_NAME')
        env_db_port_str_from_os = os.getenv('DB_PORT') # Port from env is a string

        # print(f"Raw os.getenv('DB_HOST'): {env_db_host_from_os}")
        # print(f"Raw os.getenv('DB_USER'): {env_db_user_from_os}")
        # print(f"Raw os.getenv('DB_PASSWORD') exists: {env_db_password_exists_from_os}")
        # print(f"Raw os.getenv('DB_NAME'): {env_db_name_from_os}")
        # print(f"Raw os.getenv('DB_PORT'): {env_db_port_str_from_os}")
        # print("--- End of Reading DB Environment Variables ---")

        # Railway specific defaults (derived from typical Railway setups and user-provided connection string)
        RAILWAY_DEFAULT_HOST = 'localhost'
        RAILWAY_DEFAULT_USER = 'root'
        RAILWAY_DEFAULT_PASSWORD = "anime951827"
        RAILWAY_DEFAULT_DB_NAME = "expired" # As per user's connection string example
        RAILWAY_DEFAULT_PORT_STR = '3306' # As per user's connection string example
        
        db_host = env_db_host_from_os
        if env_db_host_from_os and env_db_host_from_os.lower() == 'localhost':
            # print(f"Warning: DB_HOST from environment is '{env_db_host_from_os}'. This is often incorrect for a remote Railway database. "
                #   f"Overriding to use the default Railway proxy host: '{RAILWAY_DEFAULT_HOST}'.")
            db_host = RAILWAY_DEFAULT_HOST
        elif not env_db_host_from_os:
            # print(f"DB_HOST not set in environment. Using default Railway proxy host: '{RAILWAY_DEFAULT_HOST}'.")
            db_host = RAILWAY_DEFAULT_HOST

        db_user = env_db_user_from_os or RAILWAY_DEFAULT_USER
        db_password = env_db_password_from_os or RAILWAY_DEFAULT_PASSWORD
        db_name = env_db_name_from_os or RAILWAY_DEFAULT_DB_NAME
        
        effective_port_str = env_db_port_str_from_os or RAILWAY_DEFAULT_PORT_STR
        
        db_port = int(RAILWAY_DEFAULT_PORT_STR) 
        try:
            db_port = int(effective_port_str)
        except (ValueError, TypeError):
            # print(f"CRITICAL: The determined DB_PORT string ('{effective_port_str}') is not a valid integer. "
                #   f"Falling back to the default Railway port number: {RAILWAY_DEFAULT_PORT_STR}.")
            pass # db_port is already set
        
        # print(f"Attempting DB connection with: Host='{db_host}', Port={db_port}, User='{db_user}', Database='{db_name}'")

        conn = mysql.connector.connect(
            host=db_host,
            user=db_user,
            password=db_password,
            database=db_name,
            port=db_port
        )
        # print("Database connection successful.")
        return conn
    except mysql.connector.Error as err:
        print(f"Database connection error: {err}")
        traceback.print_exc()
        return None

# ----------------- FILE PROCESSING HELPER -----------------
def process_uploaded_file(file_storage, file_description_for_error, max_size_mb=5):
    if file_storage and file_storage.filename != '':
        filename = secure_filename(file_storage.filename)
        mimetype = file_storage.mimetype

        allowed_mimetypes = ['image/jpeg', 'image/png', 'image/jpg', 'image/webp', 'application/pdf']
        allowed_extensions = ['.jpg', '.jpeg', '.png', '.webp', '.pdf']
        file_ext = os.path.splitext(filename)[1].lower()

        is_image = mimetype in ['image/jpeg', 'image/png', 'image/jpg', 'image/webp'] or file_ext in ['.jpg', '.jpeg', '.png', '.webp']
        is_pdf = mimetype == 'application/pdf' or file_ext == '.pdf'

        if not (is_image or is_pdf):
             return None, None, None, f"Invalid file type for {file_description_for_error} ('{filename}'). Allowed: PDF, JPG, PNG, WEBP."

        file_storage.seek(0, os.SEEK_END)
        file_length = file_storage.tell()
        file_storage.seek(0)

        if file_length == 0:
            return None, None, None, f"{file_description_for_error} ('{filename}') is empty. Please upload a valid file."

        max_bytes = max_size_mb * 1024 * 1024
        if file_length > max_bytes:
            return None, None, None, f"{file_description_for_error} ('{filename}') is too large (max {max_size_mb}MB)."

        file_data = file_storage.read()
        return file_data, filename, mimetype, None
    return None, None, None, None


def save_image_to_uploads(file_storage, prefix='img'):
    """Saves an image file to the UPLOAD_FOLDER and returns the new filename."""
    if not file_storage or not file_storage.filename:
        return None, "No file provided."

    allowed_extensions = {'.png', '.jpg', '.jpeg', '.webp'}
    file_ext = os.path.splitext(file_storage.filename)[1].lower()

    if file_ext not in allowed_extensions:
        return None, f"Invalid image type. Allowed: {', '.join(allowed_extensions)}"
    
    # Create a unique filename to prevent overwrites and conflicts
    timestamp = int(time.time())
    secure_name = secure_filename(os.path.splitext(file_storage.filename)[0])
    new_filename = f"{prefix}_{secure_name}_{timestamp}{file_ext}"
    
    try:
        save_path = os.path.join(current_app.config['UPLOAD_FOLDER'], new_filename)
        file_storage.save(save_path)
        return new_filename, None
    except Exception as e:
        print(f"Error saving file: {e}")
        traceback.print_exc()
        return None, f"Server error while saving file: {e}"

def delete_image_from_uploads(filename):
    """Deletes an image file from the UPLOAD_FOLDER."""
    if not filename:
        return
    try:
        file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
        if os.path.exists(file_path):
            os.remove(file_path)
    except Exception as e:
        print(f"Error deleting file {filename}: {e}")
        traceback.print_exc()

def save_file_to_uploads(file_storage, prefix='doc'):
    """Saves a file (PDF/Image) to UPLOAD_FOLDER."""
    if not file_storage or not file_storage.filename:
        return None, "No file provided."

    allowed_extensions = {'.png', '.jpg', '.jpeg', '.webp', '.pdf'}
    file_ext = os.path.splitext(file_storage.filename)[1].lower()

    if file_ext not in allowed_extensions:
        return None, f"Invalid file type. Allowed: PDF, JPG, PNG."
    
    timestamp = int(time.time())
    secure_name = secure_filename(os.path.splitext(file_storage.filename)[0])
    new_filename = f"{prefix}_{secure_name}_{timestamp}{file_ext}"
    
    try:
        save_path = os.path.join(current_app.config['UPLOAD_FOLDER'], new_filename)
        file_storage.save(save_path)
        return new_filename, None
    except Exception as e:
        return None, str(e)
    
# --- Helper: Amount to Words ---
def amount_to_words(n):
    """Converts a number to words (e.g., 500 -> Five Hundred)."""
    try:
        n = float(n)
    except:
        return "Invalid Amount"
        
    units = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine"]
    teens = ["", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen", "Seventeen", "Eighteen", "Nineteen"]
    tens = ["", "Ten", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]
    thousands = ["", "Thousand", "Million", "Billion"]

    def num999(n):
        c = n % 10
        b = ((n % 100) - c) // 10
        a = ((n % 1000) - (b * 10) - c) // 100
        t = ""
        if a != 0:
            t += units[a] + " Hundred "
        if b <= 1:
            if n % 100 == 0: return t
            if b == 0: t += units[c]
            if b == 1:
                if c == 0: t += "Ten"
                else: t += teens[c]
        else:
            t += tens[b]
            if c > 0: t += " " + units[c]
        return t

    if n == 0: return "Zero Pesos"
    
    parts = []
    # Integer part
    i = 0
    int_part = int(n)
    while int_part > 0:
        if int_part % 1000 != 0:
            parts.insert(0, num999(int_part % 1000) + " " + thousands[i])
        int_part //= 1000
        i += 1
    
    text = " ".join(parts).strip() + " Pesos"

    # Decimal part (Centavos)
    cents = int(round((n - int(n)) * 100))
    if cents > 0:
        text += f" and {cents}/100"
    
    return text

# ----------------- EMAIL otp -----------------

def send_otp_email(user_email, otp_code):
    # !!! THERE ARE NO CHECKS HERE !!! 
    # It goes straight to defining sender and subject.

    sender_name_from_config = current_app.config.get('MAIL_SENDER_NAME', 'Padre Garcia Polytechnic College')
    subject = f"Verify Your PGPC Account - OTP: {otp_code} - {sender_name_from_config}"
    
    try:
        verify_url = url_for('auth.verify_otp_page', _external=True) 
    except RuntimeError:
        verify_url = f"{current_app.config.get('PREFERRED_URL_SCHEME', 'http')}://{current_app.config.get('SERVER_NAME', 'your-domain.com')}/verify-otp"

    try:
        html_body = render_template(
            'email/otp_verification_email.html',
            user_email=user_email,
            otp_code=otp_code,
            verify_url=verify_url,
            sender_name=sender_name_from_config,
            now=datetime.datetime.now()
        )
    except Exception as e_template:
        print(f"CRITICAL: Email template 'email/otp_verification_email.html' not found or error rendering. Error: {e_template}")
        html_body = f"<p>Dear User,</p><p>Your One-Time Password (OTP) for account verification is: <strong>{otp_code}</strong>. Please enter this OTP on the verification page. This OTP is valid for 10 minutes.</p><p>If you didn't request this, please ignore this email.</p>"
        
    # It goes straight to sending.
    return _send_email(subject, [user_email], html_body, sender_name_override=sender_name_from_config)


@auth.route('/admin/student/statement/<int:applicant_id>')
def admin_view_student_statement(applicant_id):
    # Allow multiple roles to view statement
    allowed_roles = ['admin', 'registrar', 'cashier', 'accounting', 'guidance', 'osa']
    if session.get('user_role') not in allowed_roles:
        return redirect(url_for('auth.admin'))

    conn = None; cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # 1. Fetch Student Info
        cursor.execute("""
            SELECT applicant_id, first_name, last_name, old_student_id, control_number, 
                   program_choice, application_status
            FROM applicants 
            WHERE applicant_id = %s
        """, (applicant_id,))
        student = cursor.fetchone()
        
        if not student:
            flash("Student not found.", "danger")
            return redirect(request.referrer)
            
        student['display_id'] = student['old_student_id'] or student['control_number'] or 'N/A'

        # 2. Fetch Assessments (All, sorted by date)
        cursor.execute("SELECT * FROM assessments WHERE student_id = %s ORDER BY created_at DESC", (applicant_id,))
        assessments = cursor.fetchall()

        # 3. Fetch Payments (All, sorted by date)
        cursor.execute("SELECT * FROM payments WHERE student_id = %s ORDER BY payment_date DESC", (applicant_id,))
        payments = cursor.fetchall()

        # 4. Calculate Total Balance
        total_balance = sum(item['balance'] for item in assessments)

        return render_template('admin_view_statement.html',
                               student=student,
                               assessments=assessments,
                               payments=payments,
                               total_balance=total_balance,
                               today_date=datetime.date.today().strftime('%B %d, %Y'))

    except Exception as e:
        print(f"Error viewing statement: {e}")
        flash("An error occurred.", "danger")
        return redirect(request.referrer)
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

# ----------------- BACKUP & RESTORE -----------------

@auth.route('/admin/system/backup-restore')
def admin_backup_restore_page():
    if session.get('user_role') != 'admin': # Strictly Admin only
        return redirect(url_for('auth.admin_dashboard'))
    
    _, stats = _get_all_applications_and_stats()
    return render_template('admin_backup_restore.html', stats=stats, active_page='backup_restore')

@auth.route('/admin/system/download-backup', methods=['POST'])
def admin_download_backup():
    if session.get('user_role') != 'admin': return redirect(url_for('auth.admin'))

    # Get DB Credentials from Environment
    db_host = os.getenv('DB_HOST') or 'localhost'
    db_user = os.getenv('DB_USER') or 'root'
    db_password = os.getenv('DB_PASSWORD') or ''
    db_name = os.getenv('DB_NAME') or 'expired'
    db_port = os.getenv('DB_PORT') or '3306'

    # Create a temporary file to store the dump
    with tempfile.NamedTemporaryFile(suffix='.sql', delete=False) as temp_file:
        temp_file_path = temp_file.name

    try:
        # 1. Define the full path to mysqldump.exe for MySQL Workbench/Server users
        # Default path for MySQL Server 8.0 on Windows
        mysql_dump_exe = r"C:\Program Files\MySQL\MySQL Server 8.0\bin\mysqldump.exe"

        # Check if the file exists at that path. If not, fallback to just 'mysqldump'
        if not os.path.exists(mysql_dump_exe):
            if shutil.which('mysqldump'):
                mysql_dump_exe = 'mysqldump'
            else:
                # Fallback for XAMPP users just in case
                xampp_path = r"C:\xampp\mysql\bin\mysqldump.exe"
                if os.path.exists(xampp_path):
                    mysql_dump_exe = xampp_path
                else:
                    raise Exception("mysqldump.exe not found. Please add MySQL bin to System PATH or check auth.py paths.")

        # 2. Prepare Environment Variables (Safest way to pass password)
        env = os.environ.copy()
        env['MYSQL_PWD'] = db_password

        # 3. Construct Command
        # We use quotes around the executable path in case of spaces in "Program Files"
        dump_cmd = f'"{mysql_dump_exe}" -h {db_host} -P {db_port} -u {db_user} {db_name} > "{temp_file_path}"'

        # 4. Run Subprocess
        # shell=True is required for the '>' redirection to work on Windows
        subprocess.run(dump_cmd, shell=True, env=env, check=True)

        # 5. Send the file to the user
        filename = f"backup_pgpc_{datetime.datetime.now().strftime('%Y-%m-%d_%H-%M')}.sql"
        
        with open(temp_file_path, 'rb') as f:
            data = f.read()
            
        # Clean up temp file immediately after reading
        os.remove(temp_file_path)

        return Response(
            data,
            mimetype="application/sql",
            headers={"Content-Disposition": f"attachment;filename={filename}"}
        )

    except subprocess.CalledProcessError as e:
        print(f"Backup Error: {e}")
        flash(f"Error generating backup. Details: {e}", "danger")
        return redirect(url_for('auth.admin_backup_restore_page'))
    except Exception as e:
        print(f"General Backup Error: {e}")
        flash(f"An unexpected error occurred: {e}", "danger")
        return redirect(url_for('auth.admin_backup_restore_page'))
    
@auth.route('/admin/system/restore-data', methods=['POST'])
def admin_restore_data():
    if session.get('user_role') != 'admin': return redirect(url_for('auth.admin'))

    file = request.files.get('backup_file')
    if not file or not file.filename:
        flash("No file selected.", "warning")
        return redirect(url_for('auth.admin_backup_restore_page'))

    if not file.filename.endswith('.sql'):
        flash("Invalid file type. Please upload a .sql file.", "danger")
        return redirect(url_for('auth.admin_backup_restore_page'))

    # Get DB Credentials
    db_host = os.getenv('DB_HOST') or 'localhost'
    db_user = os.getenv('DB_USER') or 'root'
    db_password = os.getenv('DB_PASSWORD') or ''
    db_name = os.getenv('DB_NAME') or 'expired'
    db_port = os.getenv('DB_PORT') or '3306'

    try:
        # 1. Define the full path to mysql.exe
        mysql_exe = r"C:\Program Files\MySQL\MySQL Server 8.0\bin\mysql.exe"

        # Check path existence
        if not os.path.exists(mysql_exe):
            if shutil.which('mysql'):
                mysql_exe = 'mysql'
            else:
                 # Fallback for XAMPP
                xampp_path = r"C:\xampp\mysql\bin\mysql.exe"
                if os.path.exists(xampp_path):
                    mysql_exe = xampp_path
                else:
                    raise Exception("mysql.exe not found.")

        # 2. Prepare environment
        env = os.environ.copy()
        env['MYSQL_PWD'] = db_password

        # 3. Construct Command List
        # Note: When using a list with subprocess.Popen, we don't need quotes around paths, Python handles it.
        restore_cmd = [
            mysql_exe, 
            '-h', db_host, 
            '-P', str(db_port), 
            '-u', db_user, 
            db_name
        ]

        # 4. Run Process
        # We pipe the file content directly to standard input
        process = subprocess.Popen(restore_cmd, stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE, env=env)
        stdout, stderr = process.communicate(input=file.read())

        if process.returncode != 0:
            error_msg = stderr.decode('utf-8')
            print(f"Restore Error Output: {error_msg}")
            flash(f"Restore failed: {error_msg}", "danger")
        else:
            flash("Database restored successfully! Please log in again if user credentials were changed.", "success")

    except Exception as e:
        print(f"General Restore Error: {e}")
        flash(f"An error occurred during restore: {e}", "danger")

    return redirect(url_for('auth.admin_backup_restore_page'))

@auth.route('/admin/users/toggle-permission', methods=['POST'])
def admin_toggle_user_permission():
    # Only Main Admin can change permissions
    if session.get('user_role') != 'admin':
        return jsonify({"success": False, "message": "Unauthorized"}), 401

    data = request.get_json()
    user_id = data.get('user_id')
    can_edit = data.get('can_edit')

    conn = None; cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # Prevent disabling the main admin's own rights
        cursor.execute("SELECT role FROM system_users WHERE id = %s", (user_id,))
        target_user = cursor.fetchone()
        if target_user and target_user['role'] == 'admin':
             return jsonify({"success": False, "message": "Cannot restrict Administrator account."}), 403

        # Update permission
        cursor.execute("UPDATE system_users SET can_edit = %s WHERE id = %s", (can_edit, user_id))
        conn.commit()
        
        status_text = "enabled" if can_edit else "disabled"
        return jsonify({"success": True, "message": f"Edit privilege {status_text}."})

    except Exception as e:
        print(f"Error toggling permission: {e}")
        return jsonify({"success": False, "message": "Database error"}), 500
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

@auth.route('/admin/manage-downloads', methods=['GET', 'POST'])
def admin_manage_downloads():
    if session.get('user_role') not in ['admin', 'registrar', 'secretary']: return redirect(url_for('auth.admin'))
    
    conn = None
    cursor = None
    
    # Configuration for download folder
    DOWNLOAD_FOLDER = os.path.join(current_app.static_folder, 'uploads', 'downloads')
    os.makedirs(DOWNLOAD_FOLDER, exist_ok=True) 

    try:
        conn = get_db_connection()
        
        # --- HANDLE POST (Upload) ---
        if request.method == 'POST':
            title = request.form.get('title')
            category = request.form.get('category')
            
            # New Targeting Fields
            target_program = request.form.get('target_program') or 'All'
            target_year_level = request.form.get('target_year_level') or 'All'
            target_section = request.form.get('target_section') or 'All'
            
            file = request.files.get('file')

            if not title or not category or not file:
                flash("Title, Category, and File are required.", "danger")
            else:
                filename = secure_filename(file.filename)
                timestamp = int(time.time())
                save_name = f"{timestamp}_{filename}"
                
                try:
                    file.save(os.path.join(DOWNLOAD_FOLDER, save_name))
                    
                    cursor = conn.cursor()
                    # Updated Query to include targeting fields
                    cursor.execute("""
                        INSERT INTO student_downloads 
                        (category, title, filename, target_program, target_year_level, target_section) 
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, (category, title, save_name, target_program, target_year_level, target_section))
                    conn.commit()
                    flash("File uploaded successfully.", "success")
                except Exception as e:
                    print(f"Upload Error: {e}")
                    flash("Error saving file.", "danger")
            
            return redirect(url_for('auth.admin_manage_downloads'))

        # --- HANDLE GET (List Files & Filter Options) ---
        cursor = conn.cursor(dictionary=True)
        
        # 1. Fetch Files
        cursor.execute("SELECT * FROM student_downloads ORDER BY category, uploaded_at DESC")
        files = cursor.fetchall()
        
        grouped_files = defaultdict(list)
        for f in files:
            grouped_files[f['category']].append(f)

        # 2. Fetch Options for Dropdowns
        cursor.execute("SELECT title FROM programs ORDER BY title")
        programs = cursor.fetchall()
        
        cursor.execute("SELECT DISTINCT section_name FROM sections ORDER BY section_name")
        sections = cursor.fetchall()

        _, stats = _get_all_applications_and_stats()
        
        return render_template('admin_manage_downloads.html', 
                               grouped_files=grouped_files, 
                               programs=programs,
                               sections=sections,
                               stats=stats, 
                               active_page='manage_content')

    except Exception as e:
        print(f"Error in manage downloads: {e}")
        traceback.print_exc()
        flash(f"Error: {e}", "danger")
        return redirect(url_for('auth.admin_manage_content'))
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

@auth.route('/admin/downloads/edit/<int:file_id>', methods=['POST'])
def admin_edit_download(file_id):
    if session.get('user_role') not in ['admin', 'registrar', 'secretary']: return redirect(url_for('auth.admin'))
    
    conn = None
    cursor = None
    DOWNLOAD_FOLDER = os.path.join(current_app.static_folder, 'uploads', 'downloads')

    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # 1. Collect Form Data
        title = request.form.get('title')
        category = request.form.get('category')
        target_program = request.form.get('target_program')
        target_year_level = request.form.get('target_year_level')
        target_section = request.form.get('target_section')
        file = request.files.get('file') # Optional

        if not title or not category:
            flash("Title and Category are required.", "danger")
            return redirect(url_for('auth.admin_manage_downloads'))

        # 2. Handle File Replacement (if a new file is uploaded)
        filename_update_sql = ""
        params = [title, category, target_program, target_year_level, target_section]

        if file and file.filename:
            # Get old filename to delete
            cursor.execute("SELECT filename FROM student_downloads WHERE id = %s", (file_id,))
            old_record = cursor.fetchone()
            
            # Save new file
            filename = secure_filename(file.filename)
            timestamp = int(time.time())
            save_name = f"{timestamp}_{filename}"
            file.save(os.path.join(DOWNLOAD_FOLDER, save_name))
            
            # Delete old file
            if old_record and old_record['filename']:
                old_path = os.path.join(DOWNLOAD_FOLDER, old_record['filename'])
                if os.path.exists(old_path):
                    os.remove(old_path)
            
            filename_update_sql = ", filename = %s"
            params.append(save_name)

        # 3. Update Database
        params.append(file_id) # Add ID for WHERE clause
        
        query = f"""
            UPDATE student_downloads 
            SET title = %s, category = %s, target_program = %s, target_year_level = %s, target_section = %s
            {filename_update_sql}
            WHERE id = %s
        """
        
        cursor.execute(query, tuple(params))
        conn.commit()
        flash("Resource updated successfully.", "success")

    except Exception as e:
        print(f"Update Error: {e}")
        flash("Error updating resource.", "danger")
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

    return redirect(url_for('auth.admin_manage_downloads'))

@auth.route('/admin/downloads/delete/<int:file_id>', methods=['POST'])
def admin_delete_download(file_id):
    if session.get('user_role') not in ['admin', 'registrar', 'secretary']: return redirect(url_for('auth.admin'))
    
    conn = None
    cursor = None
    DOWNLOAD_FOLDER = os.path.join(current_app.static_folder, 'uploads', 'downloads')

    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get filename to delete from disk
        cursor.execute("SELECT filename FROM student_downloads WHERE id = %s", (file_id,))
        file_record = cursor.fetchone()
        
        if file_record:
            # Delete from DB
            cursor.execute("DELETE FROM student_downloads WHERE id = %s", (file_id,))
            conn.commit()
            
            # Delete from Disk
            file_path = os.path.join(DOWNLOAD_FOLDER, file_record['filename'])
            if os.path.exists(file_path):
                os.remove(file_path)
                
            flash("Resource deleted.", "success")
        else:
            flash("File not found.", "warning")

    except Exception as e:
        flash(f"Error deleting file: {e}", "danger")
    finally:
        if cursor: cursor.close()
        if conn: conn.close()
        
    return redirect(url_for('auth.admin_manage_downloads'))

@auth.route('/student/messenger', methods=['GET', 'POST'])
def student_messenger():
    if 'student_id' not in session: return redirect(url_for('auth.student_login_page'))
    
    # Get specific department from URL or default to registrar
    dept = request.args.get('dept', 'registrar') 
    
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    cursor.execute("SELECT applicant_id, first_name, last_name FROM applicants WHERE student_user_id = %s", (session['student_id'],))
    app = cursor.fetchone()
    
    if request.method == 'POST':
        msg_text = request.form.get('message')
        if msg_text and app:
            cursor.execute("""
                INSERT INTO application_messages (applicant_id, sender_type, sender_name, admin_role, message_text)
                VALUES (%s, 'student', %s, %s, %s)
            """, (app['applicant_id'], f"{app['first_name']} {app['last_name']}", dept, msg_text))
            conn.commit()
        return redirect(url_for('auth.student_messenger', dept=dept))

    # Fetch ONLY messages for the selected department (The DM logic)
    cursor.execute("""
        SELECT * FROM application_messages 
        WHERE applicant_id = %s AND admin_role = %s 
        ORDER BY created_at ASC
    """, (app['applicant_id'], dept))
    messages = cursor.fetchall()
    
    # Mark as read
    cursor.execute("UPDATE application_messages SET is_read = TRUE WHERE applicant_id = %s AND sender_type = 'admin' AND admin_role = %s", (app['applicant_id'], dept))
    conn.commit()
    
    cursor.close()
    conn.close()
    return render_template('student_messenger.html', messages=messages, application=app, current_dept=dept, student_logged_in=True)


@auth.route('/admin/messenger')
def admin_messenger_list():
    role = session.get('user_role')
    if role not in ['admin', 'registrar', 'guidance', 'osa']: return redirect(url_for('auth.admin'))
    
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # ADMINS ONLY SEE MESSAGES TAGGED FOR THEIR ROLE
    cursor.execute("""
        SELECT a.applicant_id, a.first_name, a.last_name,
        (SELECT message_text FROM application_messages WHERE applicant_id = a.applicant_id AND admin_role = %s ORDER BY created_at DESC LIMIT 1) as last_msg,
        (SELECT COUNT(*) FROM application_messages WHERE applicant_id = a.applicant_id AND is_read = FALSE AND sender_type = 'student' AND admin_role = %s) as unread_count
        FROM applicants a
        WHERE EXISTS (SELECT 1 FROM application_messages WHERE applicant_id = a.applicant_id AND admin_role = %s)
        ORDER BY unread_count DESC
    """, (role, role, role))
    chat_list = cursor.fetchall()
    
    cursor.close()
    conn.close()
    return render_template('admin_messenger.html', chat_list=chat_list, active_page='messenger')

@auth.route('/admin/messenger/chat/<int:applicant_id>', methods=['GET', 'POST'])
def admin_chat_view(applicant_id):
    role = session.get('user_role')
    if role not in ['admin', 'registrar', 'guidance', 'osa']: return redirect(url_for('auth.admin'))
    
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    if request.method == 'POST':
        msg_text = request.form.get('message')
        cursor.execute("""
            INSERT INTO application_messages (applicant_id, sender_type, sender_name, admin_role, message_text)
            VALUES (%s, 'admin', %s, %s, %s)
        """, (applicant_id, session.get('user_name'), role, msg_text))
        conn.commit()
        return redirect(url_for('auth.admin_chat_view', applicant_id=applicant_id))

    cursor.execute("SELECT first_name, last_name FROM applicants WHERE applicant_id = %s", (applicant_id,))
    student = cursor.fetchone()
    
    cursor.execute("SELECT * FROM application_messages WHERE applicant_id = %s ORDER BY created_at ASC", (applicant_id,))
    messages = cursor.fetchall()
    
    # Mark student messages as read
    cursor.execute("UPDATE application_messages SET is_read = TRUE WHERE applicant_id = %s AND sender_type = 'student'", (applicant_id,))
    conn.commit()
    
    cursor.close()
    conn.close()
    return render_template('admin_chat_window.html', messages=messages, student=student, applicant_id=applicant_id)

# --- NEW/MODIFIED ACADEMIC TERM & SETTINGS MANAGEMENT ---
@auth.route('/admin/manage-academic-term', methods=['GET', 'POST'])
def admin_manage_academic_term():
    if session.get('user_role') not in ['admin', 'registrar']: return redirect(url_for('auth.admin'))

    # Handle POST for updating the enrollment year
    if request.method == 'POST':
        new_year = request.form.get('enrollment_year')
        if new_year and new_year.isdigit() and len(new_year) == 4:
            conn_update = None
            cursor_update = None
            try:
                conn_update = get_db_connection()
                cursor_update = conn_update.cursor()
                cursor_update.execute("UPDATE system_settings SET setting_value = %s WHERE setting_key = 'current_enrollment_year'", (new_year,))
                conn_update.commit()
                flash(f"Current Enrollment Year has been updated to {new_year}.", "success")
            except Exception as e:
                flash(f"Error updating enrollment year: {e}", "danger")
            finally:
                if cursor_update: cursor_update.close()
                if conn_update: conn_update.close()
        else:
            flash("Invalid year format. Please enter a 4-digit year.", "danger")
        return redirect(url_for('auth.admin_manage_academic_term'))

    # Handle GET request to display the page
    academic_terms = []
    settings = {}
    conn = None; cursor = None
    try:
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT * FROM academic_terms ORDER BY year_name ASC, semester ASC")
            academic_terms = cursor.fetchall()
            
            cursor.execute("SELECT setting_key, setting_value FROM system_settings")
            for row in cursor.fetchall():
                settings[row['setting_key']] = row['setting_value']
    except Exception as e:
        flash(f"Error fetching page data: {e}", "danger")
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

    _, stats = _get_all_applications_and_stats()
    return render_template('admin_manage_school_year.html', 
                           academic_terms=academic_terms,
                           settings=settings, 
                           stats=stats, 
                           active_page='manage_academic_term')

@auth.route('/admin/academic-term/add', methods=['POST'])
def admin_add_academic_term():
    if session.get('user_role') not in ['admin', 'registrar']: return redirect(url_for('auth.admin'))
    
    year_name = request.form.get('year_name')
    semester = request.form.get('semester')
    max_units = request.form.get('max_units', 26) # Default 26

    if not year_name or not year_name.strip() or not semester:
        flash("Academic year and semester cannot be empty.", "danger")
        return redirect(url_for('auth.admin_manage_academic_term'))

    conn = None; cursor = None
    try:
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            # Updated INSERT query
            cursor.execute("INSERT INTO academic_terms (year_name, semester, is_active, max_units) VALUES (%s, %s, %s, %s)", (year_name.strip(), semester, False, max_units))
            conn.commit()
            flash(f"Academic Term '{year_name} - {semester}' added successfully.", "success")
    except mysql.connector.Error as err:
        if err.errno == 1062: # Duplicate entry
            flash(f"Academic Term '{year_name} - {semester}' already exists.", "danger")
        else:
            flash(f"Database error: {err}", "danger")
    except Exception as e:
        flash(f"An error occurred: {e}", "danger")
    finally:
        if cursor: cursor.close()
        if conn: conn.close()
        
    return redirect(url_for('auth.admin_manage_academic_term'))

@auth.route('/admin/academic-term/update/<int:term_id>', methods=['POST'])
def admin_update_term_units(term_id):
    if session.get('user_role') not in ['admin', 'registrar']: return redirect(url_for('auth.admin'))
    
    new_max = request.form.get('max_units')
    conn = None; cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("UPDATE academic_terms SET max_units = %s WHERE id = %s", (new_max, term_id))
        conn.commit()
        flash("Max units updated.", "success")
    except Exception as e:
        flash(f"Error: {e}", "danger")
    finally:
        if cursor: cursor.close()
        if conn: conn.close()
    return redirect(url_for('auth.admin_manage_academic_term'))

@auth.route('/admin/academic-term/toggle/<int:term_id>', methods=['POST'])
def admin_toggle_academic_term(term_id):
    if session.get('user_role') not in ['admin', 'registrar']: return redirect(url_for('auth.admin'))

    action = request.form.get('action')
    conn = None; cursor = None
    try:
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor(dictionary=True, buffered=True)

            if action == 'open':
                # 1. OPENING A NEW TERM
                # Logic: Only move students who are currently 'Enrolled'.
                # 'Dropped' and 'Not Enrolled' students are ignored by the WHERE clause.

                cursor.execute("SELECT COUNT(*) as count FROM applicants WHERE application_status = 'Enrolled'")
                count_check = cursor.fetchone()
                print(f"DEBUG: Attempting to move {count_check['count']} Enrolled students to Eligible.")
                
                cursor.execute("""
                    UPDATE applicants 
                    SET application_status = 'Eligible for Enrollment', 
                        original_enrollment_status = 'Enrolled',
                        original_academic_term = CONCAT(academic_year, ' - ', enrollment_semester)
                    WHERE application_status = 'Enrolled'
                """)

                print(f"DEBUG: Rows updated: {cursor.rowcount}")
                
                # Activate the new term
                cursor.execute("UPDATE academic_terms SET is_active = FALSE")
                cursor.execute("UPDATE academic_terms SET is_active = TRUE WHERE id = %s", (term_id,))
                
                flash(f"Term opened. {cursor.rowcount} student(s) moved to 'Eligible for Enrollment'.", "success")

            elif action == 'close':
                # 2. CLOSING/REVERTING A TERM
                # Logic: Only revert students who are currently waiting in 'Eligible for Enrollment'
                # AND were originally 'Enrolled'.
                
                cursor.execute("""
                    UPDATE applicants 
                    SET application_status = 'Enrolled', 
                        original_enrollment_status = NULL,
                        original_academic_term = NULL
                    WHERE application_status = 'Eligible for Enrollment'
                      AND original_enrollment_status = 'Enrolled'
                """)
                
                # Close the specified term
                cursor.execute("UPDATE academic_terms SET is_active = FALSE WHERE id = %s", (term_id,))
                flash("Academic term has been closed. Eligible students reverted to Enrolled status.", "info")
            
            conn.commit()
    except Exception as e:
        flash(f"An error occurred: {e}", "danger")
        traceback.print_exc()
        if conn: conn.rollback()
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

    return redirect(url_for('auth.admin_manage_academic_term'))

@auth.route('/student/initiate-enrollment', methods=['POST'])
def student_initiate_enrollment():
    if 'student_id' not in session:
        return redirect(url_for('auth.student_login_page'))
        
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        student_user_id = session['student_id']
        
        # Check if student is eligible
        cursor.execute("""
            SELECT applicant_id FROM applicants 
            WHERE student_user_id = %s AND application_status = 'Eligible for Enrollment'
        """, (student_user_id,))
        app = cursor.fetchone()
        
        if app:
            # REMOVED: status update query.
            # We just send them to the form. Their status remains "Eligible for Enrollment"
            return redirect(url_for('views.enrollment_form_page', applicant_id=app['applicant_id']))
        else:
            # Fallback for those already in Enrolling status
            cursor.execute("""
                SELECT applicant_id FROM applicants 
                WHERE student_user_id = %s AND application_status = 'Enrolling'
            """, (student_user_id,))
            already = cursor.fetchone()
            if already:
                return redirect(url_for('views.enrollment_form_page', applicant_id=already['applicant_id']))
                
            flash("You are not currently eligible to initiate enrollment.", "warning")
            return redirect(url_for('views.application_status_page'))
            
    except Exception as e:
        print(f"Error initiating enrollment: {e}")
        return redirect(url_for('views.application_status_page'))
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

@auth.route('/admin/academic-term/delete/<int:term_id>', methods=['POST'])
def admin_delete_academic_term(term_id):
    if session.get('user_role') not in ['admin', 'registrar']:
        return redirect(url_for('auth.admin'))

    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor(dictionary=True)
            
            # Check if any applications are associated with this term
            cursor.execute("SELECT year_name, semester FROM academic_terms WHERE id = %s", (term_id,))
            term_to_delete = cursor.fetchone()
            if term_to_delete:
                cursor.execute("SELECT COUNT(*) as count FROM applicants WHERE academic_year = %s", (term_to_delete['year_name'],))
                app_count = cursor.fetchone()['count']

                if app_count > 0:
                    flash(f"Cannot delete term '{term_to_delete['year_name']} - {term_to_delete['semester']}' because {app_count} application(s) are associated with it.", "danger")
                    return redirect(url_for('auth.admin_manage_academic_term'))

            # If no applications are associated, proceed with deletion
            cursor.execute("DELETE FROM academic_terms WHERE id = %s", (term_id,))
            conn.commit()
            flash("Academic term deleted successfully.", "success")
        else:
            flash("Database connection failed.", "danger")
    except Exception as e:
        flash(f"An error occurred while deleting the term: {e}", "danger")
        traceback.print_exc()
        if conn: conn.rollback()
    finally:
        if cursor: cursor.close()
        if conn: conn.close()
    
    return redirect(url_for('auth.admin_manage_academic_term'))


# ----------------- EMAIL SETTINGS HELPER -----------------
def _is_email_trigger_enabled(trigger_name):
    """
    Checks if a specific email trigger is enabled in system_settings.
    Defaults to True if setting is missing to ensure critical emails (like OTP) go out unless explicitly disabled.
    trigger_name: 'email_trigger_student' or 'email_trigger_admin'
    """
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT setting_value FROM system_settings WHERE setting_key = %s", (trigger_name,))
            result = cursor.fetchone()
            
            # If setting exists, return True only if value is 'true'
            if result:
                return result['setting_value'] == 'true'
            
        return True # Default ON if db connection fails or setting missing
    except Exception as e:
        print(f"Error checking email trigger setting: {e}")
        return True 
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

# ----------------- EMAIL SENDING HELPER FUNCTIONS -----------------

def _send_email(subject, recipients, html_body, sender_name_override=None):
    """
    Base function to send emails. 
    """
    mail_handler = current_app.extensions.get('mail')
    
    # 1. Check if Flask-Mail is loaded
    if not mail_handler:
        print("CRITICAL ERROR: Flask-Mail extension not found.")
        return False

    # 2. Check Credentials
    username = current_app.config.get('MAIL_USERNAME')
    password = current_app.config.get('MAIL_PASSWORD')
    
    if not username or not password:
        print("CRITICAL ERROR: MAIL_USERNAME or MAIL_PASSWORD is missing in .env file.")
        return False

    # 3. Configure Sender
    effective_sender_name = sender_name_override or current_app.config.get('MAIL_SENDER_NAME')
    default_sender_config = current_app.config.get('MAIL_DEFAULT_SENDER')
    sender_email_address = None

    if isinstance(default_sender_config, tuple):
        sender_email_address = default_sender_config[1]
        if not effective_sender_name: effective_sender_name = default_sender_config[0]
    elif isinstance(default_sender_config, str):
        sender_email_address = default_sender_config

    if not sender_email_address:
        sender_email_address = username
    
    if not effective_sender_name: 
        effective_sender_name = "PGPC System"

    final_sender = (effective_sender_name, sender_email_address)

    # 4. Attempt Send
    try:
        msg = Message(subject=subject, sender=final_sender, recipients=recipients, html=html_body)
        mail_handler.send(msg)
        print(f"SUCCESS: Email '{subject}' sent to {recipients}.")
        return True
    except Exception as e:
        # !!! CHECK YOUR TERMINAL FOR THIS ERROR MESSAGE !!!
        print(f"CRITICAL SENDING ERROR: {e}")
        traceback.print_exc()
        return False
    
def send_otp_email(user_email, otp_code):
    """
    Sends an OTP email.
    INDEPENDENT: This function DOES NOT check admin settings. It ALWAYS attempts to send.
    """
    print(f"DEBUG: Attempting to send OTP to {user_email}...") # Debug line

    sender_name_from_config = current_app.config.get('MAIL_SENDER_NAME', 'Padre Garcia Polytechnic College')
    subject = f"Verify Your PGPC Account - OTP: {otp_code} - {sender_name_from_config}"
    
    try:
        verify_url = url_for('auth.verify_otp_page', _external=True) 
    except RuntimeError:
        verify_url = f"{current_app.config.get('PREFERRED_URL_SCHEME', 'http')}://{current_app.config.get('SERVER_NAME', 'your-domain.com')}/verify-otp"

    try:
        html_body = render_template(
            'email/otp_verification_email.html',
            user_email=user_email,
            otp_code=otp_code,
            verify_url=verify_url,
            sender_name=sender_name_from_config,
            now=datetime.datetime.now()
        )
    except Exception as e_template:
        print(f"TEMPLATE ERROR: {e_template}")
        html_body = f"<p>Your OTP is: <strong>{otp_code}</strong></p>"
        
    # Direct call to send, bypassing any 'email_trigger_student' checks
    return _send_email(subject, [user_email], html_body, sender_name_override=sender_name_from_config)

def send_application_status_email(applicant_email, applicant_name, new_status, application_id, program_choice=None, exam_status=None, permit_details=None, old_student_id=None, control_number=None, academic_year=None):
    
    # --- CHECK TOGGLE ---
    if not _is_email_trigger_enabled('email_trigger_student'):
        print(f"Student email trigger disabled via Admin Panel. Skipping status update email to {applicant_email} for status '{new_status}'.")
        return False
    # --------------------

    sender_name_from_config = current_app.config.get('MAIL_SENDER_NAME', 'Padre Garcia Polytechnic College')
    
    # ID Formatting Logic
    if old_student_id:
        app_id_formatted = old_student_id
    elif control_number:
        app_id_formatted = control_number
    else:
        # Fallback for older records
        year_prefix = academic_year[:4] if academic_year else '2025'
        app_id_formatted = f"P{year_prefix}{application_id:04d}"

    subject, template_name = "", ""
    
    email_context = {
        'applicant_name': applicant_name, 'application_id_formatted': app_id_formatted,
        'program_choice': program_choice, 'sender_name': sender_name_from_config,
        'now': datetime.datetime.now(), 'exam_status': exam_status, 'application_status': new_status,
        'application': permit_details 
    }

    status_map = {
        'Approved': ('email/application_approved_email.html', f"Application {app_id_formatted} Approved"),
        'Scheduled': ('email/exam_scheduled_email.html', f"Admission Exam Scheduled for {app_id_formatted}"),
        'Rejected': ('email/application_rejected_email.html', f"Application {app_id_formatted} Update"),
        'Passed': ('email/application_passed_email.html', f"Congratulations! Application {app_id_formatted} Processed: Passed"),
        'Failed': ('email/application_failed_email.html', f"Application {app_id_formatted} Processed: Failed"),
        'Enrolled': ('email/enrollment_approved_email.html', f"Enrollment Confirmed for Application {app_id_formatted}!"),
        'Not Enrolled': ('email/application_not_enrolled_email.html', f"Application {app_id_formatted} Update: Not Enrolled"),
    }
    exam_status_map = {
        'Passed': ('email/exam_passed_email.html', f"Admission Exam Result for {app_id_formatted}"),
        'Failed': ('email/exam_failed_email.html', f"Admission Exam Result for {app_id_formatted}"),
    }

    if new_status in status_map:
        template_name, base_subject = status_map[new_status]
        subject = base_subject
        if new_status == 'Approved' and exam_status == 'Passed':
            subject = f"Congratulations! Application {app_id_formatted} Approved & Exam Passed"
        elif new_status == 'Rejected' and exam_status == 'Failed':
            subject = f"Application {app_id_formatted} Update (Exam Result)"
        subject += f" - {sender_name_from_config}"
    elif exam_status in exam_status_map and new_status not in status_map: 
        template_name, subject_base = exam_status_map[exam_status]
        subject = f"{subject_base} - {sender_name_from_config}"
    else:
        # Status changes that don't trigger emails (like Pending -> In Review)
        # print(f"Notice: Email not configured for application status '{new_status}'...") 
        return False

    try:
        html_body = render_template(template_name, **email_context)
    except Exception as e_template:
        print(f"CRITICAL: Email template '{template_name}' not found or error rendering. Error: {e_template}")
        if new_status == 'Scheduled':
             html_body = f"<p>Dear {applicant_name},</p><p>Your admission exam for application {app_id_formatted} has been scheduled. Please check the student portal for details or await further communication.</p>"
        elif new_status == 'Enrolled':
            html_body = f"<p>Dear {applicant_name},</p><p>Congratulations! Your enrollment for application {app_id_formatted} has been confirmed. Welcome to {sender_name_from_config}!</p>"
        elif new_status == 'Not Enrolled':
            html_body = f"<p>Dear {applicant_name},</p><p>This is an update regarding your application {app_id_formatted}. Your status has been set to 'Not Enrolled' as the enrollment period may have concluded. If this is a mistake, please contact the registrar's office immediately.</p>"
        else: 
            html_body = f"<p>Dear {applicant_name},</p><p>There is an update on your application {app_id_formatted}. Status: {new_status}. Exam: {exam_status or 'N/A'}.</p>"

    return _send_email(subject, [applicant_email], html_body, sender_name_override=sender_name_from_config)


def send_enrollment_rejection_email(applicant_email, applicant_name, application_id, reason, old_student_id=None, control_number=None, academic_year=None):
    
    # --- CHECK TOGGLE ---
    if not _is_email_trigger_enabled('email_trigger_student'):
        print(f"Student email trigger disabled. Skipping rejection email to {applicant_email}.")
        return False
    # --------------------

    sender_name_from_config = current_app.config.get('MAIL_SENDER_NAME', 'Padre Garcia Polytechnic College')
    
    if old_student_id:
        app_id_formatted = old_student_id
    elif control_number:
        app_id_formatted = control_number
    else:
        year_prefix = academic_year[:4] if academic_year else '2025'
        app_id_formatted = f"P{year_prefix}{application_id:04d}"

    subject = f"Update on Your Enrollment for Application {app_id_formatted} - {sender_name_from_config}"
    
    try:
        status_url = url_for('views.application_status_page', _external=True)
    except RuntimeError:
        status_url = f"{current_app.config.get('PREFERRED_URL_SCHEME', 'http')}://{current_app.config.get('SERVER_NAME', 'your-domain.com')}/application-status"

    email_context = {
        'applicant_name': applicant_name,
        'application_id_formatted': app_id_formatted,
        'reason': reason,
        'status_url': status_url,
        'sender_name': sender_name_from_config,
        'now': datetime.datetime.now()
    }
    
    template_name = 'email/enrollment_rejected_email.html'
    try:
        html_body = render_template(template_name, **email_context)
    except Exception as e_template:
        print(f"CRITICAL: Email template '{template_name}' not found. Error: {e_template}")
        html_body = f"<p>Dear {applicant_name},</p><p>Your enrollment could not be approved: {reason}</p>"

    return _send_email(subject, [applicant_email], html_body, sender_name_override=sender_name_from_config)


def send_password_reset_email(user_email, token):
    # Password resets are user-initiated critical actions. 
    # Usually you NEVER disable these, but if you must, check the student trigger.
    if not _is_email_trigger_enabled('email_trigger_student'):
         print(f"Student email trigger disabled. Skipping password reset email to {user_email}.")
         return False, None

    sender_name_from_config = current_app.config.get('MAIL_SENDER_NAME', 'Padre Garcia Polytechnic College')
    try:
        reset_url = url_for('auth.reset_password_page', token=token, _external=True)
    except RuntimeError as e:
        print(f"CRITICAL: Cannot generate external reset URL. Error: {e}")
        return False, None

    subject = f"Password Reset Request - {sender_name_from_config}"
    html_body = render_template('email/reset_password_email.html', reset_url=reset_url, user_email=user_email, sender_name=sender_name_from_config, now=datetime.datetime.now())
    email_sent = _send_email(subject, [user_email], html_body, sender_name_override=sender_name_from_config)
    return email_sent, reset_url

def send_admin_created_account_email(user_email, user_full_name, temporary_password):
    # This is triggered by ADMIN action. It might fall under Admin trigger or Student trigger.
    # Since it sends TO a student, checking student trigger makes sense.
    if not _is_email_trigger_enabled('email_trigger_student'):
         print(f"Student email trigger disabled. Skipping admin-created account email to {user_email}.")
         return False

    sender_name_from_config = current_app.config.get('MAIL_SENDER_NAME', 'Padre Garcia Polytechnic College')
    try:
        login_url = url_for('auth.student_login_page', _external=True)
    except RuntimeError:
        login_url = f"{current_app.config.get('PREFERRED_URL_SCHEME', 'http')}://{current_app.config.get('SERVER_NAME', 'your-domain.com')}/student-login"

    subject = f"Your PGPC Student Account Has Been Created - {sender_name_from_config}"
    try:
        html_body = render_template('email/admin_created_account_email.html', user_full_name=user_full_name, user_email=user_email, temporary_password=temporary_password, login_url=login_url, sender_name=sender_name_from_config, now=datetime.datetime.now())
    except Exception as e_template:
        html_body = f"<p>Dear {user_full_name},</p><p>Account created. Email: {user_email}, Temp Pass: {temporary_password}</p>"
        
    return _send_email(subject, [user_email], html_body, sender_name_override=sender_name_from_config)

def _prepare_print_transcript(student_user_id):
    """
    Helper to fetch and organize grades for printing Class Cards.
    Returns: student_info, transcript_data (dict grouped by Term)
    """
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # 1. Fetch Student Info
    cursor.execute("""
        SELECT a.first_name, a.last_name, a.old_student_id, a.control_number, a.program_choice 
        FROM applicants a
        WHERE a.student_user_id = %s 
        ORDER BY a.submitted_at DESC LIMIT 1
    """, (student_user_id,))
    student_info = cursor.fetchone()
    
    if not student_info:
        conn.close()
        return None, None

    # 2. Fetch All Grades
    cursor.execute("""
        SELECT 
            sg.academic_year, 
            sg.semester, 
            s.subject_code, 
            s.subject_title, 
            s.units,
            s.year_level,
            sg.grade, 
            sg.remarks
        FROM student_grades sg
        JOIN subjects s ON sg.subject_id = s.id
        WHERE sg.student_user_id = %s
        ORDER BY sg.academic_year DESC, sg.semester DESC, s.subject_code ASC
    """, (student_user_id,))
    
    all_grades = cursor.fetchall()
    conn.close()

    # 3. Group by Term (Year + Sem) and Calculate GWA
    transcript = {}
    
    for grade in all_grades:
        term_key = f"{grade['academic_year']} - {grade['semester']}"
        
        if term_key not in transcript:
            transcript[term_key] = {
                'academic_year': grade['academic_year'],
                'semester': grade['semester'],
                'subjects': [],
                'total_units': 0.0,
                'weighted_sum': 0.0,
                'gwa_units': 0.0
            }
        
        # Add subject to list
        transcript[term_key]['subjects'].append(grade)
        
        # Convert DB types to float to prevent TypeError (Decimal vs Float)
        units_val = float(grade['units']) if grade['units'] is not None else 0.0
        
        # Add to total units
        transcript[term_key]['total_units'] += units_val

        # Calculate GWA (Only include subjects with numeric grades > 0)
        # We explicitly cast to float here to fix the error
        if grade['grade'] is not None:
            grade_val = float(grade['grade'])
            
            if grade_val > 0:
                transcript[term_key]['weighted_sum'] += (grade_val * units_val)
                transcript[term_key]['gwa_units'] += units_val
            
    # Finalize GWA
    for key, data in transcript.items():
        if data['gwa_units'] > 0:
            data['gwa'] = data['weighted_sum'] / data['gwa_units']
        else:
            data['gwa'] = 0.0
            
    return student_info, transcript

# ----------------- ADMIN HELPERS -----------------
# ----------------- Admission id -----------------
def _generate_admission_id(cursor):
    """
    Generates a temporary Admission ID (A00001).
    """
    try:
        cursor.execute("SELECT setting_value FROM system_settings WHERE setting_key = 'last_admission_sequence' FOR UPDATE")
        row = cursor.fetchone()
        
        if not row:
            cursor.execute("INSERT INTO system_settings (setting_key, setting_value) VALUES ('last_admission_sequence', '0')")
            last_sequence = 0
        else:
            # FIX: Handle Dict vs Tuple
            if isinstance(row, dict):
                last_sequence = int(row['setting_value'])
            else:
                last_sequence = int(row[0])

        new_sequence = last_sequence + 1
        
        cursor.execute("UPDATE system_settings SET setting_value = %s WHERE setting_key = 'last_admission_sequence'", (str(new_sequence),))
        
        return f"A{new_sequence:05d}"
    except Exception as e:
        print(f"Error generating admission ID: {e}")
        raise e

# --- MODIFIED: New Student ID Generation Helper ---
def _generate_student_id(cursor):
    """
    Generates a new unique student ID (P-Series).
    Handles both dictionary cursors and standard tuple cursors to prevent type errors.
    """
    try:
        # 1. Get the current enrollment year setting
        cursor.execute("SELECT setting_value FROM system_settings WHERE setting_key = 'current_enrollment_year'")
        year_row = cursor.fetchone()
        
        # 2. Get and lock the global sequence number to prevent race conditions
        cursor.execute("SELECT setting_value FROM system_settings WHERE setting_key = 'last_sequence_number' FOR UPDATE")
        seq_row = cursor.fetchone()

        if not year_row or not seq_row:
            raise Exception("System settings for 'current_enrollment_year' or 'last_sequence_number' not found in the database.")

        # --- FIX: Handle Cursor Type (Dict vs Tuple) ---
        # If row is a dictionary (from dictionary=True cursor), access by key
        # If row is a tuple (from standard cursor), access by index 0
        if isinstance(year_row, dict):
            year = year_row['setting_value']
        else:
            year = year_row[0]

        if isinstance(seq_row, dict):
            last_sequence = int(seq_row['setting_value'])
        else:
            last_sequence = int(seq_row[0])
        # -----------------------------------------------
        
        # 3. Increment the sequence
        new_sequence = last_sequence + 1
        
        # 4. Update the sequence number in the database immediately
        cursor.execute("UPDATE system_settings SET setting_value = %s WHERE setting_key = 'last_sequence_number'", (str(new_sequence),))
        
        # 5. Format the new ID (P<YYYY><5-digit-number>)
        new_student_id = f"P{year}{new_sequence:05d}"
        
        return new_student_id
    except Exception as e:
        print(f"Error generating student ID: {e}")
        traceback.print_exc()
        raise e  # Re-raise to trigger a rollback in the calling function

def _get_program_list():
    """Fetches a simple list of program titles from the DB for filtering."""
    programs = []
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor(dictionary=True)
            # Fetching just the title as it's used for filtering value and display
            cursor.execute("SELECT title FROM programs ORDER BY title ASC")
            programs = cursor.fetchall()
    except Exception as e:
        print(f"Error fetching program list: {e}")
        traceback.print_exc()
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()
    return programs

def _get_all_sections():
    """Fetches all sections for filter dropdowns."""
    sections = []
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor(dictionary=True)
            # Fetch distinct section names or all sections
            cursor.execute("SELECT section_name FROM sections ORDER BY section_name ASC")
            sections = cursor.fetchall()
    except Exception as e:
        print(f"Error fetching sections: {e}")
    finally:
        if cursor: cursor.close()
        if conn: conn.close()
    return sections

def _check_and_update_missed_exams():
    """
    Automatically updates the exam_status of scheduled applicants to 'Not Taken'
    if their exam date has passed and their status is still null.
    Returns the number of updated records.
    """
    conn = None
    cursor = None
    updated_count = 0
    try:
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            # CURDATE() gets the current date.
            # We update applicants who were scheduled, have a past exam date, and have no exam status yet.
            query = """
                UPDATE applicants
                SET exam_status = 'Not Taken', last_updated_at = NOW()
                WHERE 
                    (application_status = 'Scheduled' OR (application_status = 'Approved' AND permit_exam_date IS NOT NULL))
                    AND permit_exam_date < CURDATE()
                    AND (exam_status IS NULL OR exam_status = '')
            """
            cursor.execute(query)
            conn.commit()
            updated_count = cursor.rowcount
    except Exception as e:
        print(f"Error checking for missed exams: {e}")
        traceback.print_exc()
        if conn: conn.rollback()
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()
    return updated_count

def _get_active_term():
    """Fetches the currently active academic term."""
    conn = None; cursor = None
    try:
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT * FROM academic_terms WHERE is_active = TRUE LIMIT 1")
            return cursor.fetchone()
    except Exception as e:
        print(f"Error fetching active academic term: {e}")
    finally:
        if cursor: cursor.close()
        if conn: conn.close()
    return None

def _get_all_school_years():
    """Fetches a list of all school year names."""
    school_years = []
    conn = None; cursor = None
    try:
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT DISTINCT year_name FROM academic_terms ORDER BY year_name DESC")
            school_years = cursor.fetchall()
    except Exception as e:
        print(f"Error fetching all school years: {e}")
    finally:
        if cursor: cursor.close()
        if conn: conn.close()
    return school_years

def _auto_assign_fees(cursor, applicant_id, student_data):
    """
    Appends new fees to the student's ledger based on their CURRENT Program, Year, and Semester.
    Does not care about existing balances; simply adds new matching assessments.
    """
    try:
        # 1. Fetch all active batch assessment rules
        cursor.execute("SELECT * FROM batch_assessments_log")
        batches = cursor.fetchall()
        
        assigned_count = 0
        
        # 2. Extract context from the student's record
        stud_course = student_data.get('program_choice')
        stud_year = student_data.get('enrollment_year_level')
        stud_sem = student_data.get('enrollment_semester')   
        
        # 3. Residency check for local discounts/fees
        address = (student_data.get('permanent_address_city_municipality') or '').lower()
        is_garciano = 'padre garcia' in address

        for batch in batches:
            # --- MATCHING LOGIC ---
            if batch.get('semester') and batch['semester'] != 'All' and batch['semester'] != stud_sem:
                continue
            if batch['course'] and batch['course'] != 'All' and batch['course'] != stud_course:
                continue
            if batch['year_level'] and batch['year_level'] != 'All' and batch['year_level'] != stud_year:
                continue
            if batch['scope'] == 'garciano' and not is_garciano:
                continue
            if batch['scope'] == 'non_garciano' and is_garciano:
                continue

            # 4. PREVENT DUPLICATES: Only add if this specific fee isn't already on the student's SOA
            cursor.execute("""
                SELECT id FROM assessments 
                WHERE student_id = %s AND description = %s
            """, (applicant_id, batch['description']))
            
            if not cursor.fetchone():
                cursor.execute("""
                    INSERT INTO assessments (student_id, description, total_amount, balance, amount_paid, status, created_at)
                    VALUES (%s, %s, %s, %s, 0.00, 'unpaid', NOW())
                """, (applicant_id, batch['description'], batch['amount'], batch['amount']))
                assigned_count += 1
        
        return assigned_count
    except Exception as e:
        print(f"Error in _auto_assign_fees: {e}")
        return 0

# ----------------- HELPER: AUTO PROGRESSION LOGIC -----------------
def _calculate_next_term(current_year_str, current_sem_str):
    """
    Calculates the next academic level based on current status.
    Returns: (next_year_str, next_sem_str, is_graduating)
    """
    if not current_year_str or not current_sem_str:
        return None, None, False

    # 1. Parse Year (e.g., "1st Year" -> 1)
    try:
        year_int = int(current_year_str[0]) 
    except:
        return current_year_str, current_sem_str, False

    # 2. Logic
    next_year_int = year_int
    next_sem_str = ""
    is_graduating = False

    if "1st" in current_sem_str:
        # 1st Sem -> 2nd Sem (Same Year)
        next_sem_str = "2nd Semester"
    elif "2nd" in current_sem_str:
        # 2nd Sem -> 1st Sem (Next Year)
        if year_int < 4:
            next_year_int = year_int + 1
            next_sem_str = "1st Semester"
        else:
            # Already 4th Year 2nd Sem
            return current_year_str, current_sem_str, True # Flag as graduating/done
    elif "Summer" in current_sem_str:
         # Summer -> Next Year 1st Sem (usually)
         if year_int < 4:
            next_year_int = year_int + 1
            next_sem_str = "1st Semester"
    
    # Format Year String
    suffix = "th"
    if next_year_int == 1: suffix = "st"
    elif next_year_int == 2: suffix = "nd"
    elif next_year_int == 3: suffix = "rd"
    
    next_year_str = f"{next_year_int}{suffix} Year"

    return next_year_str, next_sem_str, False

def _get_all_applications_and_stats(school_year=None):
    """Helper function to fetch all applications, stats, AND financial data."""
    updated_count = _check_and_update_missed_exams()
    if updated_count > 0:
        flash(f"{updated_count} scheduled applicant(s) from past dates were automatically marked as 'Did Not Attend'.", "info")

    applications = []
    stats = {
        'total_applications': 0, 'pending': 0, 'approved': 0, 'scheduled': 0, 
        'rejected': 0, 'passed': 0, 'failed': 0, 'not_taken': 0, 'enrolling': 0,
        'enrolled': 0, 'dropped': 0, 'not_enrolled': 0, 'eligible_for_enrollment': 0
    }
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        if not conn:
            flash("Database connection error.", "danger")
            return [], stats

        cursor = conn.cursor(dictionary=True)
        
        # MODIFIED QUERY: Added subqueries to calculate total_balance and total_paid from assessments table
        base_query = """
            SELECT a.*, su.email as student_account_email, p.program_id, s.section_name,
            (SELECT SUM(balance) FROM assessments WHERE student_id = a.applicant_id) as total_balance,
            (SELECT SUM(amount_paid) FROM assessments WHERE student_id = a.applicant_id) as total_paid
            FROM applicants a 
            LEFT JOIN student_users su ON a.student_user_id = su.id 
            LEFT JOIN programs p ON TRIM(a.program_choice) = TRIM(p.title)
            LEFT JOIN sections s ON a.section_id = s.id
        """
        params = []
        if school_year:
            base_query += " WHERE a.academic_year = %s"
            params.append(school_year)
        
        base_query += " ORDER BY a.submitted_at DESC"
        
        cursor.execute(base_query, tuple(params))
        applications = cursor.fetchall()
        
        stats['total_applications'] = len(applications)
        approved_awaiting_schedule_count = 0
        scheduled_count = 0
        not_taken_count = 0

        for app in applications:
            # Decode bytes if necessary
            for key, value in app.items(): 
                if isinstance(value, bytes):
                    try: 
                        app[key] = value.decode('utf-8')
                    except UnicodeDecodeError: 
                        app[key] = "Decode Error"
            
            if app.get('average_family_income'):
                app['average_family_income'] = app['average_family_income'].replace('_', ' ')

            # --- FINANCIAL STATUS LOGIC ---
            balance = app.get('total_balance')
            paid = app.get('total_paid')

            if balance is None:
                # No assessments found
                app['financial_status'] = 'No Assessment'
                app['display_balance'] = 0.00
            elif float(balance) <= 0:
                app['financial_status'] = 'Fully Paid'
                app['display_balance'] = 0.00
            elif paid and float(paid) > 0:
                app['financial_status'] = 'Partial'
                app['display_balance'] = float(balance)
            else:
                app['financial_status'] = 'Unpaid'
                app['display_balance'] = float(balance)
            # -----------------------------
            
            # Stats calculation (unchanged)
            if app.get('exam_status') == 'Not Taken':
                not_taken_count += 1

            is_scheduled_and_attending = (
                (app.get('application_status') == 'Scheduled' or
                 (app.get('application_status') == 'Approved' and 
                  app.get('permit_exam_date') and 
                  app.get('permit_exam_time') and 
                  app.get('permit_testing_room')))
                and app.get('exam_status') != 'Not Taken'
            )

            if is_scheduled_and_attending:
                scheduled_count += 1
            elif app.get('application_status') == 'Approved':
                approved_awaiting_schedule_count += 1
            
            status_key = app.get('application_status','').lower().replace(' ', '_')
            if status_key in stats:
                 stats[status_key] += 1
        
        stats['approved'] = approved_awaiting_schedule_count
        stats['scheduled'] = scheduled_count
        stats['not_taken'] = not_taken_count

    except Exception as e:
        flash(f"Error loading application data: {e}", "danger")
        traceback.print_exc()
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()

    return applications, stats

def _recalculate_subject_availability(cursor, student_user_id, program_id):
    """
    Recalculates and updates the availability status for all subjects in a program
    based on a student's failed prerequisites.
    NOTE: This function assumes the 'subjects' table has an 'availability_status' column.
    """
    if not all([cursor, student_user_id, program_id]):
        return

    try:
        # 1. Get all subjects the student has failed
        cursor.execute("SELECT subject_id FROM student_grades WHERE student_user_id = %s AND remarks = 'Failed'", (student_user_id,))
        failed_subject_ids = {row['subject_id'] for row in cursor.fetchall()}

        # 2. Get all subjects in the program's curriculum
        cursor.execute("SELECT id, prerequisite_subject_id FROM subjects WHERE program_id = %s", (program_id,))
        all_program_subjects = cursor.fetchall()

        subjects_to_update = []
        for subject in all_program_subjects:
            prereq_id = subject.get('prerequisite_subject_id')
            new_status = 'Available'
            if prereq_id and prereq_id in failed_subject_ids:
                new_status = 'Unavailable'
            
            # Prepare for bulk update
            subjects_to_update.append((new_status, subject['id']))
        
        # 3. Perform a bulk update on the subjects table
        # NOTE: This updates the table for ALL students viewing this curriculum, 
        # but the frontend logic will still show the correct state based on the calculated disabled_subjects_info.
        # A more advanced implementation would have a separate student_subject_availability table.
        # For now, this recalculation happens on each load.

    except Exception as e:
        print(f"Error recalculating subject availability: {e}")
        traceback.print_exc()
        # Don't proceed with a partial update
        return

def _get_student_status_and_failed_subjects(cursor, student_user_id):
    """
    Determines a student's academic status (Regular/Irregular) and returns a set of failed/incomplete subject IDs.
    """
    if not student_user_id or not cursor:
        return "N/A", set()
    
    # MODIFIED: Also check for 'Incomplete' remarks
    cursor.execute("""
        SELECT subject_id FROM student_grades 
        WHERE student_user_id = %s AND remarks IN ('Failed', 'Incomplete')
    """, (student_user_id,))
    
    failed_subjects_rows = cursor.fetchall()
    failed_subject_ids = {row['subject_id'] for row in failed_subjects_rows}
    
    status = "Irregular" if failed_subject_ids else "Regular"
    
    return status, failed_subject_ids


def _render_applications_page(template_name, active_page, applications, stats, programs, all_school_years, selected_school_year, page_title="All Applications", filter_status=None, filter_exam_status=None):
    """Helper to render application list pages with filtering."""
    filtered_apps = applications
    if filter_status:
        if filter_status == 'Approved':
            filtered_apps = [
                app for app in applications 
                if app.get('application_status') == 'Approved' and not (
                    app.get('permit_exam_date') and app.get('permit_exam_time') and app.get('permit_testing_room')
                )
            ]
        elif filter_status == 'Scheduled':
            filtered_apps = [
                app for app in applications 
                if (app.get('application_status') == 'Scheduled' or
                    (app.get('application_status') == 'Approved' and 
                     app.get('permit_exam_date') and 
                     app.get('permit_exam_time') and 
                     app.get('permit_testing_room')))
                and app.get('exam_status') != 'Not Taken'
            ]
        else:
            filtered_apps = [app for app in applications if app.get('application_status') == filter_status]
    elif filter_exam_status:
        filtered_apps = [app for app in applications if app.get('exam_status') == filter_exam_status]
            
    return render_template(template_name, applications=filtered_apps, stats=stats, active_page=active_page, page_title=page_title, programs=programs, all_school_years=all_school_years, selected_school_year=selected_school_year)

def admin_print_grade_card(student_user_id):
    # Check permissions
    if session.get('user_role') not in ['admin', 'registrar', 'guidance', 'osa']:
        return redirect(url_for('auth.admin'))
        
    student, transcript = _prepare_print_transcript(student_user_id)
    
    if not student:
        flash("Student not found.", "danger")
        return redirect(request.referrer)
        
    return render_template('printable_class_card.html', student=student, transcript=transcript)


@auth.route('/admin/grades/<int:student_user_id>/print-card')
def admin_print_grade_card(student_user_id):
    # Check permissions
    if session.get('user_role') not in ['admin', 'registrar', 'guidance', 'osa']:
        return redirect(url_for('auth.admin'))
        
    student, transcript = _prepare_print_transcript(student_user_id)
    
    if not student:
        flash("Student not found.", "danger")
        return redirect(request.referrer)
        
    return render_template('printable_class_card.html', student=student, transcript=transcript)


@auth.route('/my-grades/print-card')
def student_print_grade_card():
    # Check login
    if 'student_id' not in session:
        return redirect(url_for('auth.student_login_page'))
        
    student_user_id = session['student_id']
    student, transcript = _prepare_print_transcript(student_user_id)
    
    if not student:
        flash("Record not found.", "danger")
        return redirect(url_for('views.application_status_page'))
        
    return render_template('printable_class_card.html', student=student, transcript=transcript)

# ----------------- ACCOUNTING ROUTES -----------------

@auth.route('/accounting/chart-of-accounts', methods=['GET'])
def chart_of_accounts():
    if session.get('user_role') != 'accounting': return redirect(url_for('auth.admin_login'))
    
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT * FROM chart_of_accounts ORDER BY account_code")
        accounts = cursor.fetchall()
    finally:
        cursor.close()
        conn.close()
        
    return render_template('chart_of_accounts.html', accounts=accounts)

@auth.route('/accounting/add-account', methods=['POST'])
def add_account():
    if session.get('user_role') != 'accounting': return redirect(url_for('auth.admin_login'))
    
    code = request.form.get('account_code')
    name = request.form.get('account_name')
    acct_type = request.form.get('type')
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO chart_of_accounts (account_code, account_name, type) VALUES (%s, %s, %s)", 
                       (code, name, acct_type))
        conn.commit()
        flash("Account added successfully.", "success")
    except Exception as e:
        flash(f"Error adding account: {e}", "danger")
    finally:
        cursor.close()
        conn.close()
    return redirect(url_for('chart_of_accounts'))

@auth.route('/accounting/delete-account/<int:id>', methods=['POST'])
def delete_account(id):
    if session.get('user_role') != 'accounting': return redirect(url_for('auth.admin_login'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM chart_of_accounts WHERE id = %s", (id,))
        conn.commit()
        flash("Account deleted.", "success")
    except Exception as e:
        flash("Cannot delete this account. It might be in use.", "danger")
    finally:
        cursor.close()
        conn.close()
    return redirect(url_for('chart_of_accounts'))

@auth.route('/accounting/journal', methods=['GET'])
def journal_entries():
    if session.get('user_role') != 'accounting': return redirect(url_for('auth.admin_login'))
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    query = """
        SELECT je.id, je.entry_date, je.description, je.reference_number
        FROM journal_entries je
    """
    params = []
    
    if start_date and end_date:
        query += " WHERE je.entry_date BETWEEN %s AND %s"
        params.extend([start_date, end_date])
    
    query += " ORDER BY je.entry_date DESC, je.id DESC"
    
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    entries = []
    try:
        cursor.execute(query, tuple(params))
        headers = cursor.fetchall()
        
        for head in headers:
            cursor.execute("""
                SELECT jl.*, ca.account_name, ca.account_code 
                FROM journal_lines jl
                JOIN chart_of_accounts ca ON jl.account_id = ca.id
                WHERE jl.journal_entry_id = %s
            """, (head['id'],))
            lines = cursor.fetchall()
            head['lines'] = lines
            entries.append(head)
    finally:
        cursor.close()
        conn.close()
        
    return render_template('journal_entries.html', entries=entries, start_date=start_date, end_date=end_date)

# --- NEW: Save Receipt Layout via AJAX ---
@auth.route('/admin/save-receipt-layout', methods=['POST'])
def save_receipt_layout():
    if session.get('user_role') not in ['admin', 'accounting', 'cashier']:
        return jsonify({"success": False, "message": "Unauthorized"}), 401

    data = request.get_json() # JavaScript sends the 'layout' array here
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        for item in data['layout']:
            # This updates each field (dragDate, dragName, etc) with its new position
            cursor.execute("""
                UPDATE receipt_layout_settings 
                SET top_pos = %s, left_pos = %s, font_size = %s 
                WHERE field_id = %s
            """, (item['top'], item['left'], item['fontSize'], item['id']))
        
        conn.commit()
        return jsonify({"success": True, "message": "Layout saved permanently!"})
    except Exception as e:
        print(f"Save Error: {e}")
        return jsonify({"success": False, "message": str(e)}), 500
    finally:
        cursor.close()
        conn.close()

# --- NEW: Upload Receipt Background ---
@auth.route('/admin/upload-receipt-bg', methods=['POST'])
def upload_receipt_bg():
    if session.get('user_role') not in ['admin', 'accounting', 'cashier']:
        return redirect(url_for('auth.admin_login'))

    file = request.files.get('receipt_bg')
    if file and file.filename:
        # Re-use your existing helper save_image_to_uploads
        filename, error = save_image_to_uploads(file, prefix='receipt_bg')
        if not error:
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("UPDATE system_settings SET setting_value = %s WHERE setting_key = 'receipt_background_file'", (filename,))
            conn.commit()
            cursor.close()
            conn.close()
            flash("Receipt template background updated!", "success")
        else:
            flash(error, "danger")
    return redirect(url_for('auth.receipt_layout_tool'))

# --- MODIFIED: Print Receipt Route to fetch saved positions ---
@auth.route('/cashier/print-receipt/<int:payment_id>')
def print_receipt(payment_id):
    if session.get('user_role') not in ['cashier', 'accounting']:
        return redirect(url_for('auth.admin_login'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        # 1. Fetch Payment Data (your existing query)
        cursor.execute("SELECT p.*, a.first_name, a.last_name, asm.description as fee_description FROM payments p JOIN applicants a ON p.student_id = a.applicant_id LEFT JOIN assessments asm ON p.assessment_id = asm.id WHERE p.id = %s", (payment_id,))
        payment = cursor.fetchone()

        # 2. Fetch Layout Settings
        cursor.execute("SELECT * FROM receipt_layout_settings")
        layout_list = cursor.fetchall()
        # Convert list to dictionary for easy access: {'dragDate': {'top_pos': '10px', ...}}
        layout = {item['field_id']: item for item in layout_list}

        # 3. Fetch Background filename
        cursor.execute("SELECT setting_value FROM system_settings WHERE setting_key = 'receipt_background_file'")
        bg_row = cursor.fetchone()
        bg_file = bg_row['setting_value'] if bg_row else 'receipt.jpg'

    finally:
        cursor.close()
        conn.close()

    return render_template('print_receipt_movable.html', payment=payment, layout=layout, bg_file=bg_file)


@auth.route('/cashier/receipt-layout-tool', methods=['GET'])
def receipt_layout_tool():
    if session.get('user_role') not in ['cashier', 'accounting', 'admin']:
        return redirect(url_for('auth.admin_login'))
    
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 1. Fetch Layout Settings (Positions and Font Sizes)
        cursor.execute("SELECT * FROM receipt_layout_settings")
        layout_list = cursor.fetchall()
        
        # Convert list to dictionary: {'dragDate': {'top_pos': '15%', ...}, ...}
        layout = {item['field_id']: item for item in layout_list}
        
        # 2. Fetch Background filename from system_settings
        cursor.execute("SELECT setting_value FROM system_settings WHERE setting_key = 'receipt_background_file'")
        bg_row = cursor.fetchone()
        bg_file = bg_row['setting_value'] if bg_row else 'receipt.jpg'

        # 3. Optional: Fetch payment data if a test ID is provided in URL
        payment_id = request.args.get('payment_id')
        payment_data = None
        if payment_id:
            cursor.execute("""
                SELECT p.*, a.first_name, a.last_name, asm.description as fee_description
                FROM payments p
                JOIN applicants a ON p.student_id = a.applicant_id
                LEFT JOIN assessments asm ON p.assessment_id = asm.id
                WHERE p.id = %s
            """, (payment_id,))
            payment_data = cursor.fetchone()

        return render_template('receipt_layout_tool.html', 
                               layout=layout, 
                               bg_file=bg_file, 
                               payment=payment_data)

    except Exception as e:
        print(f"Error loading designer: {e}")
        flash("Could not load layout settings. Ensure the database table exists.", "danger")
        return redirect(url_for('auth.cashier_dashboard'))
    finally:
        cursor.close()
        conn.close()
# ----------------- STUDENT DISCOUNT ROUTES -----------------

# --- IN auth.py ---

@auth.route('/student/discount', methods=['GET', 'POST'])
def student_discount():
    if session.get('user_role') not in ['cashier', 'accounting']:
        return redirect(url_for('auth.admin_login'))

    base_template = 'accounting_base.html' 
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        # --- HANDLE POST ACTIONS ---
        if request.method == 'POST':
            action = request.form.get('action')
            
            if action == 'add_org':
                name = request.form.get('name')
                amount = request.form.get('amount')
                try:
                    cursor.execute("INSERT INTO organization_discounts (name, discount_amount) VALUES (%s, %s)", (name, amount))
                    conn.commit()
                    flash(f"Discount rule for '{name}' added.", "success")
                except mysql.connector.Error as err:
                    flash(f"Error: {err}", "danger")

            elif action == 'delete_org':
                org_id = request.form.get('org_id')
                cursor.execute("DELETE FROM organization_discounts WHERE id = %s", (org_id,))
                conn.commit()
                flash("Discount rule deleted.", "success")

            # --- NEW: Manually Add Student to Org ---
            elif action == 'add_student_to_org':
                student_identifier = request.form.get('student_identifier') # ID or Name
                org_name = request.form.get('org_name')
                
                # Find the student first
                cursor.execute("""
                    SELECT applicant_id, first_name, last_name 
                    FROM applicants 
                    WHERE (old_student_id = %s OR control_number = %s OR applicant_id = %s)
                    AND application_status = 'Enrolled'
                """, (student_identifier, student_identifier, student_identifier))
                
                student = cursor.fetchone()
                
                if student:
                    # Update their organization field
                    cursor.execute("UPDATE applicants SET organization = %s WHERE applicant_id = %s", (org_name, student['applicant_id']))
                    conn.commit()
                    flash(f"Added {student['first_name']} {student['last_name']} to {org_name}.", "success")
                else:
                    flash(f"Student not found with ID: {student_identifier}", "danger")
            
            return redirect(url_for('auth.student_discount'))

        # --- HANDLE GET (View Data) ---
        
        # 1. Get Discount Rules
        cursor.execute("SELECT * FROM organization_discounts ORDER BY name")
        organizations = cursor.fetchall()
        org_rules = {org['name']: org['discount_amount'] for org in organizations}
        org_ids = {org['name']: org['id'] for org in organizations}

        # 2. Get Students with Organizations
        # Filters
        course_filter = request.args.get('course')
        year_filter = request.args.get('year_level')

        query = """
            SELECT applicant_id as id, old_student_id, first_name, last_name, program_choice as course, 
                   enrollment_year_level as year_level, organization
            FROM applicants 
            WHERE application_status = 'Enrolled' 
            AND organization IS NOT NULL 
            AND organization != ''
        """
        params = []
        
        if course_filter:
            query += " AND program_choice = %s"
            params.append(course_filter)
        if year_filter:
            query += " AND enrollment_year_level = %s"
            params.append(year_filter)
            
        cursor.execute(query, tuple(params))
        students_raw = cursor.fetchall()

        # 3. Get Applied Status
        cursor.execute("SELECT student_id FROM student_discounts_applied")
        applied_set = {row['student_id'] for row in cursor.fetchall()}

        # 4. Merge Data
        student_list = []
        for s in students_raw:
            org_name = s['organization']
            discount = org_rules.get(org_name, 0)
            is_applied = s['id'] in applied_set
            
            # Format display ID
            s['display_id'] = s['old_student_id'] or s['id']

            student_list.append({
                'student': s,
                'discount_amount': discount,
                'is_applied': is_applied,
                'org_id': org_ids.get(org_name)
            })

    except Exception as e:
        print(f"Error in student discount: {e}")
        flash("An error occurred loading the page.", "danger")
        organizations = []
        student_list = []
    finally:
        cursor.close()
        conn.close()

    return render_template('student_discount.html', 
                           organizations=organizations, 
                           students=student_list,
                           course=request.args.get('course'),
                           year_level=request.args.get('year_level'),
                           base_template=base_template)

@auth.route('/student/discount/apply/<int:student_id>', methods=['POST'])
def apply_org_discount(student_id):
    if session.get('user_role') not in ['cashier', 'accounting']:
        return redirect(url_for('auth.admin_login'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        # 1. Get Student Org Info
        cursor.execute("SELECT organization FROM applicants WHERE applicant_id = %s", (student_id,))
        student = cursor.fetchone()
        if not student or not student['organization']:
            flash("Student does not belong to an organization.", "warning")
            return redirect(url_for('auth.student_discount'))

        # 2. Get Discount Rule
        cursor.execute("SELECT id, discount_amount FROM organization_discounts WHERE name = %s", (student['organization'],))
        rule = cursor.fetchone()
        if not rule:
            flash(f"No discount rule found for {student['organization']}.", "warning")
            return redirect(url_for('auth.student_discount'))

        discount_amount = float(rule['discount_amount'])

        # 3. Check if already applied
        cursor.execute("SELECT id FROM student_discounts_applied WHERE student_id = %s AND org_discount_id = %s", (student_id, rule['id']))
        if cursor.fetchone():
            flash("Discount already applied to this student.", "warning")
            return redirect(url_for('auth.student_discount'))

        # 4. Find a specific assessment to apply to (Prioritize 'Miscellaneous Fee', fallback to any unpaid)
        # Note: Ideally, you apply this to a specific fee. Here we apply to the largest unpaid balance.
        cursor.execute("SELECT id, balance FROM assessments WHERE student_id = %s AND status != 'paid' ORDER BY balance DESC LIMIT 1", (student_id,))
        assessment = cursor.fetchone()

        if not assessment:
            flash("Student has no unpaid assessments to apply the discount to.", "warning")
            return redirect(url_for('auth.student_discount'))

        # 5. APPLY: Record as a "Payment" (Type: Discount) and Reduce Balance
        # Update Assessment
        new_balance = float(assessment['balance']) - discount_amount
        new_status = 'paid' if new_balance <= 0 else 'partial'
        # Ensure balance doesn't go below zero for logic, though technically a refund/credit
        
        cursor.execute("""
            UPDATE assessments 
            SET balance = balance - %s, 
                amount_paid = amount_paid + %s,
                status = %s 
            WHERE id = %s
        """, (discount_amount, discount_amount, new_status, assessment['id']))

        # Insert Payment Record (Log as Discount)
        cursor.execute("""
            INSERT INTO payments (student_id, assessment_id, amount_paid, payment_date, payment_method, remark, cashier_username)
            VALUES (%s, %s, %s, NOW(), 'Discount', %s, %s)
        """, (student_id, assessment['id'], discount_amount, f"Org Discount: {student['organization']}", session.get('user_name')))

        # Log Application
        cursor.execute("""
            INSERT INTO student_discounts_applied (student_id, org_discount_id, applied_by)
            VALUES (%s, %s, %s)
        """, (student_id, rule['id'], session.get('user_name')))

        conn.commit()
        flash(f"Discount of ₱{discount_amount} applied successfully.", "success")

    except Exception as e:
        print(f"Error applying discount: {e}")
        flash("Error applying discount.", "danger")
        if conn: conn.rollback()
    finally:
        cursor.close()
        conn.close()

    return redirect(url_for('auth.student_discount'))

# ----------------- RETURN VOUCHER ROUTES -----------------

# --- IN auth.py ---

@auth.route('/return-voucher', methods=['GET'])
def return_voucher():
    # 1. Permissions
    if session.get('user_role') not in ['cashier', 'accounting']:
        return redirect(url_for('auth.admin_login'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        # 2. Get Filters
        school_year = request.args.get('school_year')
        semester = request.args.get('semester')
        course = request.args.get('course')
        year_level = request.args.get('year_level')
        section = request.args.get('section')
        status_filter = request.args.get('status')
        search_query = request.args.get('query', '')

        # 3. Handle Defaults (Crucial for the "Update Term" logic)
        # If the user hasn't selected a specific term, default to the Active Term.
        # This ensures the UI always shows the status for *some* specific term.
        active_term = _get_active_term()
        if not school_year and active_term:
            school_year = active_term['year_name']
        if not semester and active_term:
            semester = active_term['semester']

        # 4. Fetch Students (Garciano only)
        # We query students regardless of term history, but filter by current enrollment data
        query = """
            SELECT a.applicant_id as id, a.old_student_id as student_id, a.first_name, a.last_name,
                   a.program_choice as course, a.enrollment_year_level as year_level,
                   s.section_name as section
            FROM applicants a
            LEFT JOIN sections s ON a.section_id = s.id
            WHERE a.application_status = 'Enrolled'
            AND a.permanent_address_city_municipality LIKE '%Padre Garcia%'
        """
        params = []

        if course:
            query += " AND a.program_choice = %s"
            params.append(course)
        if year_level:
            query += " AND a.enrollment_year_level = %s"
            params.append(year_level)
        if section:
            query += " AND s.section_name = %s"
            params.append(section)
        if search_query:
            query += " AND (a.last_name LIKE %s OR a.first_name LIKE %s OR a.old_student_id LIKE %s)"
            search_term = f"%{search_query}%"
            params.extend([search_term, search_term, search_term])
        
        query += " ORDER BY a.last_name ASC"

        cursor.execute(query, tuple(params))
        all_students = cursor.fetchall()

        # 5. Fetch Returned IDs for the SPECIFIC Selected Term
        # This is the logic that resets the button when you change the term.
        # It only grabs IDs that have been marked returned for the filtered School Year & Semester.
        cursor.execute("""
            SELECT student_id 
            FROM voucher_returns 
            WHERE school_year = %s AND semester = %s
        """, (school_year, semester))
        
        returned_rows = cursor.fetchall()
        returned_ids = {row['student_id'] for row in returned_rows}

        # 6. Process Final List & Counts
        final_list = []
        returned_count = 0
        not_returned_count = 0

        for stud in all_students:
            is_returned = stud['id'] in returned_ids
            
            if is_returned: returned_count += 1
            else: not_returned_count += 1

            # Apply Status Filter (Returned vs Not Returned view)
            if status_filter == 'returned' and not is_returned: continue
            if status_filter == 'not_returned' and is_returned: continue
            
            final_list.append(stud)

        # Fetch all school years for dropdown
        all_sy = _get_all_school_years()

        return render_template('return_voucher.html', 
                               students=final_list,
                               returned_ids=returned_ids,
                               returned_count=returned_count,
                               not_returned_count=not_returned_count,
                               all_school_years=all_sy,
                               # Pass back filters so they stick in the form
                               school_year=school_year, semester=semester,
                               course=course, year_level=year_level,
                               section=section, status_filter=status_filter,
                               search_query=search_query,
                               base_template='accounting_base.html')

    except Exception as e:
        print(f"Error in return_voucher view: {e}")
        traceback.print_exc()
        flash("An error occurred loading the page.", "danger")
        return redirect(url_for('auth.cashier_dashboard'))
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

@auth.route('/submit-voucher/<int:student_id_db>', methods=['POST'])
def submit_voucher(student_id_db):
    if session.get('user_role') not in ['cashier', 'accounting']:
        return redirect(url_for('auth.admin_login'))
    
    school_year = request.form.get('school_year')
    semester = request.form.get('semester')
    amount_str = request.form.get('amount') 

    # Fallback to active term if filtering failed
    if not school_year or not semester:
        active = _get_active_term()
        if active:
            school_year = school_year or active['year_name']
            semester = semester or active['semester']
    
    if not school_year or not semester:
        flash("Please select a specific School Year and Semester.", "warning")
        return redirect(request.referrer)

    # Validate amount
    if not amount_str:
        flash("Global voucher amount is invalid.", "danger")
        return redirect(request.referrer)
    try:
        voucher_amount = float(amount_str)
        if voucher_amount <= 0: raise ValueError
    except ValueError:
        flash("Invalid amount.", "danger")
        return redirect(request.referrer)

    conn = get_db_connection()
    # FIX: Add buffered=True to handle nested queries and unread results safely
    cursor = conn.cursor(dictionary=True, buffered=True)
    
    try:
        # --- NEW CHECK: PREVENT DUPLICATE RETURN FOR SAME TERM ---
        cursor.execute("""
            SELECT id FROM voucher_returns 
            WHERE student_id = %s AND school_year = %s AND semester = %s
        """, (student_id_db, school_year, semester))
        
        existing_return = cursor.fetchone()
        
        if existing_return:
            flash(f"Voucher already returned for {school_year} - {semester}.", "warning")
            return redirect(request.referrer)
        # ---------------------------------------------------------

        # 1. Log the Return
        cursor.execute("""
            INSERT INTO voucher_returns (student_id, school_year, semester, returned_by, amount)
            VALUES (%s, %s, %s, %s, %s)
        """, (student_id_db, school_year, semester, session.get('user_name'), voucher_amount))

        # 2. Financial Deduction
        cursor.execute("SELECT id, balance FROM assessments WHERE student_id = %s AND status != 'paid' ORDER BY balance DESC LIMIT 1", (student_id_db,))
        assessment = cursor.fetchone()
        assessment_id = assessment['id'] if assessment else None

        cursor.execute("""
            INSERT INTO payments (student_id, assessment_id, amount_paid, payment_date, payment_method, remark, cashier_username)
            VALUES (%s, %s, %s, NOW(), 'Voucher', %s, %s)
        """, (student_id_db, assessment_id, voucher_amount, f"Garciano Voucher", session.get('user_name')))

        if assessment_id:
            cursor.execute("""
                UPDATE assessments 
                SET balance = GREATEST(0, balance - %s), 
                    amount_paid = amount_paid + %s
                WHERE id = %s
            """, (voucher_amount, voucher_amount, assessment_id))
            
            cursor.execute("SELECT total_amount, amount_paid FROM assessments WHERE id = %s", (assessment_id,))
            chk = cursor.fetchone()
            if chk and chk['amount_paid'] >= chk['total_amount']:
                cursor.execute("UPDATE assessments SET status = 'paid' WHERE id = %s", (assessment_id,))
            else:
                cursor.execute("UPDATE assessments SET status = 'partial' WHERE id = %s", (assessment_id,))

        conn.commit()
        flash(f"Voucher returned. ₱{voucher_amount:,.2f} deducted.", "success")

    except Exception as e:
        print(f"Error submitting voucher: {e}")
        flash(f"Error: {e}", "danger")
    finally:
        if cursor: cursor.close()
        if conn: conn.close()
        
    return redirect(request.referrer)

@auth.route('/clear-filters')
def clear_filters():
    # Redirect to the referring page without query parameters
    referer = request.referrer
    if referer:
        return redirect(referer.split('?')[0])
    return redirect(url_for('auth.cashier_dashboard'))

# ----------------- CASHIER DASHBOARD -----------------
@auth.route('/cashier/dashboard')
def cashier_dashboard():
    # 1. Allow both roles
    if session.get('user_role') not in ['cashier', 'accounting']:
        return redirect(url_for('auth.admin_login'))

    # 2. Determine correct base template
    base_template = 'accounting_base.html'

    conn = None
    cursor = None
    stats = {
        'monthly_collections': 0,
        'today_collections': 0,
        'total_students': 0,
        'padre_garcia_students': 0
    }
    recent_activities = []
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("SELECT SUM(amount_paid) as total FROM payments WHERE DATE(payment_date) = CURDATE()")
        today_res = cursor.fetchone()
        stats['today_collections'] = today_res['total'] if today_res and today_res['total'] else 0

        cursor.execute("SELECT SUM(amount_paid) as total FROM payments WHERE MONTH(payment_date) = MONTH(CURRENT_DATE()) AND YEAR(payment_date) = YEAR(CURRENT_DATE())")
        month_res = cursor.fetchone()
        stats['monthly_collections'] = month_res['total'] if month_res and month_res['total'] else 0

        cursor.execute("SELECT COUNT(*) as count FROM applicants WHERE application_status = 'Enrolled'")
        total_stud_res = cursor.fetchone()
        stats['total_students'] = total_stud_res['count'] if total_stud_res else 0

        cursor.execute("SELECT COUNT(*) as count FROM applicants WHERE application_status = 'Enrolled' AND permanent_address_city_municipality LIKE '%Padre Garcia%'")
        local_res = cursor.fetchone()
        stats['padre_garcia_students'] = local_res['count'] if local_res else 0

        cursor.execute("""
            SELECT p.*, a.first_name, a.last_name 
            FROM payments p
            JOIN applicants a ON p.student_id = a.applicant_id
            ORDER BY p.payment_date DESC, p.id DESC LIMIT 5
        """)
        raw_activities = cursor.fetchall()
        
        for act in raw_activities:
            recent_activities.append({
                'student': {'first_name': act['first_name'], 'last_name': act['last_name']},
                'amount_paid': act['amount_paid'],
                'remark': act['remark'] or 'Payment',
                'payment_date': act['payment_date']
            })

    except Exception as e:
        print(f"Error loading cashier dashboard: {e}")
        traceback.print_exc()
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

    # 3. Pass base_template to render
    return render_template('cashier_dashboard.html', 
                           stats=stats, 
                           recent_activities=recent_activities, 
                           current_year=datetime.datetime.now().year,
                           base_template=base_template)

# ----------------- CASHIER SUPPORTING ROUTES -----------------
@auth.route('/cashier/student-records', methods=['GET'])
def student_records():
    # 1. Security Check
    if session.get('user_role') not in ['cashier', 'accounting']: 
        return redirect(url_for('auth.admin_login'))
    
    base_template = 'accounting_base.html'
    
    # 2. Fetch Filter Options for the UI
    all_school_years = _get_all_school_years()
    programs = _get_program_list()
    sections = _get_all_sections()
    
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # 3. Fetch Students who are/were in the enrollment process
    # This includes current students, dropouts, and those awaiting re-enrollment
    sql = """
        SELECT a.applicant_id as id, a.old_student_id as student_id, a.first_name, a.last_name,
               a.program_choice as course, a.enrollment_year_level as year_level,
               a.enrollment_semester as semester, a.academic_year, a.application_status,
               s.section_name as section
        FROM applicants a
        LEFT JOIN sections s ON a.section_id = s.id
        WHERE a.application_status IN ('Enrolled', 'Dropped', 'Not Enrolled', 'Eligible for Enrollment', 'Enrolling')
        ORDER BY a.last_name ASC
    """
    
    cursor.execute(sql)
    students = cursor.fetchall()
    
    # 4. Calculate Financial Balances for each student
    for stud in students:
        cursor.execute("SELECT SUM(balance) as total_bal FROM assessments WHERE student_id = %s", (stud['id'],))
        bal_res = cursor.fetchone()
        stud['total_balance'] = float(bal_res['total_bal']) if bal_res['total_bal'] else 0.0
        
        # Determine human-readable payment status
        if stud['total_balance'] == 0: 
            stud['payment_status'] = 'Paid'
        elif stud['total_balance'] > 0: 
            stud['payment_status'] = 'Partial'
        else: 
            stud['payment_status'] = 'Overpaid'

    cursor.close()
    conn.close()
    
    return render_template('student_records.html', 
                           students=students, 
                           all_school_years=all_school_years,
                           programs=programs,
                           sections=sections,
                           base_template=base_template)

@auth.route('/cashier/export-student-records')
def export_student_records():
    # 1. Security Check
    if session.get('user_role') not in ['cashier', 'accounting']:
        return redirect(url_for('auth.admin_login'))

    # 2. Retrieve Filter Parameters from the URL
    query_search = request.args.get('query', '').lower()
    program_filter = request.args.get('program')
    status_filter = request.args.get('status') # Admission Status (Enrolled, Dropped, etc.)

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        # 3. Base SQL Query
        # We select relevant student info and use a subquery to get the Total Balance
        sql = """
            SELECT a.applicant_id as id, 
                   a.old_student_id as student_id, 
                   a.first_name, 
                   a.last_name,
                   a.program_choice as course, 
                   a.enrollment_year_level as year_level,
                   s.section_name as section,
                   a.application_status,
                   (SELECT SUM(balance) FROM assessments WHERE student_id = a.applicant_id) as total_balance
            FROM applicants a
            LEFT JOIN sections s ON a.section_id = s.id
            WHERE a.application_status IN ('Enrolled', 'Dropped', 'Not Enrolled', 'Eligible for Enrollment', 'Enrolling')
        """
        
        cursor.execute(sql)
        all_students = cursor.fetchall()

        # 4. Filter Data in Python (to handle complex string matching and calculated statuses)
        filtered_data = []

        for stud in all_students:
            # Handle Null balance
            balance = float(stud['total_balance']) if stud['total_balance'] else 0.0
            
            # Determine Payment Status string for the Excel file
            payment_status = 'Unpaid'
            if balance == 0:
                payment_status = 'Paid'
            elif balance > 0:
                payment_status = 'Partial'
            else:
                payment_status = 'Overpaid'

            # --- Apply Filters ---
            # A. Search Filter
            if query_search:
                search_pool = f"{stud['last_name']} {stud['first_name']} {stud['student_id']}".lower()
                if query_search not in search_pool:
                    continue
            
            # B. Program Filter
            if program_filter and stud['course'] != program_filter:
                continue
            
            # C. Admission Status Filter
            if status_filter and stud['application_status'] != status_filter:
                continue

            # 5. Map Data to Excel Columns
            filtered_data.append({
                'Student ID': stud['student_id'] or 'N/A',
                'Last Name': stud['last_name'].upper(),
                'First Name': stud['first_name'].upper(),
                'Program': stud['course'],
                'Year Level': stud['year_level'],
                'Section': stud['section'] or 'Unassigned',
                'Admission Status': stud['application_status'],
                'Total Balance': balance,
                'Payment Status': payment_status
            })

        # 6. Check if data exists
        if not filtered_data:
            flash("No records found matching the current filters to export.", "warning")
            return redirect(url_for('auth.student_records'))

        # 7. Generate Excel using Pandas
        df = pd.DataFrame(filtered_data)
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Student Financial Records')
            
            # Auto-adjust column widths for better readability
            worksheet = writer.sheets['Student Financial Records']
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2
        
        output.seek(0)
        filename = f"Student_Records_Export_{datetime.date.today()}.xlsx"

        return send_file(output, 
                         as_attachment=True, 
                         download_name=filename, 
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        print(f"Export Error: {e}")
        flash("An error occurred while generating the Excel file.", "danger")
        return redirect(url_for('auth.student_records'))
    finally:
        cursor.close()
        conn.close()

@auth.route('/my-payment-history')
def student_view_payment_history():
    if 'student_id' not in session:
        flash("Please log in to view payment history.", "warning")
        return redirect(url_for('auth.student_login_page'))

    student_user_id = session['student_id']
    conn = None
    cursor = None
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # 1. FIX: Select ALL columns (*) so the template has access to names, photos, etc.
        cursor.execute("SELECT * FROM applicants WHERE student_user_id = %s ORDER BY submitted_at DESC LIMIT 1", (student_user_id,))
        student = cursor.fetchone()
        
        if not student:
            flash("No student record found.", "danger")
            return redirect(url_for('views.application_status_page'))
            
        applicant_id = student['applicant_id']

        # Fetch Payments with Assessment Description
        cursor.execute("""
            SELECT p.*, a.description as assessment_description
            FROM payments p
            LEFT JOIN assessments a ON p.assessment_id = a.id
            WHERE p.student_id = %s
            ORDER BY p.payment_date DESC, p.id DESC
        """, (applicant_id,))
        payments = cursor.fetchall()
        
        total_paid = sum(p['amount_paid'] for p in payments)
        
        # 2. FIX: Pass 'application=student' to the template
        return render_template('student_payment_history.html', 
                               payments=payments, 
                               total_paid=total_paid,
                               application=student,  
                               student_logged_in=True)

    except Exception as e:
        print(f"Error fetching payment history: {e}")
        # traceback.print_exc() # Uncomment for more details in terminal
        flash("An error occurred.", "danger")
        return redirect(url_for('views.view_student_statement'))
    finally:
        if cursor: cursor.close()
        if conn: conn.close()


@auth.route('/cashier/payment/<int:student_id_db>', methods=['GET', 'POST'])
def update_payment(student_id_db):
    # 1. Security Check
    if session.get('user_role') not in ['cashier', 'accounting']:
        return redirect(url_for('auth.admin_login'))
    
    base_template = 'accounting_base.html'
    
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # --- HANDLE POST (Record Payment) ---
        if request.method == 'POST':
            assessment_id = request.form.get('assessment_id')
            amount = float(request.form.get('amount'))
            remark = request.form.get('remark')
            payment_method = request.form.get('payment_method')
            date_str = request.form.get('payment_date')
            
            # A. Insert Payment Record
            cursor.execute("""
                INSERT INTO payments (student_id, assessment_id, amount_paid, payment_date, payment_method, remark, cashier_username)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (student_id_db, assessment_id, amount, date_str, payment_method, remark, session.get('user_name')))
            
            # Capture the new payment ID for the receipt
            new_payment_id = cursor.lastrowid

            # B. Update Assessment Balance (if linked to one)
            if assessment_id:
                # Deduct amount from balance, add to amount_paid
                cursor.execute("UPDATE assessments SET amount_paid = amount_paid + %s, balance = balance - %s WHERE id = %s", 
                               (amount, amount, assessment_id))
                
                # Check if fully paid to update status
                cursor.execute("SELECT total_amount, amount_paid FROM assessments WHERE id = %s", (assessment_id,))
                ass_data = cursor.fetchone()
                if ass_data:
                    new_status = 'paid' if ass_data['amount_paid'] >= ass_data['total_amount'] else 'partial'
                    cursor.execute("UPDATE assessments SET status = %s WHERE id = %s", (new_status, assessment_id))

            conn.commit()
            
            # C. Check System Setting for Receipt Issuance
            cursor.execute("SELECT setting_value FROM system_settings WHERE setting_key = 'issuance_of_receipt'")
            setting_row = cursor.fetchone()
            
            # Default to True (Print) if setting doesn't exist, otherwise check value
            should_print = True
            if setting_row and setting_row['setting_value'] == 'false':
                should_print = False

            # D. Redirect Logic
            if should_print:
                flash("Payment recorded successfully! Preparing receipt...", "success")
                return redirect(url_for('auth.print_receipt', payment_id=new_payment_id))
            else:
                flash("Payment recorded successfully. (Receipt printing is disabled)", "success")
                # Reload the same page to show updated balance and history
                return redirect(url_for('auth.update_payment', student_id_db=student_id_db))

        # --- HANDLE GET (View Page) ---
        
        # 1. Fetch Student Details
        cursor.execute("""
            SELECT a.applicant_id, a.old_student_id, a.control_number, 
                   a.first_name, a.last_name, a.program_choice as course,
                   a.enrollment_year_level as year_level, s.section_name as section, a.email_address as email
            FROM applicants a
            LEFT JOIN sections s ON a.section_id = s.id
            WHERE a.applicant_id = %s
        """, (student_id_db,))
        student = cursor.fetchone()
        
        if student:
            # Set display ID for UI
            student['display_id'] = student['old_student_id'] or student['control_number'] or 'N/A'

        # 2. Fetch Assessments (Fees)
        cursor.execute("SELECT * FROM assessments WHERE student_id = %s ORDER BY created_at DESC", (student_id_db,))
        assessments = cursor.fetchall()

        # 3. Fetch Payment History
        cursor.execute("""
            SELECT p.*, a.description as assessment_description
            FROM payments p
            LEFT JOIN assessments a ON p.assessment_id = a.id
            WHERE p.student_id = %s
            ORDER BY p.payment_date DESC, p.id DESC
        """, (student_id_db,))
        payment_history = cursor.fetchall()
        
    except Exception as e:
        print(f"Error in update_payment: {e}")
        traceback.print_exc()
        flash("An error occurred loading the payment page.", "danger")
        student = None
        assessments = []
        payment_history = []
    finally:
        if cursor: cursor.close()
        if conn: conn.close()
    
    return render_template('update_payment.html', 
                           student=student, 
                           assessments=assessments, 
                           payment_history=payment_history,
                           today_date=datetime.date.today().strftime('%Y-%m-%d'), 
                           base_template=base_template)

# --- IN auth.py ---

@auth.route('/cashier/student-history/<int:student_id>')
def cashier_view_history(student_id):
    if session.get('user_role') not in ['cashier', 'accounting']:
        return redirect(url_for('auth.admin_login'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 1. Fetch Student Basic Info
        cursor.execute("""
            SELECT applicant_id, old_student_id, control_number, first_name, last_name, 
                   program_choice as course, enrollment_year_level as year_level
            FROM applicants 
            WHERE applicant_id = %s
        """, (student_id,))
        student = cursor.fetchone()
        
        if student:
            student['display_id'] = student['old_student_id'] or student['control_number'] or 'N/A'

        # 2. Fetch Detailed Payment History
        # We join assessments to get the description
        # We calculate total paid as well
        cursor.execute("""
            SELECT p.*, a.description as assessment_description, 
                   # Generate a receipt number string if not exists (using ID)
                   CONCAT('BTG-', LPAD(p.id, 6, '0')) as receipt_number
            FROM payments p
            LEFT JOIN assessments a ON p.assessment_id = a.id
            WHERE p.student_id = %s
            ORDER BY p.payment_date DESC, p.id DESC
        """, (student_id,))
        history = cursor.fetchall()

        total_paid = sum(row['amount_paid'] for row in history)

    except Exception as e:
        print(f"Error fetching history: {e}")
        student = None
        history = []
        total_paid = 0
    finally:
        cursor.close()
        conn.close()

    return render_template('view_payment_history.html', 
                           student=student, 
                           history=history, 
                           total_paid=total_paid,
                           base_template='accounting_base.html')

@auth.route('/cashier/batch-assessment', methods=['GET', 'POST'])
def create_batch_assessment():
    # 1. Security Check: Allow Cashier and Accounting roles
    if session.get('user_role') not in ['cashier', 'accounting']: 
        return redirect(url_for('auth.admin_login'))
    
    base_template = 'accounting_base.html'
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # 2. Handle POST Request (Creating a new Batch Assessment)
    if request.method == 'POST':
        desc = request.form.get('description')
        amount = request.form.get('total_amount')
        target_sem = request.form.get('semester')
        target_scope = request.form.get('scope')
        target_course = request.form.get('course')
        target_yl = request.form.get('year_level')

        # Logic to apply immediately to currently enrolled students matching criteria
        # We no longer filter by academic_year
        query = "SELECT applicant_id FROM applicants WHERE application_status = 'Enrolled'"
        params = []
        
        if target_sem:
            query += " AND enrollment_semester = %s"
            params.append(target_sem)
        if target_course:
            query += " AND program_choice = %s"
            params.append(target_course)
        if target_yl:
            query += " AND enrollment_year_level = %s"
            params.append(target_yl)
        if target_scope == 'garciano':
            query += " AND LOWER(permanent_address_city_municipality) LIKE '%%padre garcia%%'"
        elif target_scope == 'non_garciano':
            query += " AND LOWER(permanent_address_city_municipality) NOT LIKE '%%padre garcia%%'"
            
        cursor.execute(query, tuple(params))
        students = cursor.fetchall()
        
        count = 0
        for s in students:
            # Check for duplicates before inserting for each student
            cursor.execute("SELECT id FROM assessments WHERE student_id = %s AND description = %s", (s['applicant_id'], desc))
            if not cursor.fetchone():
                cursor.execute("""
                    INSERT INTO assessments (student_id, description, total_amount, balance, status, created_at)
                    VALUES (%s, %s, %s, %s, 'unpaid', NOW())
                """, (s['applicant_id'], desc, amount, amount))
                count += 1
            
        # Save this rule to the Log for future auto-assignment (target_year set to 'All')
        cursor.execute("""
            INSERT INTO batch_assessments_log (description, amount, target_year, semester, course, year_level, scope, student_count, created_at)
            VALUES (%s, %s, 'All', %s, %s, %s, %s, %s, NOW())
        """, (desc, amount, target_sem or 'All', target_course or 'All', target_yl or 'All', target_scope or 'All', count))
        
        conn.commit()
        flash(f"Assessment '{desc}' created. Applied to {count} currently enrolled students.", "success")
        return redirect(url_for('auth.create_batch_assessment'))
    
    # 3. Handle GET Request (Loading the page)
    cursor.execute("SELECT * FROM batch_assessments_log ORDER BY created_at DESC")
    history = cursor.fetchall()
    
    active_term = _get_active_term()
    active_sem = active_term['semester'] if active_term else ''
    
    cursor.close()
    conn.close()
    
    return render_template('create_batch_assessment.html', 
                           history=history, 
                           active_semester=active_sem,
                           base_template=base_template)

@auth.route('/cashier/batch-assessment/edit/<int:batch_id>', methods=['POST'])
def edit_batch_assessment(batch_id):
    if session.get('user_role') not in ['cashier', 'accounting']:
        return redirect(url_for('auth.admin_login'))

    new_desc = request.form.get('description')
    new_amount = request.form.get('amount')
    new_sem = request.form.get('semester')
    new_scope = request.form.get('scope')
    new_course = request.form.get('course')
    new_level = request.form.get('year_level')

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("SELECT description FROM batch_assessments_log WHERE id = %s", (batch_id,))
        old_batch = cursor.fetchone()

        if not old_batch:
            flash("Batch record not found.", "danger")
            return redirect(url_for('auth.create_batch_assessment'))

        old_desc = old_batch['description']

        # REMOVED: updated_at = NOW() from the query below
        cursor.execute("""
            UPDATE batch_assessments_log 
            SET description = %s, 
                amount = %s, 
                semester = %s,
                scope = %s, 
                course = %s, 
                year_level = %s
            WHERE id = %s
        """, (new_desc, new_amount, new_sem, new_scope, new_course, new_level, batch_id))

        # Cascade Update to individual Assessments (Description & Amount only)
        cursor.execute("""
            UPDATE assessments 
            SET description = %s, 
                total_amount = %s,
                balance = %s - amount_paid
            WHERE description = %s 
            AND status != 'paid'
        """, (new_desc, new_amount, new_amount, old_desc))

        conn.commit()
        flash(f"Batch log updated and student balances adjusted.", "success")

    except Exception as e:
        flash(f"Error: {e}", "danger")
        if conn: conn.rollback()
    finally:
        cursor.close()
        if conn: conn.close()

    return redirect(url_for('auth.create_batch_assessment'))

@auth.route('/cashier/batch-assessment/delete/<int:batch_id>', methods=['POST'])
def delete_batch_assessment(batch_id):
    if session.get('user_role') not in ['cashier', 'accounting']:
        return redirect(url_for('auth.admin_login'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        # 1. Get details to delete linked assessments
        cursor.execute("SELECT description FROM batch_assessments_log WHERE id = %s", (batch_id,))
        batch = cursor.fetchone()

        if batch:
            # 2. Delete Log
            cursor.execute("DELETE FROM batch_assessments_log WHERE id = %s", (batch_id,))
            
            # 3. Delete individual assessments (Optional: Only delete if no payment made?)
            # Here we delete ONLY if no payments have been made (amount_paid = 0) to preserve financial history.
            cursor.execute("""
                DELETE FROM assessments 
                WHERE description = %s AND amount_paid = 0
            """, (batch['description'],))
            
            deleted_count = cursor.rowcount
            conn.commit()
            flash(f"Batch deleted. {deleted_count} unpaid student assessments were removed.", "success")
        else:
             flash("Batch not found.", "danger")

    except Exception as e:
        print(f"Error deleting batch: {e}")
        flash("Error deleting batch.", "danger")
    finally:
        cursor.close()
        conn.close()

    return redirect(url_for('auth.create_batch_assessment'))

@auth.route('/cashier/generate-report', methods=['GET'])
def generate_report():
    # 1. Check Permissions
    if session.get('user_role') not in ['cashier', 'accounting']:
        return redirect(url_for('auth.admin_login'))
    
    # 2. Determine Base Template
    base_template = 'accounting_base.html'

    # 3. Get Filter Parameters
    filter_type = request.args.get('filter_type')
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    
    course = request.args.get('course')
    year_level = request.args.get('year_level')
    section = request.args.get('section')

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # 4. Determine Date Range based on Filter Type
    start_date = None
    end_date = None
    active_filter_display = "Custom Range"
    today = datetime.date.today()

    if filter_type == 'daily':
        start_date = today
        end_date = today
        active_filter_display = f"Daily Sales ({today.strftime('%b %d, %Y')})"
    elif filter_type == 'weekly':
        end_date = today
        start_date = today - timedelta(days=7)
        active_filter_display = "Last 7 Days"
    elif filter_type == 'monthly':
        # First day of current month
        start_date = today.replace(day=1)
        # Last day of current month
        last_day = calendar.monthrange(today.year, today.month)[1]
        end_date = today.replace(day=last_day)
        active_filter_display = f"Current Month ({today.strftime('%B %Y')})"
    else:
        # Custom Range from inputs
        if start_date_str:
            start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
        if end_date_str:
            end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date()

    # 5. Build Query
    query = """
        SELECT p.payment_date, p.amount_paid, p.remark, p.payment_method,
               a.first_name, a.last_name, a.program_choice as course, 
               a.enrollment_year_level as year_level, a.enrollment_semester as semester,
               s.section_name as section
        FROM payments p
        JOIN applicants a ON p.student_id = a.applicant_id
        LEFT JOIN sections s ON a.section_id = s.id
        WHERE 1=1
    """
    params = []

    if start_date:
        query += " AND DATE(p.payment_date) >= %s"
        params.append(start_date)
    if end_date:
        query += " AND DATE(p.payment_date) <= %s"
        params.append(end_date)
        
    if course and course != 'All Courses':
        query += " AND a.program_choice = %s"
        params.append(course)
    if year_level and year_level != 'All Years':
        query += " AND a.enrollment_year_level = %s"
        params.append(year_level)
    if section and section != 'All Sections':
        query += " AND s.section_name = %s"
        params.append(section)

    query += " ORDER BY p.payment_date DESC, p.id DESC"

    try:
        cursor.execute(query, tuple(params))
        reports = cursor.fetchall()
        
        # Re-structure for template consistency (putting student data inside a 'student' dict)
        formatted_reports = []
        total_amount = 0.0
        
        for r in reports:
            total_amount += float(r['amount_paid'])
            formatted_reports.append({
                'payment_date': r['payment_date'],
                'amount_paid': r['amount_paid'],
                'remark': r['remark'],
                'student': {
                    'first_name': r['first_name'],
                    'last_name': r['last_name'],
                    'course': r['course'],
                    'year_level': r['year_level'],
                    'semester': r['semester'],
                    'section': r['section'] or 'N/A'
                }
            })

    except Exception as e:
        print(f"Error generating report: {e}")
        formatted_reports = []
        total_amount = 0.0
    finally:
        cursor.close()
        conn.close()

    # 6. Render
    return render_template('generate_report.html', 
                           reports=formatted_reports, 
                           total="{:,.2f}".format(total_amount), 
                           report_date=datetime.date.today().strftime('%B %d, %Y'),
                           active_filter_display=active_filter_display,
                           # Pass back filters to keep form populated
                           filter=filter_type,
                           start_date=start_date,
                           end_date=end_date,
                           course=course,
                           base_template=base_template)

# --- IN auth.py ---

@auth.route('/cashier/activity-history', methods=['GET'])
def activity_history():
    # 1. Permissions
    if session.get('user_role') not in ['cashier', 'accounting']:
        return redirect(url_for('auth.admin_login'))
    
    base_template = 'accounting_base.html'

    # 2. Get Filters & View Mode
    school_year = request.args.get('school_year')
    semester = request.args.get('semester')
    course = request.args.get('course')
    year_level = request.args.get('year_level')
    section = request.args.get('section')
    remark = request.args.get('remark')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # Check if "View All" is requested
    view_all = request.args.get('view_all') == 'true'

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # 3. Build Query
    query = """
        SELECT p.*, 
               a.first_name, a.last_name, a.program_choice as course, 
               a.enrollment_year_level as year_level, a.enrollment_semester as semester,
               a.academic_year,
               s.section_name as section
        FROM payments p
        JOIN applicants a ON p.student_id = a.applicant_id
        LEFT JOIN sections s ON a.section_id = s.id
        WHERE 1=1
    """
    params = []

    if school_year:
        query += " AND a.academic_year = %s"
        params.append(school_year)
    if semester:
        query += " AND a.enrollment_semester = %s"
        params.append(semester)
    if course and course != 'All Courses':
        query += " AND a.program_choice = %s"
        params.append(course)
    if year_level and year_level != 'All Years':
        query += " AND a.enrollment_year_level = %s"
        params.append(year_level)
    if section and section != 'All Sections':
        query += " AND s.section_name = %s"
        params.append(section)
    if remark:
        query += " AND p.remark = %s"
        params.append(remark)
    if start_date:
        query += " AND DATE(p.payment_date) >= %s"
        params.append(start_date)
    if end_date:
        query += " AND DATE(p.payment_date) <= %s"
        params.append(end_date)

    query += " ORDER BY p.payment_date DESC, p.id DESC"
    
    # 4. Apply LIMIT if not viewing all
    if not view_all:
        query += " LIMIT 100"

    try:
        cursor.execute(query, tuple(params))
        raw_history = cursor.fetchall()
        
        history = []
        for row in raw_history:
            history.append({
                'payment_date': row['payment_date'],
                'amount_paid': row['amount_paid'],
                'type': row['payment_method'],
                'remark': row['remark'],
                'cashier_username': row['cashier_username'],
                'student': {
                    'first_name': row['first_name'],
                    'last_name': row['last_name'],
                    'course': row['course'],
                    'section': row['section'] or 'N/A',
                    'semester': row['semester']
                }
            })
            
        # Check if there are more records than displayed (simple check)
        cursor.execute("SELECT COUNT(*) as count FROM payments")
        total_records = cursor.fetchone()['count']
        has_more = total_records > 100 and not view_all

    except Exception as e:
        print(f"Error fetching activity history: {e}")
        history = []
        has_more = False
    finally:
        cursor.close()
        conn.close()

    return render_template('activity_history.html', 
                           history=history, 
                           all_school_years=_get_all_school_years(),
                           report_date=datetime.date.today().strftime('%B %d, %Y'),
                           # Pass back filters
                           school_year=school_year, semester_filter=semester,
                           course_filter=course, year_level=year_level,
                           remark_filter=remark, start_date=start_date, end_date=end_date,
                           # View state
                           view_all=view_all,
                           has_more=has_more,
                           base_template=base_template)

@auth.route('/cashier/export-activity')
def export_activity_excel():
    if session.get('user_role') not in ['cashier', 'accounting']:
        return redirect(url_for('auth.admin_login'))

    # 1. Get Filters (Same as activity_history)
    school_year = request.args.get('school_year')
    semester = request.args.get('semester')
    course = request.args.get('course')
    year_level = request.args.get('year_level')
    section = request.args.get('section')
    remark = request.args.get('remark')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # 2. Build Query
    query = """
        SELECT p.payment_date AS 'Date',
               CONCAT(a.last_name, ', ', a.first_name) AS 'Student Name',
               a.program_choice AS 'Course',
               s.section_name AS 'Section',
               p.amount_paid AS 'Amount',
               p.payment_method AS 'Type',
               p.remark AS 'Remark',
               p.cashier_username AS 'Cashier'
        FROM payments p
        JOIN applicants a ON p.student_id = a.applicant_id
        LEFT JOIN sections s ON a.section_id = s.id
        WHERE 1=1
    """
    params = []

    if school_year:
        query += " AND a.academic_year = %s"
        params.append(school_year)
    if semester:
        query += " AND a.enrollment_semester = %s"
        params.append(semester)
    if course and course != 'All Courses':
        query += " AND a.program_choice = %s"
        params.append(course)
    if year_level and year_level != 'All Years':
        query += " AND a.enrollment_year_level = %s"
        params.append(year_level)
    if section and section != 'All Sections':
        query += " AND s.section_name = %s"
        params.append(section)
    if remark:
        query += " AND p.remark = %s"
        params.append(remark)
    if start_date:
        query += " AND DATE(p.payment_date) >= %s"
        params.append(start_date)
    if end_date:
        query += " AND DATE(p.payment_date) <= %s"
        params.append(end_date)

    query += " ORDER BY p.payment_date DESC"

    try:
        cursor.execute(query, tuple(params))
        data = cursor.fetchall()
        
        if not data:
            flash("No data to export.", "warning")
            return redirect(request.referrer)

        # 3. Create DataFrame
        df = pd.DataFrame(data)
        
        # Format output
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Activity Log')
            # Auto-adjust columns
            worksheet = writer.sheets['Activity Log']
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2
        
        output.seek(0)
        filename = f"Payment_Activity_{datetime.date.today()}.xlsx"

        return send_file(output, 
                         as_attachment=True, 
                         download_name=filename, 
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        print(f"Export error: {e}")
        flash("An error occurred during export.", "danger")
        return redirect(request.referrer)
    finally:
        cursor.close()
        conn.close()

# ----------------- ACCOUNTING DASHBOARD -----------------
@auth.route('/accounting/dashboard')
def accounting_dashboard():
    if session.get('user_role') != 'accounting':
        return redirect(url_for('auth.admin_login'))

    # For now, we can redirect to the Journal Entries as the main dashboard,
    # or render a specific summary page if you create one.
    # Here is a basic summary implementation:
    
    conn = None
    cursor = None
    total_revenue = 0
    total_expenses = 0
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Example: Calculate totals from journal lines linked to Revenue/Expense accounts
        # Assuming a 'chart_of_accounts' table exists with 'type' column
        
        cursor.execute("""
            SELECT SUM(jl.credit) as revenue 
            FROM journal_lines jl
            JOIN chart_of_accounts ca ON jl.account_id = ca.id
            WHERE ca.type = 'Revenue'
        """)
        rev_res = cursor.fetchone()
        total_revenue = rev_res['revenue'] if rev_res and rev_res['revenue'] else 0

        cursor.execute("""
            SELECT SUM(jl.debit) as expenses 
            FROM journal_lines jl
            JOIN chart_of_accounts ca ON jl.account_id = ca.id
            WHERE ca.type = 'Expense'
        """)
        exp_res = cursor.fetchone()
        total_expenses = exp_res['expenses'] if exp_res and exp_res['expenses'] else 0

    except Exception as e:
        print(f"Error loading accounting dashboard: {e}")
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

    return render_template('accounting_dashboard.html', 
                           total_revenue=total_revenue, 
                           total_expenses=total_expenses,
                           net_income=total_revenue - total_expenses)

# ----------------- ADMIN ROUTES -----------------

@auth.route('/admin/dashboard', methods=['GET'])
def admin_dashboard():
    role = session.get('user_role')
    if role not in ['admin', 'registrar']:
        if role in ['guidance', 'osa']:
            return redirect(url_for('auth.admin_enrolled_applications'))
        flash("⚠️ Log in for dashboard.", "warning")
        return redirect(url_for('auth.admin'))
    
    selected_sy = request.args.get('sy')
    all_sy = _get_all_school_years()
    applications, stats = _get_all_applications_and_stats(school_year=selected_sy)
    
    active_term = _get_active_term()
    active_school_year = f"{active_term['year_name']} - {active_term['semester']}" if active_term else "Not Set"
        
    return render_template('admin_dashboard2.html', 
                           applications=applications, 
                           stats=stats, 
                           active_page='dashboard',
                           active_school_year=active_school_year,
                           all_school_years=all_sy,
                           selected_school_year=selected_sy)


# --- ADMIN APPLICATION VIEW ROUTES ---
@auth.route('/admin/applications/all')
def admin_all_applications():
    if session.get('user_role') not in ['admin', 'registrar']: return redirect(url_for('auth.admin'))
    selected_sy = request.args.get('sy')
    all_sy = _get_all_school_years()
    applications, stats = _get_all_applications_and_stats(school_year=selected_sy)
    programs = _get_program_list()
    return _render_applications_page('admin_all_applications.html', 'all_applications', applications, stats, programs, all_sy, selected_sy, page_title="All Applications")

@auth.route('/admin/applications/pending')
def admin_pending_applications():
    if session.get('user_role') not in ['admin', 'registrar']: return redirect(url_for('auth.admin'))
    selected_sy = request.args.get('sy')
    all_sy = _get_all_school_years()
    applications, stats = _get_all_applications_and_stats(school_year=selected_sy)
    programs = _get_program_list()
    return _render_applications_page('admin_pending.html', 'pending', applications, stats, programs, all_sy, selected_sy, page_title="Pending Applications", filter_status='Pending')

@auth.route('/admin/applications/approved')
def admin_approved_applications():
    if session.get('user_role') not in ['admin', 'registrar']: return redirect(url_for('auth.admin'))
    selected_sy = request.args.get('sy')
    all_sy = _get_all_school_years()
    applications, stats = _get_all_applications_and_stats(school_year=selected_sy)
    programs = _get_program_list()
    return _render_applications_page('admin_approved.html', 'approved', applications, stats, programs, all_sy, selected_sy, page_title="Approved Applications", filter_status='Approved')

@auth.route('/admin/applications/scheduled')
def admin_scheduled_applications():
    if session.get('user_role') not in ['admin', 'registrar']: return redirect(url_for('auth.admin'))
    selected_sy = request.args.get('sy')
    all_sy = _get_all_school_years()
    applications, stats = _get_all_applications_and_stats(school_year=selected_sy)
    programs = _get_program_list()
    return _render_applications_page('admin_scheduled.html', 'scheduled', applications, stats, programs, all_sy, selected_sy, page_title="Scheduled Applications", filter_status='Scheduled')

@auth.route('/admin/applications/not-taken')
def admin_not_taken_applications():
    if session.get('user_role') not in ['admin', 'registrar']: return redirect(url_for('auth.admin'))
    selected_sy = request.args.get('sy')
    all_sy = _get_all_school_years()
    applications, stats = _get_all_applications_and_stats(school_year=selected_sy)
    programs = _get_program_list()
    return _render_applications_page('admin_not_taken.html', 'not_taken', applications, stats, programs, all_sy, selected_sy, page_title="Did Not Attend Exam", filter_exam_status='Not Taken')

@auth.route('/admin/applications/rejected')
def admin_rejected_applications():
    if session.get('user_role') not in ['admin', 'registrar']: return redirect(url_for('auth.admin'))
    selected_sy = request.args.get('sy')
    all_sy = _get_all_school_years()
    applications, stats = _get_all_applications_and_stats(school_year=selected_sy)
    programs = _get_program_list()
    return _render_applications_page('admin_rejected.html', 'rejected', applications, stats, programs, all_sy, selected_sy, page_title="Rejected Applications", filter_status='Rejected')

@auth.route('/admin/applications/passed')
def admin_passed_applications():
    if session.get('user_role') not in ['admin', 'registrar']: return redirect(url_for('auth.admin'))
    selected_sy = request.args.get('sy')
    all_sy = _get_all_school_years()
    applications, stats = _get_all_applications_and_stats(school_year=selected_sy)
    programs = _get_program_list()
    return _render_applications_page('admin_passed.html', 'passed', applications, stats, programs, all_sy, selected_sy, page_title="Passed Applications", filter_status='Passed')


@auth.route('/admin/applications/eligible')
def admin_eligible_applications():
    if session.get('user_role') not in ['admin', 'registrar']: return redirect(url_for('auth.admin'))
    selected_sy = request.args.get('sy')
    all_sy = _get_all_school_years()
    applications, stats = _get_all_applications_and_stats(school_year=selected_sy)
    programs = _get_program_list()
    sections = _get_all_sections() # <-- NEW
    return render_template('admin_eligible.html', 
                           applications=[app for app in applications if app.get('application_status') == 'Eligible for Enrollment'], 
                           stats=stats, programs=programs, sections=sections, # <-- Pass sections
                           all_school_years=all_sy, selected_school_year=selected_sy, 
                           page_title="Eligible for Enrollment", active_page='eligible')

@auth.route('/admin/applications/enrolling')
def admin_enrolling_applications():
    if session.get('user_role') not in ['admin', 'registrar']: return redirect(url_for('auth.admin'))
    selected_sy = request.args.get('sy')
    all_sy = _get_all_school_years()
    applications, stats = _get_all_applications_and_stats(school_year=selected_sy)
    programs = _get_program_list()
    sections = _get_all_sections() # <-- NEW
    return render_template('admin_enrolling.html', 
                           applications=[app for app in applications if app.get('application_status') == 'Enrolling'],
                           stats=stats, programs=programs, sections=sections, # <-- Pass sections
                           all_school_years=all_sy, selected_school_year=selected_sy, 
                           page_title="Enrolling Applications", active_page='enrolling')

@auth.route('/admin/applications/enrolled')
def admin_enrolled_applications():
    allowed_roles = ['admin', 'registrar', 'guidance', 'osa', 'president']
    if session.get('user_role') not in allowed_roles: return redirect(url_for('auth.admin'))
    selected_sy = request.args.get('sy')
    all_sy = _get_all_school_years()
    applications, stats = _get_all_applications_and_stats(school_year=selected_sy)
    programs = _get_program_list()
    sections = _get_all_sections() # <-- NEW
    return render_template('admin_enrolled.html', 
                           applications=[app for app in applications if app.get('application_status') == 'Enrolled'],
                           stats=stats, programs=programs, sections=sections, # <-- Pass sections
                           all_school_years=all_sy, selected_school_year=selected_sy, 
                           page_title="Enrolled Students", active_page='enrolled')

@auth.route('/admin/applications/dropped')
def admin_dropped_applications():
    if session.get('user_role') not in ['admin', 'registrar']: return redirect(url_for('auth.admin'))
    selected_sy = request.args.get('sy')
    all_sy = _get_all_school_years()
    applications, stats = _get_all_applications_and_stats(school_year=selected_sy)
    programs = _get_program_list()
    return _render_applications_page('admin_dropped.html', 'dropped', applications, stats, programs, all_sy, selected_sy, page_title="Dropped Students", filter_status='Dropped')

@auth.route('/admin/applications/failed')
def admin_failed_applications():
    if session.get('user_role') not in ['admin', 'registrar']: return redirect(url_for('auth.admin'))
    selected_sy = request.args.get('sy')
    all_sy = _get_all_school_years()
    applications, stats = _get_all_applications_and_stats(school_year=selected_sy)
    programs = _get_program_list()
    return _render_applications_page('admin_failed.html', 'failed', applications, stats, programs, all_sy, selected_sy, page_title="Failed Applications", filter_status='Failed')

@auth.route('/admin/applications/not-enrolled')
def admin_not_enrolled_applications():
    if session.get('user_role') not in ['admin', 'registrar']: return redirect(url_for('auth.admin'))
    selected_sy = request.args.get('sy')
    all_sy = _get_all_school_years()
    applications, stats = _get_all_applications_and_stats(school_year=selected_sy)
    programs = _get_program_list()
    return _render_applications_page('admin_not_enrolled.html', 'not_enrolled', applications, stats, programs, all_sy, selected_sy, page_title="Not Enrolled Applications", filter_status='Not Enrolled')

@auth.route('/admin/add-application')
def admin_add_application_page():
    if session.get('user_role') not in ['admin', 'registrar']: return redirect(url_for('auth.admin'))
    
    active_term = _get_active_term()
    if not active_term:
        flash("Cannot add application: Admissions are currently closed (no active school year).", "warning")
        return redirect(url_for('auth.admin_dashboard'))

    _, stats = _get_all_applications_and_stats() # We only need stats for the sidebar
    today_date = datetime.date.today().strftime('%Y-%m-%d')
    programs_list = _get_program_list()
    return render_template('admin_new_applicants.html', stats=stats, active_page='add_application', today_date_for_form=today_date, programs=programs_list, active_school_year=active_term['year_name'])

@auth.route('/admin/application/<int:applicant_id>/update-inventory', methods=['POST'])
def admin_update_inventory(applicant_id):
    # 1. Check Role
    if session.get('user_role') not in ['admin', 'registrar', 'guidance', 'osa']:
        return jsonify({"success": False, "message": "Unauthorized role."}), 401

    # 2. Check Permission
    if not session.get('can_edit_student'):
        return jsonify({"success": False, "message": "❌ Permission Denied: Edit privilege is disabled for your account."}), 403

    conn = None; cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # 3. Handle Address Unification (Address fields are often the cause of errors)
        # We use .get() to avoid KeyErrors if field is missing in form
        street = request.form.get('address_street', '').strip()
        city = request.form.get('address_city', '').strip()
        province = request.form.get('address_province', '').strip()

        update_clauses = []
        update_values = []

        # Only update address if fields are present in the form submission
        if 'address_city' in request.form:
            update_clauses.append("permanent_address_street_barangay = %s")
            update_values.append(street)
            
            update_clauses.append("permanent_address_city_municipality = %s")
            update_values.append(city)
            
            update_clauses.append("permanent_address_province = %s")
            update_values.append(province)

            # Combined Inventory Address
            full_address_parts = [part for part in [street, city, province] if part]
            full_address_string = ", ".join(full_address_parts)
            update_clauses.append("inventory_complete_address = %s")
            update_values.append(full_address_string)

        # 4. Standard Inventory Fields
        inventory_fields = [
            'inventory_gender', 'inventory_age', 'inventory_religion', 
            'inventory_mobile_number', 'inventory_facebook_account', 
            'inventory_health_condition', 'inventory_interest_hobbies',
            'inventory_pre_elementary_school', 'inventory_pre_elementary_dates', 'inventory_pre_elementary_awards',
            'inventory_elementary_school', 'inventory_elementary_dates', 'inventory_elementary_awards',
            'inventory_secondary_school', 'inventory_secondary_dates', 'inventory_secondary_awards',
            'inventory_vocational_school', 'inventory_vocational_dates', 'inventory_vocational_awards',
            'inventory_father_name', 'inventory_father_age', 'inventory_father_status', 
            'inventory_father_education', 'inventory_father_occupation', 'inventory_father_contact',
            'inventory_mother_name', 'inventory_mother_age', 'inventory_mother_status', 
            'inventory_mother_education', 'inventory_mother_occupation', 'inventory_mother_contact',
            'inventory_guardian_name', 'inventory_guardian_relationship', 'inventory_guardian_contact',
            'inventory_parents_marital_status', 'inventory_number_of_children', 
            'inventory_number_of_brothers', 'inventory_number_of_sisters', 
            'inventory_family_income', 'inventory_family_description',
            'inventory_favorite_colors', 'inventory_favorite_sports', 'inventory_favorite_foods',
            'inventory_emergency_contact_name', 'inventory_emergency_contact_number', 'inventory_emergency_contact_relationship'
        ]

        # Sync Map (Inventory -> Main Profile)
        sync_map = {
            'inventory_gender': 'sex',
            'inventory_religion': 'religion',
            'inventory_mobile_number': 'mobile_number',
            'inventory_father_name': 'father_name',
            'inventory_father_occupation': 'father_occupation',
            'inventory_father_contact': 'father_contact_number',
            'inventory_mother_name': 'mother_maiden_name',
            'inventory_mother_occupation': 'mother_occupation',
            'inventory_mother_contact': 'mother_contact_number',
            'inventory_guardian_name': 'guardian_name',
            'inventory_guardian_contact': 'guardian_contact_number',
            'inventory_family_income': 'average_family_income'
        }

        for field in inventory_fields:
            if field in request.form:
                val = request.form.get(field)
                
                # Update Inventory Field
                update_clauses.append(f"{field} = %s")
                update_values.append(val)

                # Sync to Main Profile
                if field in sync_map:
                    main_col = sync_map[field]
                    update_clauses.append(f"{main_col} = %s")
                    update_values.append(val)

        # 5. Execute Update
        if update_clauses:
            update_clauses.append("last_updated_at = %s")
            update_values.append(datetime.datetime.now())
            
            update_values.append(applicant_id) # For WHERE clause

            sql = f"UPDATE applicants SET {', '.join(update_clauses)} WHERE applicant_id = %s"
            
            # DEBUG PRINT: Check what is being sent to DB
            # print(f"Executing SQL: {sql}")
            # print(f"Values: {update_values}")

            cursor.execute(sql, tuple(update_values))
            conn.commit()
            return jsonify({"success": True, "message": "Inventory updated successfully."})
        else:
            return jsonify({"success": True, "message": "No fields to update."})

    except mysql.connector.Error as db_err:
        print(f"Database Error in admin_update_inventory: {db_err}")
        return jsonify({"success": False, "message": f"Database Error: {db_err}"}), 500
    except Exception as e:
        print(f"General Error in admin_update_inventory: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "message": f"Server error: {str(e)}"}), 500
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

@auth.route('/admin/reports/grades', methods=['GET'])
def admin_grade_reports():
    if session.get('user_role') not in ['admin', 'registrar']:
        return redirect(url_for('auth.admin'))

    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Fetch options for dropdowns
        cursor.execute("SELECT DISTINCT year_name FROM academic_terms ORDER BY year_name DESC")
        school_years = cursor.fetchall()
        
        cursor.execute("SELECT program_id, title FROM programs ORDER BY title")
        programs = cursor.fetchall()
        
        cursor.execute("SELECT id, section_name FROM sections WHERE is_active=TRUE ORDER BY section_name")
        sections = cursor.fetchall()
        
        # Get active term for default selection
        active_term = _get_active_term()
        
    except Exception as e:
        print(f"Error loading report page: {e}")
        flash("Error loading form data.", "danger")
        return redirect(url_for('auth.admin_dashboard'))
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

    return render_template('admin_export_grade_sheet.html', 
                           school_years=school_years, 
                           programs=programs, 
                           sections=sections,
                           active_term=active_term)

@auth.route('/api/get-program-subjects/<program_id>/<year>/<semester>')
def api_get_program_subjects(program_id, year, semester):
    # API to populate Subject dropdown dynamically
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # Map string semester to int if necessary, or matches DB
    sem_map = {'1st Semester': 1, '2nd Semester': 2, 'Summer': 3}
    sem_int = sem_map.get(semester, 1)
    
    # Clean up year level (e.g., "1st Year" -> 1)
    year_int = 1
    if '1' in year: year_int = 1
    elif '2' in year: year_int = 2
    elif '3' in year: year_int = 3
    elif '4' in year: year_int = 4
    
    query = """
        SELECT id, subject_code, subject_title 
        FROM subjects 
        WHERE program_id = %s AND year_level = %s AND semester = %s
        ORDER BY subject_code
    """
    cursor.execute(query, (program_id, year_int, sem_int))
    subjects = cursor.fetchall()
    
    cursor.close()
    conn.close()
    return jsonify(subjects)



@auth.route('/admin/enrollment-review/<int:applicant_id>')
def admin_enrollment_review_page(applicant_id):
    if session.get('user_role') not in ['admin', 'registrar']: return redirect(url_for('auth.admin'))

    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        if not conn:
            flash("Database connection error.", "danger")
            return redirect(url_for('auth.admin_enrolling_applications'))

        cursor = conn.cursor(dictionary=True)
        # MODIFIED: Removed approval columns from this query as they are now term-specific
        cursor.execute("""
            SELECT a.*, p.program_id
            FROM applicants a
            LEFT JOIN programs p ON a.program_choice = p.title
            WHERE a.applicant_id = %s
        """, (applicant_id,))
        application = cursor.fetchone()

        if not application or application['application_status'] not in ['Enrolling', 'Enrolled']:
            flash("Enrollment application not found or is not in a valid status for review.", "warning")
            return redirect(url_for('auth.admin_enrolling_applications'))
        
        # NEW LOGIC: Handle term selection and fetch term-specific data
        all_academic_terms = []
        cursor.execute("SELECT id, year_name, semester FROM academic_terms ORDER BY year_name DESC, semester DESC")
        all_academic_terms = cursor.fetchall()

        selected_term_id = request.args.get('term_id', type=int)
        if not selected_term_id:
             active_term = _get_active_term()
             if active_term:
                 selected_term_id = active_term['id']
             elif all_academic_terms:
                 selected_term_id = all_academic_terms[0]['id']

        enrollment_record = None
        if selected_term_id and application.get('student_user_id'):
            cursor.execute("""
                SELECT * FROM student_enrollment_records
                WHERE student_user_id = %s AND academic_term_id = %s
            """, (application['student_user_id'], selected_term_id))
            enrollment_record = cursor.fetchone()

        student_user_id = application.get('student_user_id')
        program_id = application.get('program_id')

        # --- MODIFIED LOGIC ---
        has_grades = False
        subjects_by_year = {}
        student_grades = {}
        student_academic_status = None # Default to None

        if student_user_id and program_id:
            cursor.execute("SELECT 1 FROM student_grades WHERE student_user_id = %s LIMIT 1", (student_user_id,))
            if cursor.fetchone():
                has_grades = True

            # Determine academic status if student is currently enrolled (even if no grades yet) OR has a grade history.
            if application['application_status'] == 'Enrolled' or has_grades:
                student_academic_status, _ = _get_student_status_and_failed_subjects(cursor, student_user_id)
        # --- END MODIFIED LOGIC ---

            cursor.execute("""
                SELECT id, subject_code, subject_title, year_level, semester 
                FROM subjects WHERE program_id = %s ORDER BY year_level, semester, subject_code
            """, (program_id,))
            all_subjects = cursor.fetchall()

            cursor.execute("""
                SELECT subject_id, grade, remarks FROM student_grades WHERE student_user_id = %s
            """, (student_user_id,))
            grades_list = cursor.fetchall()
            student_grades = {grade['subject_id']: grade for grade in grades_list}

            for subject in all_subjects:
                year = subject['year_level']
                semester = subject['semester']
                if year not in subjects_by_year:
                    subjects_by_year[year] = {}
                if semester not in subjects_by_year[year]:
                    subjects_by_year[year][semester] = []
                subjects_by_year[year][semester].append(subject)

        inventory_fields_map = {
            'inventory_gender': 'Gender', 'inventory_age': 'Age', 'inventory_religion': 'Religion', 
            'inventory_complete_address': 'Complete Address', 'inventory_mobile_number': 'Mobile Number', 
            'inventory_facebook_account': 'Facebook Account', 'inventory_interest_hobbies': 'Interests/Hobbies', 
            'inventory_health_condition': 'Health Condition', 'inventory_pre_elementary_school': 'Pre-Elementary School', 
            'inventory_pre_elementary_dates': 'Pre-Elementary Dates', 'inventory_pre_elementary_awards': 'Pre-Elementary Awards', 
            'inventory_elementary_school': 'Elementary School', 'inventory_elementary_dates': 'Elementary Dates', 
            'inventory_elementary_awards': 'Elementary Awards', 'inventory_secondary_school': 'Secondary School', 
            'inventory_secondary_dates': 'Secondary Dates', 'inventory_secondary_awards': 'Secondary Awards', 
            'inventory_vocational_school': 'Vocational School', 'inventory_vocational_dates': 'Vocational Dates', 
            'inventory_vocational_awards': 'Vocational Awards', 'inventory_father_name': "Father's Name", 
            'inventory_father_age': "Father's Age", 'inventory_father_status': "Father's Status", 
            'inventory_father_education': "Father's Education", 'inventory_father_occupation': "Father's Occupation", 
            'inventory_father_contact': "Father's Contact", 'inventory_mother_name': "Mother's Name", 
            'inventory_mother_age': "Mother's Age", 'inventory_mother_status': "Mother's Status", 
            'inventory_mother_education': "Mother's Education", 'inventory_mother_occupation': "Mother's Occupation", 
            'inventory_mother_contact': "Mother's Contact", 'inventory_parents_marital_status': "Parents' Marital Status", 
            'inventory_number_of_children': 'Number of Children', 'inventory_number_of_brothers': 'Number of Brothers', 
            'inventory_number_of_sisters': 'Number of Sisters', 'inventory_guardian_name': "Guardian's Name", 
            'inventory_guardian_relationship': "Guardian's Relationship", 'inventory_guardian_contact': "Guardian's Contact", 
            'inventory_family_income': 'Family Income', 'inventory_favorite_colors': 'Favorite Colors', 
            'inventory_favorite_sports': 'Favorite Sports', 'inventory_favorite_foods': 'Favorite Foods', 
            'inventory_family_description': 'Family Description', 'inventory_emergency_contact_name': 'Emergency Contact Name', 
            'inventory_emergency_contact_number': 'Emergency Contact Number', 'inventory_emergency_contact_relationship': 'Emergency Contact Relationship'
        }

        enrollment_details = {
            'Student Type': application.get('enrollment_student_type'),
            'Year Level': application.get('enrollment_year_level'),
            'Semester': application.get('enrollment_semester')
        }

        inventory_data = {label: application.get(db_field) for db_field, label in inventory_fields_map.items() if application.get(db_field)}

        documents = {
            'Initial': {
                '2x2 Photo': bool(application.get('photo')),
                'SHS Diploma': bool(application.get('shs_diploma_file')),
                'SHS Card': bool(application.get('shs_card_file')),
                'Birth Certificate': bool(application.get('birth_certificate_file'))
            },
            'Enrollment': {
                'Form 138': bool(application.get('shs_card_file')),
                'SHS Diploma (Enrollment)': bool(application.get('shs_diploma_file')),
                'Good Moral Cert.': bool(application.get('good_moral_file')),
                'PSA Birth Cert.': bool(application.get('birth_certificate_file')),
                '2x2 Photos': bool(application.get('photos_2x2_file')),
                'Entrance Fee Proof': bool(application.get('entrance_fee_proof_file')),
                "Voter's ID": bool(application.get('voters_id_file')),
                'Verified CBS': bool(application.get('cbs_file')),
                'Brgy. Certification': bool(application.get('brgy_cert_file'))
            }
        }

        _, stats = _get_all_applications_and_stats()

        return render_template('admin_enrollment_review.html', 
                               application=application, 
                               stats=stats,
                               enrollment_details=enrollment_details,
                               inventory_data=inventory_data,
                               documents=documents,
                               has_grades=has_grades,
                               subjects_by_year=subjects_by_year,
                               student_grades=student_grades,
                               student_academic_status=student_academic_status,
                               all_academic_terms=all_academic_terms,
                               selected_term_id=selected_term_id,
                               enrollment_record=enrollment_record,
                               active_page='enrolling')

    except Exception as e:
        flash(f"Error loading enrollment review page: {e}", "danger")
        traceback.print_exc()
        return redirect(url_for('auth.admin_enrolling_applications'))
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()


# --- NEW ROUTE FOR HANDLING APPROVAL CHECKBOXES ---
@auth.route('/admin/enrollment/<int:applicant_id>/toggle-approval', methods=['POST'])
def admin_toggle_enrollment_approval(applicant_id):
    if session.get('user_role') not in ['admin', 'registrar']:
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    
    data = request.get_json()
    role = data.get('role')
    status = data.get('status', False)
    academic_term_id = data.get('term_id')

    if not all([role, academic_term_id]):
         return jsonify({"success": False, "message": "Missing role or term ID."}), 400

    valid_roles = ['president', 'accounting', 'dean']
    if role not in valid_roles:
        return jsonify({"success": False, "message": "Invalid role specified."}), 400

    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({"success": False, "message": "Database connection error."}), 500
        
        cursor = conn.cursor(dictionary=True)
        
        # Get student_user_id from applicant_id
        cursor.execute("SELECT student_user_id FROM applicants WHERE applicant_id = %s", (applicant_id,))
        applicant = cursor.fetchone()
        if not applicant or not applicant.get('student_user_id'):
            return jsonify({"success": False, "message": "Applicant not found or not linked to a student account."}), 404
        
        student_user_id = applicant['student_user_id']

        # Build query dynamically but safely
        flag_column = f"is_approved_by_{role}"
        date_column = f"{role}_approval_date"
        approval_date = datetime.datetime.now() if status else None

        # Use INSERT ... ON DUPLICATE KEY UPDATE to handle both new and existing records
        update_sql = f"""
            INSERT INTO student_enrollment_records (student_user_id, academic_term_id, {flag_column}, {date_column})
            VALUES (%s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
                {flag_column} = VALUES({flag_column}),
                {date_column} = VALUES({date_column})
        """
        
        insert_cursor = conn.cursor()
        insert_cursor.execute(update_sql, (student_user_id, academic_term_id, status, approval_date))
        conn.commit()
        insert_cursor.close()

        message = f"{role.title()} approval has been {'recorded' if status else 'revoked'} for this term."
        return jsonify({"success": True, "message": message})

    except mysql.connector.Error as err:
        if conn: conn.rollback()
        print(f"DB Error toggling approval for applicant {applicant_id}: {err}")
        traceback.print_exc()
        return jsonify({"success": False, "message": f"Database error: {err.msg}"}), 500
    except Exception as e:
        if conn: conn.rollback()
        print(f"Error toggling approval for applicant {applicant_id}: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "message": "A server error occurred."}), 500
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()
# --- END NEW ROUTE ---

@auth.route('/admin/enrollment/<int:applicant_id>/action', methods=['POST'])
def admin_enrollment_action(applicant_id):
    # 1. Security Check: Restrict to authorized roles
    if session.get('user_role') not in ['admin', 'registrar']:
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    
    data = request.get_json()
    action = data.get('action') # 'approve', 'reject', or 'drop'
    rejection_reason = data.get('reason')
    academic_term_id = data.get('term_id')
    
    # Clean manual_section_id (Ensures empty strings/None are handled for MySQL INT column)
    manual_section_id = data.get('manual_section_id')
    if manual_section_id == "" or manual_section_id == "None" or manual_section_id is None:
        manual_section_id = None

    if not action or not academic_term_id:
        return jsonify({"success": False, "message": "Invalid action or missing term ID."}), 400

    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # 2. Fetch the student's CURRENT data (Includes levels set by student during re-enrollment)
        cursor.execute("""
            SELECT student_user_id, first_name, last_name, email_address, 
                   program_choice, enrollment_year_level, enrollment_semester, 
                   enrollment_student_type, section_id, old_student_id, 
                   control_number, academic_year, permanent_address_city_municipality,
                   admin_notes
            FROM applicants WHERE applicant_id = %s
        """, (applicant_id,))
        app_data = cursor.fetchone()

        if not app_data:
            return jsonify({"success": False, "message": "Application not found."}), 404
        
        student_user_id = app_data['student_user_id']
        generated_id_msg = "" 
        fee_msg = ""

        # ==========================================================
        # ACTION: APPROVE (Finalize Enrollment & Balance)
        # ==========================================================
        if action == 'approve':
            new_status_record = 'Enrolled'
            
            # A. SECTION FINALIZATION
            if manual_section_id:
                # Use Admin override
                cursor.execute("UPDATE applicants SET section_id = %s, is_section_permanent = TRUE WHERE applicant_id = %s", 
                               (manual_section_id, applicant_id))
            elif not app_data['section_id']:
                # Attempt auto-assign if no section exists (New students or shifters)
                cursor.execute("SELECT program_id FROM programs WHERE title = %s", (app_data['program_choice'],))
                prog_res = cursor.fetchone()
                if prog_res:
                    _assign_section_automatically(cursor, applicant_id, prog_res['program_id'], app_data['enrollment_year_level'])
            
            # B. PERMANENT STUDENT ID GENERATION
            # If New Student and ID still starts with 'A' (Temporary), generate a permanent 'P' ID.
            if str(app_data.get('enrollment_student_type')).capitalize() == 'New':
                current_id = str(app_data.get('old_student_id') or app_data.get('control_number') or "")
                if not current_id.startswith('P'):
                    try:
                        new_permanent_id = _generate_student_id(cursor)
                        cursor.execute("UPDATE applicants SET old_student_id = %s, control_number = %s WHERE applicant_id = %s", 
                                       (new_permanent_id, new_permanent_id, applicant_id))
                        cursor.execute("UPDATE student_users SET old_student_id = %s WHERE id = %s", 
                                       (new_permanent_id, student_user_id))
                        generated_id_msg = f" New ID: {new_permanent_id} assigned."
                    except Exception as e:
                        print(f"ID Generation Error: {e}")

            # C. CUMULATIVE FINANCIAL TRIGGER
            # Assigns batch fees matching the student's CURRENT level.
            # Adds rows to the ledger even if an old balance exists.
            fees_added = _auto_assign_fees(cursor, applicant_id, app_data)
            if fees_added > 0:
                fee_msg = f" {fees_added} new assessments added to ledger."

            # D. Finalize Applicant Status
            cursor.execute("""
                UPDATE applicants 
                SET application_status = 'Enrolled', 
                    original_enrollment_status = NULL, 
                    last_updated_at = NOW() 
                WHERE applicant_id = %s
            """, (applicant_id,))

        # ==========================================
        # ACTION: DROP
        # ==========================================
        elif action == 'drop':
            new_status_record = 'Dropped'
            cursor.execute("UPDATE applicants SET application_status = 'Dropped', last_updated_at = NOW() WHERE applicant_id = %s", (applicant_id,))
        
        # ==========================================
        # ACTION: REJECT (Send back to student)
        # ==========================================
        elif action == 'reject':
            new_status_record = 'Enrolling' # Student can fix and re-submit
            cursor.execute("UPDATE applicants SET application_status = 'Enrolling', last_updated_at = NOW() WHERE applicant_id = %s", (applicant_id,))
            if rejection_reason:
                current_notes = app_data.get('admin_notes', '') or ''
                new_note = f"\n[{datetime.datetime.now().strftime('%Y-%m-%d')}] Enrollment Rejected: {rejection_reason}"
                cursor.execute("UPDATE applicants SET admin_notes = %s WHERE applicant_id = %s", (current_notes + new_note, applicant_id))

        # 3. Update the Official Term Record (History)
        cursor.execute("""
            UPDATE student_enrollment_records 
            SET enrollment_status = %s 
            WHERE student_user_id = %s AND academic_term_id = %s
        """, (new_status_record, student_user_id, academic_term_id))

        conn.commit()
        
        # 4. Notify Student via Email (Asynchronous call recommended if setup)
        applicant_name = f"{app_data['first_name']} {app_data['last_name']}"
        send_application_status_email(app_data['email_address'], applicant_name, new_status_record, applicant_id)
        
        return jsonify({
            "success": True, 
            "message": f"Action '{action}' processed.{generated_id_msg}{fee_msg}",
            "reload": True
        })

    except Exception as e:
        if conn: conn.rollback()
        print(f"CRITICAL ERROR in admin_enrollment_action: {e}")
        return jsonify({"success": False, "message": f"Server Error: {str(e)}"}), 500
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

# NEW: Admin Document Upload Route
@auth.route('/admin/enrollment/<int:applicant_id>/upload-doc', methods=['POST'])
def admin_upload_enrollment_doc(applicant_id):
    if session.get('user_role') not in ['admin', 'registrar']:
        return jsonify({"success": False, "message": "Unauthorized"}), 401

    doc_type = request.form.get('doc_type')
    file_storage = request.files.get('file')

    if not doc_type or not file_storage or not file_storage.filename:
        return jsonify({"success": False, "message": "Missing document type or file."}), 400

    # CORRECTED: This map now points to the correct, prefixed DB columns from your schema.
    doc_map = {
        'shs_card': ('enrollment_shs_card', 'Form 138 (SHS Card)'),
        'shs_diploma': ('enrollment_shs_diploma', 'SHS Diploma'),
        'good_moral': ('enrollment_good_moral', 'Good Moral Cert.'),
        'psa_birth': ('enrollment_psa_birth', 'PSA Birth Cert.'),
        'photos_2x2': ('enrollment_photos_2x2', '2x2 Photos'),
        'entrance_fee_proof': ('enrollment_entrance_fee_proof', 'Entrance Fee Proof'),
        'voters_id': ('enrollment_voters_id', "Voter's ID"),
        'cbs': ('enrollment_cbs', 'Verified CBS'),
        'brgy_cert': ('enrollment_brgy_cert', 'Brgy. Certification')
    }

    if doc_type not in doc_map:
        return jsonify({"success": False, "message": "Invalid document type."}), 400

    db_prefix, doc_desc = doc_map[doc_type]
    
    file_data, filename, mimetype, error = process_uploaded_file(file_storage, doc_desc, max_size_mb=5)
    if error:
        return jsonify({"success": False, "message": error}), 400

    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({"success": False, "message": "Database connection error."}), 500
        
        cursor = conn.cursor()
        
        # Dynamically build the update query
        update_sql = f"""
            UPDATE applicants SET 
                {db_prefix}_file = %s,
                {db_prefix}_filename = %s,
                {db_prefix}_mimetype = %s,
                last_updated_at = NOW()
            WHERE applicant_id = %s
        """
        cursor.execute(update_sql, (file_data, filename, mimetype, applicant_id))
        conn.commit()
        
        return jsonify({"success": True, "message": f"'{doc_desc}' uploaded successfully."})

    except Exception as e:
        if conn: conn.rollback()
        print(f"Error uploading document for applicant {applicant_id}: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "message": "Server error during file upload."}), 500
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()


@auth.route('/admin/application/<int:applicant_id>/permit-details', methods=['POST'])
def admin_save_permit_details(applicant_id):
    if session.get('user_role') not in ['admin', 'registrar']: return jsonify({"success": False, "message": "Unauthorized"}), 401
    
    form_permit_exam_date = request.form.get('permit_exam_date')
    form_permit_exam_time = request.form.get('permit_exam_time')
    form_permit_testing_room = request.form.get('permit_testing_room')

    conn = None; cursor = None
    
    try:
        conn = get_db_connection()
        if not conn: return jsonify({"success": False, "message": "Database connection error"}), 500
        
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT applicant_id, old_student_id, control_number FROM applicants WHERE applicant_id = %s", (applicant_id,))
        application = cursor.fetchone()

        if not application:
            return jsonify({"success": False, "message": "Application not found."}), 404

        permit_exam_date_to_db = None
        if form_permit_exam_date and form_permit_exam_date.strip():
            try: 
                permit_exam_date_to_db = datetime.datetime.strptime(form_permit_exam_date, '%Y-%m-%d').date()
            except ValueError: 
                return jsonify({"success": False, "message": "Invalid date format for exam date. Use YYYY-MM-DD."}), 400
        
        # When saving/updating permit details (i.e., rescheduling), reset exam_status to NULL
        cursor.execute("""
            UPDATE applicants 
            SET permit_exam_date=%s, permit_exam_time=%s, permit_testing_room=%s, exam_status=NULL, last_updated_at=NOW() 
            WHERE applicant_id=%s
        """, (
            permit_exam_date_to_db,
            form_permit_exam_time.strip() if form_permit_exam_time else None, 
            form_permit_testing_room.strip() if form_permit_testing_room else None, 
            applicant_id
        ))
        conn.commit()
        
        display_id = application.get('old_student_id') or application.get('control_number') or f"App #{applicant_id}"

        if cursor.rowcount > 0: 
            cursor.execute("SELECT permit_control_no FROM applicants WHERE applicant_id = %s", (applicant_id,))
            current_pcn_data = cursor.fetchone()
            current_pcn = current_pcn_data['permit_control_no'] if current_pcn_data else None

            saved_data = {
                "permit_control_no": current_pcn,
                "permit_exam_date": str(permit_exam_date_to_db) if permit_exam_date_to_db else None,
                "permit_exam_time": form_permit_exam_time.strip() if form_permit_exam_time else None,
                "permit_testing_room": form_permit_testing_room.strip() if form_permit_testing_room else None
            }
            return jsonify({"success": True, "message": f"Permit details for {display_id} saved and exam status reset.", "data": saved_data })

        cursor.execute("SELECT COUNT(*) as count FROM applicants WHERE applicant_id = %s", (applicant_id,))
        check_app_exists = cursor.fetchone()
        if check_app_exists and check_app_exists['count'] > 0:
             return jsonify({"success": True, "message": "Permit details saved (no change detected)."})
        
        return jsonify({"success": False, "message": "Application not found during update attempt."}), 404
    except Exception as e:
        print(f"Error saving permit details for {applicant_id}: {e}"); traceback.print_exc()
        return jsonify({"success": False, "message": "Server error saving permit details."}), 500
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()


@auth.route('/admin/application/<int:applicant_id>/update-ids', methods=['POST'])
def admin_update_application_ids(applicant_id):
    if session.get('user_role') not in ['admin', 'registrar']:
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    
    old_student_id = request.form.get('old_student_id', '').strip() or None
    control_number = request.form.get('control_number', '').strip() or None

    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({"success": False, "message": "Database connection error."}), 500
        
        cursor = conn.cursor(dictionary=True)
        
        # Security check: If a new old_student_id is provided, check if it's already in use by another student account
        if old_student_id:
            cursor.execute(
                "SELECT u.id FROM student_users u JOIN applicants a ON u.id = a.student_user_id WHERE u.old_student_id = %s AND a.applicant_id != %s", 
                (old_student_id, applicant_id)
            )
            if cursor.fetchone():
                return jsonify({"success": False, "message": f"The ID '{old_student_id}' is already in use by another student."}), 409

        # First, find the student_user_id
        cursor.execute("SELECT student_user_id FROM applicants WHERE applicant_id = %s", (applicant_id,))
        app_data = cursor.fetchone()

        if not app_data or not app_data.get('student_user_id'):
             return jsonify({"success": False, "message": "Could not find associated student account."}), 404

        student_user_id = app_data['student_user_id']
        
        # Update applicants table
        cursor.execute("""
            UPDATE applicants 
            SET old_student_id = %s, control_number = %s, last_updated_at = NOW() 
            WHERE applicant_id = %s
        """, (old_student_id, control_number, applicant_id))
        
        # Update student_users table
        cursor.execute("""
            UPDATE student_users
            SET old_student_id = %s, updated_at = NOW()
            WHERE id = %s
        """, (old_student_id, student_user_id))
        
        conn.commit()
        
        return jsonify({"success": True, "message": "Student IDs updated successfully."})

    except mysql.connector.Error as err:
        if conn: conn.rollback()
        print(f"DB Error updating IDs for applicant {applicant_id}: {err}")
        traceback.print_exc()
        return jsonify({"success": False, "message": f"Database error: {err.msg}"}), 500
    except Exception as e:
        if conn: conn.rollback()
        print(f"Error updating IDs for applicant {applicant_id}: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "message": "A server error occurred."}), 500
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()

# ----------------- NEW: ADMIN GRADES MANAGEMENT SECTION -----------------
@auth.route('/admin/grades/<int:student_user_id>')
def admin_manage_grades_page(student_user_id):
    if session.get('user_role') not in ['admin', 'registrar', 'guidance', 'osa']:
        return redirect(url_for('auth.admin'))

    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        if not conn:
            flash("Database connection error.", "danger")
            return redirect(url_for('auth.admin_enrolled_applications'))

        cursor = conn.cursor(dictionary=True)
        
        # 1. Get Student Info
        cursor.execute("""
            SELECT 
                u.id as student_user_id, u.old_student_id, 
                a.first_name, a.last_name, a.program_choice, p.program_id, 
                a.academic_year, a.application_status, a.control_number,
                a.enrollment_year_level, a.enrollment_semester
            FROM student_users u
            JOIN applicants a ON u.id = a.student_user_id
            JOIN programs p ON a.program_choice = p.title
            WHERE u.id = %s
            ORDER BY a.submitted_at DESC
            LIMIT 1
        """, (student_user_id,))
        student = cursor.fetchone()

        if not student:
            flash("Student not found or has no application.", "warning")
            return redirect(url_for('auth.admin_enrolled_applications'))

        # 2. Get all subjects for the student's program
        cursor.execute("""
            SELECT s.id, s.subject_code, s.subject_title, s.units, s.year_level, s.semester, s.prerequisite_subject_id, 
                   p.subject_code as prerequisite_code 
            FROM subjects s
            LEFT JOIN subjects p ON s.prerequisite_subject_id = p.id
            WHERE s.program_id = %s 
            ORDER BY s.year_level, s.semester, s.subject_code
        """, (student['program_id'],))
        all_subjects = cursor.fetchall()

        # 3. Get existing grades (Updated to include percentage)
        cursor.execute("""
            SELECT subject_id, grade, percentage, remarks, academic_year, semester 
            FROM student_grades 
            WHERE student_user_id = %s
        """, (student_user_id,))
        grades_list = cursor.fetchall()
        
        # Map grades by subject_id for easy lookup in template
        student_grades = {grade['subject_id']: grade for grade in grades_list}
        
        # 4. Determine academic status
        student_academic_status, failed_or_incomplete_ids = _get_student_status_and_failed_subjects(cursor, student_user_id)
        
        # 5. Get student's current term context
        student_current_year = None
        student_current_sem = None
        if student.get('enrollment_year_level') and student.get('enrollment_semester'):
            year_map = {'1st Year': 1, '2nd Year': 2, '3rd Year': 3, '4th Year': 4, '5th Year': 5}
            sem_map = {'1st Semester': 1, '2nd Semester': 2, 'Summer': 3, 'Summer / Midyear': 3}
            student_current_year = year_map.get(student['enrollment_year_level'])
            student_current_sem = sem_map.get(student['enrollment_semester'])

        # 6. Calculate Subject Availability
        subject_id_to_code_map = {s['id']: s['subject_code'] for s in all_subjects}
        disabled_subjects_info = {}
        for subject in all_subjects:
            prereq_id = subject.get('prerequisite_subject_id')
            subject_year = subject.get('year_level')
            subject_sem = subject.get('semester')
            is_disabled = False
            reason = ""

            if prereq_id and prereq_id in failed_or_incomplete_ids:
                prereq_code = subject_id_to_code_map.get(prereq_id, 'Unknown')
                is_disabled = True
                reason = f"Prerequisite '{prereq_code}' is Failed or Incomplete."

            if not is_disabled and student_current_year is not None and student_current_sem is not None:
                if subject_year > student_current_year:
                    is_disabled = True
                    reason = "Subject is for a higher year level."
                elif subject_year == student_current_year and subject_sem > student_current_sem:
                    is_disabled = True
                    reason = "Subject is for a later semester in the current year."
            
            if is_disabled:
                disabled_subjects_info[subject['id']] = reason

        # 7. Organize subjects by Year and Semester
        subjects_by_year = {}
        for subject in all_subjects:
            year = subject['year_level']
            semester = subject['semester']
            if year not in subjects_by_year:
                subjects_by_year[year] = {}
            if semester not in subjects_by_year[year]:
                subjects_by_year[year][semester] = []
            subjects_by_year[year][semester].append(subject)

        # 8. Fetch Max Units Allowed
        max_units_allowed = 26 
        if student.get('academic_year') and student.get('enrollment_semester'):
             cursor.execute("SELECT max_units FROM academic_terms WHERE year_name = %s AND semester = %s", 
                            (student['academic_year'], student['enrollment_semester']))
             term_data = cursor.fetchone()
             if term_data:
                 max_units_allowed = term_data['max_units']

        _, stats = _get_all_applications_and_stats()

        return render_template('admin_manage_grades.html', 
                               student=student, 
                               subjects_by_year=subjects_by_year, 
                               student_grades=student_grades,
                               stats=stats,
                               student_academic_status=student_academic_status,
                               disabled_subjects_info=disabled_subjects_info,
                               active_school_year=student.get('academic_year'),
                               max_units_allowed=max_units_allowed, 
                               active_page='enrolling')
    except Exception as e:
        flash(f"Error loading grades page: {e}", "danger")
        traceback.print_exc()
        return redirect(url_for('auth.admin_enrolling_applications'))
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()
        
@auth.route('/admin/application/<int:applicant_id>/update-level-section', methods=['POST'])
def admin_update_student_level_section(applicant_id):
    # 1. Check Permissions
    if session.get('user_role') not in ['admin', 'registrar']:
        return jsonify({"success": False, "message": "Unauthorized"}), 401

    # 2. Get Data
    new_program = request.form.get('program_choice') # <-- Capture Program
    new_year_level = request.form.get('year_level')
    new_section_id = request.form.get('section_id')

    if not new_year_level or not new_section_id or not new_program:
        return jsonify({"success": False, "message": "Program, Year Level and Section are required."}), 400

    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # 3. Update Database
        # We update program_choice, year level, section_id, and ensure is_section_permanent is True (since admin set it manually)
        cursor.execute("""
            UPDATE applicants 
            SET program_choice = %s,
                enrollment_year_level = %s, 
                section_id = %s, 
                is_section_permanent = TRUE,
                last_updated_at = NOW()
            WHERE applicant_id = %s
        """, (new_program, new_year_level, new_section_id, applicant_id))
        
        conn.commit()
        
        # Get Section Name for response message
        cursor.execute("SELECT section_name FROM sections WHERE id = %s", (new_section_id,))
        sec_res = cursor.fetchone()
        sec_name = sec_res[0] if sec_res else "Unknown"

        return jsonify({
            "success": True, 
            "message": f"Student placement updated: {new_program} / {new_year_level} / Section {sec_name}."
        })

    except Exception as e:
        print(f"Error updating level/section: {e}")
        return jsonify({"success": False, "message": "Server error."}), 500
    finally:
        if cursor: cursor.close()
        if conn: conn.close()



@auth.route('/admin/application/<int:applicant_id>/notes', methods=['POST', 'GET'])
def admin_save_application_notes(applicant_id):
    if session.get('user_role') not in ['admin', 'registrar']: return jsonify({"success": False, "message": "Unauthorized"}), 401
    
    conn = None; cursor = None
    try:
        conn = get_db_connection()
        if not conn: return jsonify({"success": False, "message": "Database connection error"}), 500
        
        if request.method == 'GET':
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT admin_notes FROM applicants WHERE applicant_id = %s", (applicant_id,))
            result = cursor.fetchone()
            if result:
                return jsonify({"success": True, "notes": result.get('admin_notes') or ""})
            return jsonify({"success": False, "message": "Application not found."}), 404

        if request.method == 'POST':
            data = request.get_json()
            if not isinstance(data, dict):
                return jsonify({"success": False, "message": "Invalid JSON format."}), 400
            notes = data.get('notes', '')

            # 1. Update the note in database
            cursor = conn.cursor()
            cursor.execute("UPDATE applicants SET admin_notes = %s, last_updated_at = NOW() WHERE applicant_id = %s", (notes, applicant_id))
            conn.commit()
            
            rows_affected = cursor.rowcount
            cursor.close() # Close this cursor before opening another

            # 2. Handle Email Notification (if enabled and note is not empty)
            email_sent_msg = ""
            if _is_email_trigger_enabled('email_trigger_notes') and notes.strip():
                cursor = conn.cursor(dictionary=True)
                cursor.execute("SELECT email_address, first_name, last_name, old_student_id, control_number FROM applicants WHERE applicant_id = %s", (applicant_id,))
                student = cursor.fetchone()
                cursor.close()
                
                if student and student['email_address']:
                    subject = "Update on Your Application: New Admin Note"
                    body = f"""
                    <p>Dear {student['first_name']} {student['last_name']},</p>
                    <p>An administrator has updated the notes on your application:</p>
                    <blockquote style="background-color: #f3f4f6; padding: 15px; border-left: 4px solid #4f46e5; margin: 10px 0;">
                        {notes}
                    </blockquote>
                    <p>Please log in to your student portal to view the full status of your application.</p>
                    """
                    if _send_email(subject, [student['email_address']], body):
                        email_sent_msg = " Email notification sent."
                    else:
                        email_sent_msg = " Email notification failed."

            # 3. Prepare response message (re-fetch ID if needed, or use passed ID)
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT old_student_id, control_number FROM applicants WHERE applicant_id = %s", (applicant_id,))
            app_ids = cursor.fetchone()
            display_id = app_ids['old_student_id'] if app_ids else (app_ids['control_number'] if app_ids else f"App #{applicant_id}")

            if rows_affected > 0: 
                return jsonify({"success": True, "message": f"Notes for {display_id} saved.{email_sent_msg}"})
            
            # Even if rowcount is 0 (same note saved), we return success
            return jsonify({"success": True, "message": f"Notes saved (no changes detected).{email_sent_msg}"})
            
    except Exception as e:
        print(f"Error handling notes for {applicant_id}: {e}"); traceback.print_exc()
        return jsonify({"success": False, "message": "Server error processing notes."}), 500
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()

@auth.route('/admin/grades/<int:student_user_id>/save', methods=['POST'])
def admin_save_grades(student_user_id):
    if session.get('user_role') not in ['admin', 'registrar']:
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    
    data = request.get_json()
    grades_to_save = data.get('grades')

    if not grades_to_save:
        return jsonify({"success": True, "message": "No grades to save."})

    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({"success": False, "message": "Database connection error."}), 500
        
        cursor = conn.cursor()
        
        # UPDATED SQL QUERY
        sql = """
            INSERT INTO student_grades (student_user_id, subject_id, grade, percentage, remarks, academic_year, semester, created_at, updated_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
            ON DUPLICATE KEY UPDATE 
                grade = VALUES(grade), 
                percentage = VALUES(percentage),
                remarks = VALUES(remarks),
                updated_at = NOW()
        """
        
        records_to_insert = []
        for grade_info in grades_to_save:
            grade_to_save = grade_info.get('grade') 
            
            # --- PERCENTAGE HANDLING ---
            percentage_to_save = grade_info.get('percentage')
            # Convert to None if empty string or None, otherwise ensure Int
            if percentage_to_save == '' or percentage_to_save is None:
                percentage_to_save = None
            else:
                try:
                    percentage_to_save = int(percentage_to_save)
                except:
                    percentage_to_save = None
            # ---------------------------

            remarks_to_save = grade_info.get('remarks')

            # Auto "Incomplete" Logic
            if grade_to_save is None and not remarks_to_save:
                remarks_to_save = 'Incomplete'
            if grade_to_save is not None and not remarks_to_save:
                remarks_to_save = 'Incomplete'
            
            db_grade_value = grade_to_save if grade_to_save is not None else 0.00
            
            # Ensure the order matches the SQL VALUES (%s...) above exactly
            records_to_insert.append((
                student_user_id,
                grade_info['subject_id'],
                db_grade_value,
                percentage_to_save, # <--- Passed here
                remarks_to_save,
                grade_info['academic_year'],
                grade_info['semester']
            ))

        cursor.executemany(sql, records_to_insert)
        conn.commit()
        
        return jsonify({"success": True, "message": f"{len(records_to_insert)} grade record(s) saved successfully."})

    except Exception as e:
        if conn: conn.rollback()
        print(f"Error saving grades for student {student_user_id}: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "message": "A server error occurred while saving grades."}), 500
    finally:
        if cursor: cursor.close()
        if conn: conn.close()


# ----------------- GRADE EXPORT/IMPORT ROUTES -----------------

@auth.route('/admin/reports/generate-grade-sheet', methods=['POST'])
def admin_generate_grade_sheet():
    if session.get('user_role') not in ['admin', 'registrar']:
        return redirect(url_for('auth.admin'))

    # 1. Get Form Data
    section_id = request.form.get('section_id')
    subject_id = request.form.get('subject_id')
    instructor_name = request.form.get('instructor_name', 'TBA') 
    academic_year = request.form.get('academic_year')
    semester = request.form.get('semester')

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        # 2. Fetch Metadata (Section Name, Subject Details)
        cursor.execute("SELECT section_name FROM sections WHERE id = %s", (section_id,))
        section_data = cursor.fetchone()
        section_name = section_data['section_name'] if section_data else "Unknown"

        cursor.execute("SELECT subject_code, subject_title, units FROM subjects WHERE id = %s", (subject_id,))
        subject_data = cursor.fetchone()
        
        # 3. Fetch Students and Grades (Including Percentage)
        query = """
            SELECT a.last_name, a.first_name, a.middle_name, sg.grade, sg.percentage, sg.remarks
            FROM applicants a
            JOIN student_users su ON a.student_user_id = su.id
            LEFT JOIN student_grades sg ON su.id = sg.student_user_id AND sg.subject_id = %s
            WHERE a.section_id = %s 
            AND a.application_status = 'Enrolled'
            ORDER BY a.sex ASC, a.last_name ASC
        """
        cursor.execute(query, (subject_id, section_id))
        students = cursor.fetchall()

        # 4. Setup Excel using OpenPyXL
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report on Grades"

        # --- STYLING CONSTANTS ---
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center')
        bold_font = Font(bold=True, name='Arial', size=10)
        regular_font = Font(name='Arial', size=10)

        # --- HEADER ---
        ws.merge_cells('A1:K1')
        ws['A1'] = "PADRE GARCIA POLYTECHNIC COLLEGE"
        ws['A1'].font = Font(bold=True, size=14, name='Arial')
        ws['A1'].alignment = center_align
        
        ws.merge_cells('A2:K2')
        ws['A2'] = "Brgy. Castillo, Padre Garcia, Batangas"
        ws['A2'].alignment = center_align

        ws.merge_cells('A3:K3')
        ws['A3'] = "REPORT ON GRADES"
        ws['A3'].font = Font(bold=True, size=12)
        ws['A3'].alignment = center_align

        ws.merge_cells('A4:K4')
        ws['A4'] = "OVERALL GRADE"
        ws['A4'].font = Font(bold=True, size=12)
        ws['A4'].alignment = center_align

        # --- METADATA ROW ---
        ws['A6'] = "Name :"
        ws.merge_cells('B6:E6')
        ws['B6'] = instructor_name.upper()
        ws['B6'].font = bold_font
        ws['B6'].border = Border(bottom=Side(style='thin'))
        
        ws['F6'] = semester.upper()
        ws['F6'].alignment = center_align
        ws['F6'].border = Border(bottom=Side(style='thin'))
        
        ws['G6'] = "Semester / Academic Year:"
        ws.merge_cells('H6:J6')
        ws['H6'] = academic_year
        ws['H6'].border = Border(bottom=Side(style='thin'))
        ws['H6'].alignment = center_align

        ws['A7'] = "Course/Code:"
        ws.merge_cells('B7:E7')
        ws['B7'] = f"{subject_data['subject_code']} - {subject_data['subject_title']}"
        ws['B7'].font = bold_font
        ws['B7'].border = Border(bottom=Side(style='thin'))
        
        ws['J7'] = "Units:"
        ws['K7'] = subject_data['units']
        ws['K7'].alignment = center_align
        ws['K7'].border = Border(bottom=Side(style='thin'))

        ws['A8'] = "Course/Sec:"
        ws.merge_cells('B8:E8')
        ws['B8'] = section_name
        ws['B8'].font = bold_font
        ws['B8'].border = Border(bottom=Side(style='thin'))

        # --- TABLE COLUMNS SETUP ---
        ws['A10'] = "No"
        ws.merge_cells('B10:C10')
        ws['B10'] = "Name of Student"
        ws['D10'] = "PERCENTAGE"
        ws['E10'] = "GRADE POINT"
        
        ws['G10'] = "No"
        ws.merge_cells('H10:I10')
        ws['H10'] = "Name of Student"
        ws['J10'] = "PERCENTAGE"
        ws['K10'] = "GRADE POINT"

        for col in ['A', 'B', 'D', 'E', 'G', 'H', 'J', 'K']:
            cell = ws[f'{col}10']
            cell.border = thin_border
            cell.font = Font(bold=True, size=8)
            cell.alignment = center_align

        # --- FILLING STUDENT DATA ---
        row_start = 11
        max_rows_per_col = 30 
        
        for i, student in enumerate(students):
            if i < max_rows_per_col:
                curr = row_start + i
                c_no, c_name, c_perc, c_grade = 'A', 'B', 'D', 'E'
                ws.merge_cells(f'B{curr}:C{curr}')
            else:
                curr = row_start + (i - max_rows_per_col)
                c_no, c_name, c_perc, c_grade = 'G', 'H', 'J', 'K'
                ws.merge_cells(f'H{curr}:I{curr}')

            full_name = f"{student['last_name']}, {student['first_name']}"
            if student['middle_name']:
                full_name += f" {student['middle_name'][0]}."
            
            grade = student['grade']
            percentage = student['percentage']
            
            disp_grade = f"{grade:.2f}" if grade is not None and grade > 0 else ""
            disp_perc = str(percentage) if percentage else ""

            is_red = False
            if student['remarks'] == 'Failed' or (grade and grade > 3.0):
                is_red = True
            
            font_style = Font(color="FF0000") if is_red else regular_font

            ws[f'{c_no}{curr}'] = i + 1
            ws[f'{c_name}{curr}'] = full_name.upper()
            ws[f'{c_perc}{curr}'] = disp_perc
            ws[f'{c_grade}{curr}'] = disp_grade

            for c in [c_no, c_name, c_perc, c_grade]:
                cell = ws[f'{c}{curr}']
                cell.border = thin_border
                cell.font = regular_font
                cell.alignment = center_align
            
            ws[f'{c_name}{curr}'].alignment = Alignment(horizontal='left', vertical='center')
            
            if is_red:
                ws[f'{c_perc}{curr}'].font = font_style
                ws[f'{c_grade}{curr}'].font = font_style

        # --- FOOTER SIGNATORIES ---
        footer_row = row_start + max_rows_per_col + 2
        ws[f'A{footer_row}'] = "Prepared by:"
        ws[f'G{footer_row}'] = f"Date Submitted: {datetime.date.today().strftime('%B %d, %Y')}"
        
        sig_row = footer_row + 3
        
        ws.merge_cells(f'A{sig_row}:C{sig_row}')
        ws[f'A{sig_row}'] = instructor_name.upper()
        ws[f'A{sig_row}'].font = Font(bold=True, underline='single')
        ws[f'A{sig_row}'].alignment = center_align
        
        ws.merge_cells(f'A{sig_row+1}:C{sig_row+1}')
        ws[f'A{sig_row+1}'] = "Instructor"
        ws[f'A{sig_row+1}'].alignment = center_align

        ws[f'G{footer_row}'] = "Recommending Approval:"
        ws.merge_cells(f'G{sig_row}:I{sig_row}')
        ws[f'G{sig_row}'] = "_______________________"
        ws[f'G{sig_row}'].alignment = center_align
        
        ws.merge_cells(f'G{sig_row+1}:I{sig_row+1}')
        ws[f'G{sig_row+1}'] = "College Dean"
        ws[f'G{sig_row+1}'].alignment = center_align

        # Grading System Legend
        legend_row = sig_row + 4
        ws.merge_cells(f'A{legend_row}:K{legend_row+5}')
        legend_cell = ws[f'A{legend_row}']
        legend_cell.value = "GRADING SYSTEM\n1.00 - 96-100 Excellent  |  1.25 - 94-95 Very Good\n1.50 - 91-93 Very Good  |  1.75 - 88-90 Good\n2.00 - 85-87 Good  |  2.25 - 82-84 Good\n2.50 - 79-81 Satisfactory  |  2.75 - 76-78 Satisfactory\n3.00 - 75 Passing  |  5.00 - Failed\nINC - Incomplete  |  UD - Unofficially Dropped"
        legend_cell.font = Font(size=8)
        legend_cell.alignment = center_align
        legend_cell.border = Border(top=Side(style='medium'), bottom=Side(style='medium'), left=Side(style='medium'), right=Side(style='medium'))

        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['H'].width = 25
        ws.column_dimensions['C'].width = 2
        ws.column_dimensions['I'].width = 2

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Grades_{section_name.replace(' ', '_')}_{subject_data['subject_code']}.xlsx"
        return send_file(output, as_attachment=True, download_name=filename, 
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        print(f"Error generating excel: {e}")
        traceback.print_exc()
        flash("An error occurred generating the report.", "danger")
        return redirect(url_for('auth.admin_grade_reports'))
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

@auth.route('/admin/grades/<int:student_user_id>/export')
def admin_export_grades(student_user_id):
    if session.get('user_role') not in ['admin', 'registrar', 'guidance', 'osa']:
        return redirect(url_for('auth.admin'))

    year_filter = request.args.get('year')
    semester_filter = request.args.get('semester')

    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # 1. Get Student Info
        cursor.execute("SELECT first_name, last_name FROM applicants WHERE student_user_id = %s", (student_user_id,))
        student = cursor.fetchone()
        student_name = f"{student['first_name']}_{student['last_name']}" if student else "Student"

        # 2. Get Program ID
        cursor.execute("""
            SELECT p.program_id 
            FROM student_users u
            JOIN applicants a ON u.id = a.student_user_id
            JOIN programs p ON a.program_choice = p.title
            WHERE u.id = %s
        """, (student_user_id,))
        program_data = cursor.fetchone()
        
        if not program_data:
            flash("Student program not found.", "danger")
            return redirect(request.referrer)

        # 3. Build Query - NOW INCLUDES PERCENTAGE
        base_query = """
            SELECT 
                s.subject_code AS 'Code',
                s.subject_title AS 'Subject Title',
                sg.percentage AS 'Percentage',
                sg.grade AS 'Grade'
            FROM subjects s
            LEFT JOIN student_grades sg ON s.id = sg.subject_id AND sg.student_user_id = %s
            WHERE s.program_id = %s
        """
        params = [student_user_id, program_data['program_id']]

        if year_filter and year_filter != 'all':
            base_query += " AND s.year_level = %s"
            params.append(year_filter)
        
        if semester_filter and semester_filter != 'all':
            base_query += " AND s.semester = %s"
            params.append(semester_filter)

        base_query += " ORDER BY s.year_level, s.semester, s.subject_code"

        cursor.execute(base_query, tuple(params))
        data = cursor.fetchall()

        # 4. Process Data
        for row in data:
            # Clean up Grade Point
            if row['Grade'] is None or row['Grade'] == 0.00:
                row['Grade'] = ""
            # Clean up Percentage
            if row['Percentage'] is None or row['Percentage'] == 0:
                row['Percentage'] = ""
        
        df = pd.DataFrame(data)

        # 5. Create Excel
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Grades')
            
            worksheet = writer.sheets['Grades']
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = length + 4

        output.seek(0)
        
        filter_suffix = ""
        if year_filter and year_filter != 'all': filter_suffix += f"_Y{year_filter}"
        if semester_filter and semester_filter != 'all': filter_suffix += f"_S{semester_filter}"

        filename = f"Grades_{student_name}{filter_suffix}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print(f"Export Error: {e}")
        traceback.print_exc()
        flash("Error exporting grades.", "danger")
        return redirect(request.referrer)
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

@auth.route('/admin/reports/import-grade-sheet', methods=['POST'])
def admin_import_grade_sheet():
    if session.get('user_role') not in ['admin', 'registrar']:
        return redirect(url_for('auth.admin'))

    section_id = request.form.get('section_id')
    subject_id = request.form.get('subject_id')
    academic_year = request.form.get('academic_year')
    semester = request.form.get('semester')
    file = request.files.get('file')

    if not all([section_id, subject_id, file]):
        flash("Missing required fields or file.", "danger")
        return redirect(url_for('auth.admin_grade_reports'))

    updated_count = 0
    errors = []

    try:
        # 1. LOAD WORKBOOK WITH data_only=True
        # This tells openpyxl to ignore formulas (e.g. =SUM(A1:A5)) and read the 
        # last calculated value (e.g. 15) stored by Excel.
        # Required for linked cells or formula-based grades.
        wb = load_workbook(file, data_only=True)
        ws = wb.active 

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Standard Layout: (Name Col, Percentage Col, Grade Col)
        # Left Side: Name=B(2), Perc=D(4), Grade=E(5)
        # Right Side: Name=H(8), Perc=J(10), Grade=K(11)
        layout_cols = [(2, 4, 5), (8, 10, 11)]
        
        # We start reading from row 11 based on the template structure
        for row_idx in range(11, 60): 
            for name_col, perc_col, grade_col in layout_cols:
                
                # --- A. READ NAME ---
                cell_name = ws.cell(row=row_idx, column=name_col).value
                if not cell_name or str(cell_name).strip() == "":
                    continue # Skip empty rows
                
                name_str = str(cell_name).strip()
                
                # Skip Headers/Garbage
                invalid_starts = ["NAME", "GRADE", "PERCENTAGE", "NO", ":", "."]
                if any(name_str.upper().startswith(x) for x in invalid_starts):
                    continue
                # Skip if name is just a number (row index)
                try:
                    float(name_str)
                    continue
                except ValueError:
                    pass

                # --- B. READ GRADE (With Flexible Column Logic) ---
                raw_grade = ws.cell(row=row_idx, column=grade_col).value
                raw_perc = ws.cell(row=row_idx, column=perc_col).value

                # FIX: Handle "Colon" layout (Name | : | Grade) shown in user screenshot
                # If we read a colon, assume the data is shifted one column to the right
                if str(raw_grade).strip() == ":":
                    raw_grade = ws.cell(row=row_idx, column=grade_col + 1).value
                
                if str(raw_perc).strip() == ":":
                    raw_perc = ws.cell(row=row_idx, column=perc_col + 1).value

                # --- C. VALIDATE EXCEL ERRORS ---
                # data_only=True returns strings like '#REF!' if formula failed
                excel_errors = ["#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A", "#NUM!", "#NULL!"]
                if str(raw_grade) in excel_errors:
                    errors.append(f"Row {row_idx}: Grade for '{name_str}' has an Excel error ({raw_grade}). Please open file in Excel and fix formulas.")
                    continue

                # --- D. PROCESS NUMERIC VALUES ---
                grade_val = 0.00
                is_valid_grade = False

                if raw_grade is not None and str(raw_grade).strip() not in ["", "-"]:
                    try:
                        grade_val = float(raw_grade)
                        is_valid_grade = True
                    except (ValueError, TypeError):
                        # If value exists but isn't a number (and wasn't a colon/header), flag it
                        # errors.append(f"Row {row_idx}: Invalid grade '{raw_grade}' for {name_str}")
                        pass # Or ignore/treat as incomplete

                percentage_val = None
                try:
                    if raw_perc is not None and str(raw_perc).strip() not in ["", "-", ":"]:
                        percentage_val = int(float(raw_perc))
                except:
                    percentage_val = None

                # Only process if we found a valid name format
                if ',' not in name_str:
                    continue

                # --- E. SAVE TO DATABASE ---
                parts = name_str.split(',')
                last_name = parts[0].strip()
                # Handle cases like "DELA CRUZ, JUAN" vs "DELA CRUZ, JUAN JR."
                first_name_part = parts[1].strip().split(' ')[0] 

                # Find student in THIS section
                cursor.execute("""
                    SELECT a.student_user_id 
                    FROM applicants a
                    WHERE a.section_id = %s 
                    AND a.last_name LIKE %s 
                    AND a.first_name LIKE %s
                    AND a.application_status = 'Enrolled'
                    LIMIT 1
                """, (section_id, f"%{last_name}%", f"%{first_name_part}%"))
                
                student = cursor.fetchone()

                if student:
                    remarks = 'Passed'
                    if grade_val > 3.0: remarks = 'Failed'
                    elif grade_val == 0.0 or not is_valid_grade: remarks = 'Incomplete'
                    elif grade_val == 5.0: remarks = 'Failed'

                    cursor.execute("""
                        INSERT INTO student_grades (student_user_id, subject_id, grade, percentage, remarks, academic_year, semester, created_at, updated_at)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
                        ON DUPLICATE KEY UPDATE 
                            grade = VALUES(grade), 
                            percentage = VALUES(percentage),
                            remarks = VALUES(remarks),
                            updated_at = NOW()
                    """, (student['student_user_id'], subject_id, grade_val, percentage_val, remarks, academic_year, semester))
                    updated_count += 1
                else:
                    # Optional: log students not found in DB
                    # errors.append(f"Student not found: {name_str}")
                    pass

        conn.commit()
        
        if updated_count > 0:
            flash(f"✅ Successfully imported/updated {updated_count} grades.", "success")
        else:
            flash("⚠️ No valid student grades were found to import.", "warning")

        if errors:
            # Show first 3 errors to avoid clutter
            error_html = "<br>".join(errors[:3])
            if len(errors) > 3: error_html += f"<br>...and {len(errors)-3} more issues."
            flash(f"Import completed with warnings:<br>{error_html}", "warning")

    except Exception as e:
        print(f"Import Error: {e}")
        import traceback
        traceback.print_exc()
        flash("An error occurred during import. Ensure the file is not corrupted.", "danger")
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

    return redirect(url_for('auth.admin_grade_reports'))

@auth.route('/admin/grades/<int:student_user_id>/import', methods=['POST'])
def admin_import_grades(student_user_id):
    if session.get('user_role') not in ['admin', 'registrar']:
        return jsonify({"success": False, "message": "Unauthorized"}), 401

    file = request.files.get('file')
    if not file or not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({"success": False, "message": "Invalid file. Please upload an Excel file."}), 400

    conn = None
    cursor = None
    try:
        df = pd.read_excel(file)
        df.columns = df.columns.str.strip()

        # Support flexible column names
        df.rename(columns={
            'Subject Code': 'Code', 
            'Subject Title': 'Subject Title',
            'Percentage': 'Percentage',
            'Grade': 'Grade'
        }, inplace=True)

        if 'Code' not in df.columns:
             return jsonify({"success": False, "message": "Invalid format. 'Code' column is missing."}), 400

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # 1. Get Program ID
        cursor.execute("""
            SELECT p.program_id 
            FROM student_users u
            JOIN applicants a ON u.id = a.student_user_id
            JOIN programs p ON a.program_choice = p.title
            WHERE u.id = %s
        """, (student_user_id,))
        program_data = cursor.fetchone()
        
        if not program_data:
            return jsonify({"success": False, "message": "Student program not found."}), 404
            
        program_id = program_data['program_id']

        # 2. Get Subject Map
        cursor.execute("SELECT id, subject_code, semester FROM subjects WHERE program_id = %s", (program_id,))
        subject_map = {row['subject_code'].strip().upper(): {'id': row['id'], 'sem_num': row['semester']} for row in cursor.fetchall()}

        # 3. Get Active Year
        active_term = _get_active_term()
        default_ay = active_term['year_name'] if active_term else '2025-2026'

        # UPDATED QUERY: Includes Percentage
        insert_query = """
            INSERT INTO student_grades (student_user_id, subject_id, grade, percentage, remarks, academic_year, semester, created_at, updated_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
            ON DUPLICATE KEY UPDATE 
                grade = VALUES(grade), 
                percentage = VALUES(percentage),
                remarks = VALUES(remarks),
                academic_year = VALUES(academic_year),
                semester = VALUES(semester),
                updated_at = NOW()
        """

        params_list = []

        for _, row in df.iterrows():
            code = str(row['Code']).strip().upper()
            
            if code in subject_map:
                subj_info = subject_map[code]
                subject_id = subj_info['id']
                
                # --- Parse Values ---
                raw_grade = row.get('Grade')
                raw_perc = row.get('Percentage')
                
                grade_to_save = 0.00
                perc_to_save = None
                remarks_to_save = ''

                # Percentage Logic
                if not pd.isna(raw_perc) and str(raw_perc).strip() != '':
                    try:
                        perc_to_save = int(float(raw_perc))
                    except:
                        perc_to_save = None

                # Grade Logic & Remarks
                if pd.isna(raw_grade) or str(raw_grade).strip() == '':
                    grade_to_save = 0.00
                    # Only clear remarks if both are empty
                    if perc_to_save is None:
                         remarks_to_save = '' 
                    else:
                         # If we have percentage but no grade, infer 'Incomplete' or calculate grade?
                         # For now, defaulting to Incomplete if grade is missing
                         remarks_to_save = 'Incomplete' 
                else:
                    try:
                        grade_val = float(raw_grade)
                        grade_to_save = grade_val
                        
                        if 1.0 <= grade_val <= 3.0: remarks_to_save = "Passed"
                        elif grade_val > 3.0: remarks_to_save = "Failed"
                        elif grade_val == 0.0: remarks_to_save = "Incomplete"
                        else: remarks_to_save = "Failed"

                    except ValueError:
                        grade_to_save = 0.00
                        remarks_to_save = str(raw_grade).strip()

                # Semester Context
                ay = default_ay
                sem_num = subj_info['sem_num']
                if sem_num == 1: sem_str = '1st Semester'
                elif sem_num == 2: sem_str = '2nd Semester'
                elif sem_num == 3: sem_str = 'Summer / Midyear'
                else: sem_str = '1st Semester'

                params_list.append((student_user_id, subject_id, grade_to_save, perc_to_save, remarks_to_save, ay, sem_str))

        if params_list:
            insert_cursor = conn.cursor()
            insert_cursor.executemany(insert_query, params_list)
            conn.commit()
            insert_cursor.close()

        return jsonify({"success": True, "message": "Grades imported successfully."})

    except Exception as e:
        print(f"Import Error: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "message": f"Server error: {str(e)}"}), 500
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

# ----------------- EXPORT APPLICATIONS LIST ROUTE -----------------
@auth.route('/admin/export-applications/<export_type>')
def admin_export_applications_list(export_type):
    if session.get('user_role') not in ['admin', 'registrar']:
        return redirect(url_for('auth.admin'))

    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # Base Query
        query = """
            SELECT 
                COALESCE(old_student_id, control_number, CONCAT('App #', applicant_id)) AS `Student ID`,
                CONCAT(first_name, ' ', last_name) AS `Applicant Name`,
                program_choice AS `Program`,
                email_address AS `Email`,
                mobile_number AS `Mobile`,
                application_status AS `Status`,
                exam_status AS `Exam Status`,
                DATE_FORMAT(submitted_at, '%Y-%m-%d') AS `Date Submitted`
            FROM applicants
        """
        
        params = []
        where_clause = ""

        # Filtering Logic based on export_type
        if export_type == 'pending':
            where_clause = " WHERE application_status = 'Pending'"
        elif export_type == 'approved':
            # Approved but not yet scheduled
            where_clause = " WHERE application_status = 'Approved' AND (permit_exam_date IS NULL OR permit_exam_date = '')"
        elif export_type == 'scheduled':
            where_clause = """ 
                WHERE (application_status = 'Scheduled' 
                OR (application_status = 'Approved' AND permit_exam_date IS NOT NULL))
                AND (exam_status IS NULL OR exam_status != 'Not Taken')
            """
        elif export_type == 'not_taken':
            where_clause = " WHERE exam_status = 'Not Taken'"
        elif export_type == 'rejected':
            where_clause = " WHERE application_status = 'Rejected'"
        elif export_type == 'passed':
            where_clause = " WHERE application_status = 'Passed'"
        elif export_type == 'failed':
            where_clause = " WHERE application_status = 'Failed'"
        elif export_type == 'eligible': 
            where_clause = " WHERE application_status = 'Eligible for Enrollment'"
        elif export_type == 'enrolling':
            where_clause = " WHERE application_status = 'Enrolling'"
        elif export_type == 'enrolled':
            where_clause = " WHERE application_status = 'Enrolled'"
        elif export_type == 'dropped':
            where_clause = " WHERE application_status = 'Dropped'"
        elif export_type == 'not_enrolled':
            where_clause = " WHERE application_status = 'Not Enrolled'"
        elif export_type == 'all':
            where_clause = "" # No filter

        final_query = query + where_clause + " ORDER BY submitted_at DESC"
        
        cursor.execute(final_query, params)
        data = cursor.fetchall()

        if not data:
            flash("No records found to export.", "info")
            return redirect(request.referrer)

        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Generate Excel
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Applications')
            # Simple column width adjustment
            worksheet = writer.sheets['Applications']
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2

        output.seek(0)
        filename = f"List_{export_type.title()}_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print(f"Export List Error: {e}")
        traceback.print_exc()
        flash("An error occurred during export.", "danger")
        return redirect(request.referrer)
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

# ----------------- ADMIN CONTENT MANAGEMENT -----------------
@auth.route('/admin/manage-content')
def admin_manage_content():
    if session.get('user_role') not in ['admin', 'registrar', 'secretary']: 
        return redirect(url_for('auth.admin'))
    _, stats = _get_all_applications_and_stats()
    return render_template('admin_manage_content.html', stats=stats, active_page='manage_content')

# --- Programs ---
@auth.route('/admin/manage-programs')
def admin_manage_programs():
    if session.get('user_role') not in ['admin', 'registrar','secretary']: return redirect(url_for('auth.admin'))
    _, stats = _get_all_applications_and_stats()
    programs = []
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor(dictionary=True)
            # ADDED: max_units to the SELECT query
            cursor.execute("SELECT program_id, title, description, hero_image_filename, max_units FROM programs ORDER BY title")
            programs = cursor.fetchall()
    except Exception as e:
        flash(f"Error fetching programs: {e}", "danger")
        traceback.print_exc()
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()
    
    return render_template('admin_manage_programs.html', programs=programs, stats=stats, active_page='manage_programs')

@auth.route('/admin/program/form', defaults={'program_id': None}, methods=['GET', 'POST'])
@auth.route('/admin/program/form/<program_id>', methods=['GET', 'POST'])
def admin_program_form(program_id):
    if session.get('user_role') not in ['admin', 'registrar', 'secretary']:  return redirect(url_for('auth.admin'))
    
    from .views import _get_program_details_from_db

    conn = None; cursor = None
    try:
        if request.method == 'POST':
            conn = get_db_connection()
            if not conn:
                flash("Database connection failed.", "danger")
                return redirect(url_for('auth.admin_manage_programs'))
            
            cursor = conn.cursor()
            
            # --- Get form data ---
            new_program_id = request.form.get('program_id').lower()
            title = request.form.get('title')
            description = request.form.get('description')
            
            # NEW: Get Max Units
            max_units = request.form.get('max_units')
            
            objectives = request.form.getlist('objectives[]')
            careers = request.form.getlist('careers[]')
            course_codes = request.form.getlist('course_codes[]')
            course_names = request.form.getlist('course_names[]')
            course_descs = request.form.getlist('course_descs[]')
            
            # ... (Image handling code remains the same) ...
            image_file = request.files.get('hero_image')
            new_image_filename = None
            if image_file and image_file.filename:
                new_image_filename, error = save_image_to_uploads(image_file, prefix='program')
                if error:
                    flash(f"Image upload error: {error}", "danger")
                    return redirect(request.url)

            if program_id: # --- UPDATE ---
                # ... (Old image deletion logic remains same) ...
                if new_image_filename:
                    cursor.execute("SELECT hero_image_filename FROM programs WHERE program_id = %s", (program_id,))
                    old_filename_row = cursor.fetchone()
                    if old_filename_row and old_filename_row[0]:
                        delete_image_from_uploads(old_filename_row[0])

                image_sql_part = ", hero_image_filename = %s" if new_image_filename else ""
                
                # MODIFIED: Update query to include max_units
                update_sql = f"UPDATE programs SET title = %s, description = %s, max_units = %s{image_sql_part} WHERE program_id = %s"
                update_params = [title, description, max_units]
                if new_image_filename: update_params.append(new_image_filename)
                update_params.append(program_id)
                
                cursor.execute(update_sql, tuple(update_params))

                # ... (Related tables delete logic remains same) ...
                cursor.execute("DELETE FROM program_objectives WHERE program_id = %s", (program_id,))
                cursor.execute("DELETE FROM program_careers WHERE program_id = %s", (program_id,))
                cursor.execute("DELETE FROM program_courses WHERE program_id = %s", (program_id,))
            
            else: # --- INSERT ---
                # MODIFIED: Insert query to include max_units
                cursor.execute(
                    "INSERT INTO programs (program_id, title, description, max_units, hero_image_filename, admission_link_endpoint) VALUES (%s, %s, %s, %s, %s, %s)",
                    (new_program_id, title, description, max_units, new_image_filename, 'views.existing_or_not')
                )
            
            target_id = program_id or new_program_id
            
            # ... (Re-insert related data logic remains same) ...
            if objectives:
                cursor.executemany("INSERT INTO program_objectives (program_id, objective_text) VALUES (%s, %s)", [(target_id, o) for o in objectives if o.strip()])
            if careers:
                cursor.executemany("INSERT INTO program_careers (program_id, career_text) VALUES (%s, %s)", [(target_id, c) for c in careers if c.strip()])
            if course_codes and len(course_codes) == len(course_names) == len(course_descs):
                courses_data = [
                    (target_id, code, name, desc) for code, name, desc in zip(course_codes, course_names, course_descs) if code.strip() and name.strip()
                ]
                cursor.executemany("INSERT INTO program_courses (program_id, course_code, course_name, course_description) VALUES (%s, %s, %s, %s)", courses_data)

            conn.commit()
            flash(f"Program '{title}' saved successfully!", "success")
            return redirect(url_for('auth.admin_manage_programs'))

        # --- GET Request ---
        # ... (GET logic remains the same) ...
        program_data = None
        if program_id:
            program_data = _get_program_details_from_db(program_id)
            if not program_data:
                flash(f"Program with ID '{program_id}' not found.", "danger")
                return redirect(url_for('auth.admin_manage_programs'))
        
        _, stats = _get_all_applications_and_stats()
        return render_template('admin_program_form.html', program=program_data, stats=stats, active_page='manage_programs')

    except Exception as e:
        # ... (Error handling remains same) ...
        flash(f"An error occurred: {e}", "danger")
        traceback.print_exc()
        if conn: conn.rollback()
        return redirect(url_for('auth.admin_manage_programs'))
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()

# --- HELPER: Save Generic File (PDF/Image) ---
def save_file_to_uploads(file_storage, prefix='doc'):
    """Saves a file (PDF/Image) to UPLOAD_FOLDER."""
    if not file_storage or not file_storage.filename:
        return None, "No file provided."

    allowed_extensions = {'.png', '.jpg', '.jpeg', '.webp', '.pdf'}
    file_ext = os.path.splitext(file_storage.filename)[1].lower()

    if file_ext not in allowed_extensions:
        return None, f"Invalid file type. Allowed: PDF, JPG, PNG."
    
    timestamp = int(time.time())
    secure_name = secure_filename(os.path.splitext(file_storage.filename)[0])
    new_filename = f"{prefix}_{secure_name}_{timestamp}{file_ext}"
    
    try:
        save_path = os.path.join(current_app.config['UPLOAD_FOLDER'], new_filename)
        file_storage.save(save_path)
        return new_filename, None
    except Exception as e:
        return None, str(e)

def save_file_to_uploads(file_storage, prefix='hero'):
    if not file_storage or not file_storage.filename:
        return None, "No file provided."

    # EXTENDED allowed list to include mp4
    allowed_extensions = {'.png', '.jpg', '.jpeg', '.webp', '.pdf', '.mp4', '.mov'}
    file_ext = os.path.splitext(file_storage.filename)[1].lower()

    if file_ext not in allowed_extensions:
        return None, f"Invalid file type. Allowed: Photos or MP4 Video."
    
    timestamp = int(time.time())
    secure_name = secure_filename(os.path.splitext(file_storage.filename)[0])
    new_filename = f"{prefix}_{secure_name}_{timestamp}{file_ext}"
    
    try:
        save_path = os.path.join(current_app.config['UPLOAD_FOLDER'], new_filename)
        file_storage.save(save_path)
        return new_filename, None
    except Exception as e:
        return None, str(e)

# --- ROUTE: Admin Manage Sidebar (Updated to fetch uploads & programs) ---
@auth.route('/admin/manage-sidebar-content', methods=['GET', 'POST'])
def admin_manage_sidebar_content():
    if session.get('user_role') not in ['admin', 'registrar', 'secretary']: return redirect(url_for('auth.admin'))
    
    conn = None; cursor = None
    try:
        conn = get_db_connection()
        
        if request.method == 'POST':
            # ... (Existing POST logic for text content remains here) ...
            cursor = conn.cursor()
            fields = [
                'sidebar_info_title', 'sidebar_info_1', 'sidebar_info_2', 'sidebar_info_3',
                'sidebar_contact_title', 'sidebar_phone_label', 
                'sidebar_email_label', 'sidebar_email_value', 
                'sidebar_hours_label', 'sidebar_hours_value',
                'status_msg_pending', 'status_msg_review', 'status_msg_approved', 
                'status_msg_scheduled', 
                'status_msg_rejected', 'status_msg_passed', 'status_msg_failed', 
                'status_msg_enrolling', 'status_msg_enrolled', 'status_msg_dropped', 
                'status_msg_not_enrolled',
                'requirements_header_text'
            ]
            for key in fields:
                value = request.form.get(key, '')
                cursor.execute("""
                    INSERT INTO page_content (content_key, content_value) VALUES (%s, %s)
                    ON DUPLICATE KEY UPDATE content_value = %s
                """, (key, value, value))
            conn.commit()
            flash("Content updated successfully.", "success")
            return redirect(url_for('auth.admin_manage_sidebar_content'))

        # GET Request
        cursor = conn.cursor(dictionary=True)
        
        # 1. Fetch Page Content
        cursor.execute("SELECT * FROM page_content")
        content = {row['content_key']: row['content_value'] for row in cursor.fetchall()}

        # 2. Fetch Requirements
        cursor.execute("SELECT * FROM status_requirements ORDER BY id ASC")
        requirements = defaultdict(list)
        for row in cursor.fetchall():
            requirements[row['status_key']].append(row)

        # 3. NEW: Fetch Programs (for the dropdown)
        cursor.execute("SELECT program_id, title FROM programs ORDER BY title")
        programs = cursor.fetchall()

        # 4. NEW: Fetch Status Uploads (Updated with COLLATE)
        cursor.execute("""
            SELECT u.*, p.title as program_title 
            FROM status_uploads u 
            LEFT JOIN programs p ON u.program_id COLLATE utf8mb4_0900_ai_ci = p.program_id COLLATE utf8mb4_0900_ai_ci
            ORDER BY u.uploaded_at DESC
        """)
        uploads_raw = cursor.fetchall()
        uploads = defaultdict(list)
        for row in uploads_raw:
            uploads[row['status_key']].append(row)

        _, stats = _get_all_applications_and_stats()
        
        return render_template('admin_manage_sidebar_content.html', 
                               content=content, 
                               requirements=requirements, 
                               programs=programs,  # Pass programs
                               uploads=uploads,    # Pass uploads
                               stats=stats, 
                               active_page='manage_sidebar_content')

    except Exception as e:
        flash(f"Error: {e}", "danger")
        traceback.print_exc()
        return redirect(url_for('auth.admin_manage_content'))
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

# --- NEW ROUTE: Upload Status File ---
@auth.route('/admin/status-file/upload', methods=['POST'])
def admin_upload_status_file():
    if session.get('user_role') not in ['admin', 'registrar']: return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    status_key = request.form.get('status_key')
    display_name = request.form.get('display_name')
    program_id = request.form.get('program_id') # Can be empty string for "All"
    if program_id == "": program_id = None 

    file = request.files.get('file')

    if not status_key or not display_name or not file:
        flash("Missing required fields.", "danger")
        return redirect(url_for('auth.admin_manage_sidebar_content'))

    filename, error = save_file_to_uploads(file, prefix='status_doc')
    if error:
        flash(f"Upload error: {error}", "danger")
        return redirect(url_for('auth.admin_manage_sidebar_content'))

    file_type = 'image' if filename.endswith(('.png', '.jpg', '.jpeg', '.webp')) else 'pdf'

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            INSERT INTO status_uploads (status_key, program_id, file_filename, display_name, file_type) 
            VALUES (%s, %s, %s, %s, %s)
        """, (status_key, program_id, filename, display_name, file_type))
        conn.commit()
        flash("File uploaded successfully.", "success")
    except Exception as e:
        print(f"DB Error: {e}")
        flash("Database error saving file info.", "danger")
    finally:
        cursor.close()
        conn.close()

    return redirect(url_for('auth.admin_manage_sidebar_content'))

# --- NEW ROUTE: Delete Status File ---
@auth.route('/admin/status-file/delete/<int:file_id>', methods=['POST'])
def admin_delete_status_file(file_id):
    if session.get('user_role') not in ['admin', 'registrar']: return redirect(url_for('auth.admin'))
    
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT file_filename FROM status_uploads WHERE id = %s", (file_id,))
        file_record = cursor.fetchone()
        
        if file_record:
            delete_image_from_uploads(file_record['file_filename']) # Helper deletes files
            cursor.execute("DELETE FROM status_uploads WHERE id = %s", (file_id,))
            conn.commit()
            flash("File deleted.", "success")
    except Exception as e:
        flash(f"Error: {e}", "danger")
    finally:
        cursor.close()
        conn.close()
        
    return redirect(url_for('auth.admin_manage_sidebar_content'))

# --- NEW ROUTE: Download Status File ---
@auth.route('/download/status-file/<filename>')
def download_status_file(filename):
    # Ideally add checks here if user is logged in or authorized to view this status
    return send_from_directory(current_app.config['UPLOAD_FOLDER'], filename, as_attachment=True)


# --- NEW: Add Requirement ---
@auth.route('/admin/requirement/add', methods=['POST'])
def admin_add_requirement():
    if session.get('user_role') not in ['admin', 'registrar']: return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    status_key = request.form.get('status_key')
    req_text = request.form.get('requirement_text')
    
    if not status_key or not req_text:
        flash("Missing data.", "danger")
        return redirect(url_for('auth.admin_manage_sidebar_content'))

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO status_requirements (status_key, requirement_text) VALUES (%s, %s)", (status_key, req_text))
        conn.commit()
        flash("Requirement added.", "success")
    except Exception as e:
        flash(f"Error: {e}", "danger")
    finally:
        cursor.close()
        conn.close()
        
    return redirect(url_for('auth.admin_manage_sidebar_content'))

# --- NEW: Delete Requirement ---
@auth.route('/admin/requirement/delete/<int:req_id>', methods=['POST'])
def admin_delete_requirement(req_id):
    if session.get('user_role') not in ['admin', 'registrar']: return redirect(url_for('auth.admin'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM status_requirements WHERE id = %s", (req_id,))
        conn.commit()
        flash("Requirement removed.", "success")
    except Exception as e:
        flash(f"Error: {e}", "danger")
    finally:
        cursor.close()
        conn.close()
        
    return redirect(url_for('auth.admin_manage_sidebar_content'))



@auth.route('/admin/program/delete/<program_id>', methods=['POST'])
def admin_delete_program(program_id):
    if session.get('user_role') not in ['admin', 'registrar', 'secretary']: return redirect(url_for('auth.admin'))
    conn = None; cursor = None
    try:
        conn = get_db_connection()
        if not conn:
            flash("Database connection failed.", "danger")
            return redirect(url_for('auth.admin_manage_programs'))
        
        cursor = conn.cursor(dictionary=True)
        # First, get the image filename to delete it from the filesystem
        cursor.execute("SELECT hero_image_filename FROM programs WHERE program_id = %s", (program_id,))
        program = cursor.fetchone()
        if program and program['hero_image_filename']:
            delete_image_from_uploads(program['hero_image_filename'])
        
        # Delete from all related tables. The DB should have cascading deletes, but doing it manually is safer.
        cursor.execute("DELETE FROM program_objectives WHERE program_id = %s", (program_id,))
        cursor.execute("DELETE FROM program_careers WHERE program_id = %s", (program_id,))
        cursor.execute("DELETE FROM program_courses WHERE program_id = %s", (program_id,))
        cursor.execute("DELETE FROM programs WHERE program_id = %s", (program_id,))
        
        conn.commit()
        flash(f"Program '{program_id.upper()}' and all its data have been deleted.", "success")
    except Exception as e:
        flash(f"Error deleting program: {e}", "danger")
        traceback.print_exc()
        if conn: conn.rollback()
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()
    
    return redirect(url_for('auth.admin_manage_programs'))

# --- Subjects (Curriculum) ---
@auth.route('/admin/program/<program_id>/subjects')
def admin_manage_subjects(program_id):
    if session.get('user_role') not in ['admin', 'registrar', 'secretary']: return redirect(url_for('auth.admin'))
    
    _, stats = _get_all_applications_and_stats()
    subjects = []
    program = None # Initialize program variable
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor(dictionary=True)
            # Get program details (Added total_units to query)
            cursor.execute("SELECT title, total_units FROM programs WHERE program_id = %s", (program_id,))
            program = cursor.fetchone()

            # Get all subjects
            cursor.execute("""
                SELECT s.*, p.subject_code as prerequisite_code 
                FROM subjects s
                LEFT JOIN subjects p ON s.prerequisite_subject_id = p.id
                WHERE s.program_id = %s 
                ORDER BY s.year_level, s.semester, s.subject_code
            """, (program_id,))
            subjects = cursor.fetchall()
    except Exception as e:
        flash(f"Error fetching subjects: {e}", "danger")
        traceback.print_exc()
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()
    
    return render_template('admin_program_subjects.html', 
                           subjects=subjects, 
                           stats=stats, 
                           program_id=program_id,
                           program=program, # Pass the full program object
                           active_page='manage_programs')

@auth.route('/admin/program/<program_id>/update-units', methods=['POST'])
def admin_update_program_units(program_id):
    if session.get('user_role') not in ['admin', 'registrar', 'secretary']: return redirect(url_for('auth.admin'))
    
    total_units = request.form.get('total_units')
    
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("UPDATE programs SET total_units = %s WHERE program_id = %s", (total_units, program_id))
        conn.commit()
        flash("Total units updated successfully.", "success")
    except Exception as e:
        flash(f"Error updating units: {e}", "danger")
    finally:
        if cursor: cursor.close()
        if conn: conn.close()
        
    return redirect(url_for('auth.admin_manage_subjects', program_id=program_id))

@auth.route('/edit-application/<int:applicant_id>', methods=['GET', 'POST'])
def edit_application_page(applicant_id):
    if 'student_id' not in session:
        flash("⚠️ Please log in to edit your application.", "warning")
        return redirect(url_for('auth.student_login_page'))

    student_user_id = session['student_id']
    conn = None
    cursor = None

    try:
        conn = get_db_connection()
        if not conn:
            flash("Database connection error.", "danger")
            return redirect(url_for('views.application_status_page'))
        
        cursor = conn.cursor(dictionary=True)

        # Security Check: Ensure the logged-in user owns this application
        cursor.execute("SELECT * FROM applicants WHERE applicant_id = %s AND student_user_id = %s", (applicant_id, student_user_id))
        application = cursor.fetchone()

        if not application:
            flash("Application not found or you do not have permission to edit it.", "danger")
            return redirect(url_for('views.application_status_page'))
        
        # Check if application is in an editable state
        if application['application_status'] not in ['Pending', 'In Review']:
            flash(f"This application cannot be edited because its status is '{application['application_status']}'.", "warning")
            return redirect(url_for('views.application_status_page'))

        if request.method == 'POST':
            # --- Handle Form Submission (UPDATE logic) ---
            update_clauses = []
            update_values = []
            
            # Process text fields
            field_list = [
                'program_choice', 'last_name', 'first_name', 'middle_name', 'date_of_birth', 'place_of_birth',
                'sex', 'civil_status', 'religion', 'citizenship', 'mobile_number', 'email_address',
                'permanent_address_street_barangay', 'permanent_address_city_municipality',
                'permanent_address_province', 'permanent_address_postal_code',
                'cultural_minority_group', 'physical_disability',
                'father_name', 'father_occupation', 'father_company_address', 'father_contact_number',
                'mother_maiden_name', 'mother_occupation', 'mother_company_address', 'mother_contact_number',
                'guardian_name', 'guardian_occupation', 'guardian_company_address', 'guardian_contact_number',
                'senior_high_school', 'senior_high_school_address', 'senior_high_school_track_strand',
                'senior_high_school_year_from', 'senior_high_school_year_to',
                'tertiary_school', 'tertiary_school_address', 'tertiary_course',
                'tertiary_year_from', 'tertiary_year_to'
            ]
            for field in field_list:
                val = request.form.get(field)
                update_clauses.append(f"{field} = %s")
                update_values.append(val.strip() if val and isinstance(val, str) else val)

            # Process file fields (only if a new file is uploaded)
            file_fields = {
                'photo': 'photo',
                'shs_diploma_file_input': 'shs_diploma',
                'shs_card_file_input': 'shs_card',
                'birth_certificate_file_input': 'birth_certificate'
            }
            for form_key, db_prefix in file_fields.items():
                file_storage = request.files.get(form_key)
                if file_storage and file_storage.filename:
                    data, fname, mtype, err = process_uploaded_file(file_storage, f"New {db_prefix} file")
                    if err:
                        flash(f"⚠️ {err}", "danger")
                        return redirect(url_for('auth.edit_application_page', applicant_id=applicant_id))
                    
                    if db_prefix == 'photo':
                        # Special handling for the 'photo' field which has columns 'photo' and 'photo_mimetype'
                        update_clauses.append("photo = %s")
                        update_values.append(data)
                        update_clauses.append("photo_mimetype = %s")
                        update_values.append(mtype)
                    else:
                        # Standard handling for other document fields with _file, _filename, _mimetype pattern
                        update_clauses.append(f"{db_prefix}_file = %s")
                        update_values.append(data)
                        update_clauses.append(f"{db_prefix}_filename = %s")
                        update_values.append(fname)
                        update_clauses.append(f"{db_prefix}_mimetype = %s")
                        update_values.append(mtype)

            # Add last_updated timestamp and build final query
            update_clauses.append("last_updated_at = %s")
            update_values.append(datetime.datetime.now())
            
            update_query = f"UPDATE applicants SET {', '.join(update_clauses)} WHERE applicant_id = %s"
            update_values.append(applicant_id)

            update_cursor = conn.cursor()
            update_cursor.execute(update_query, tuple(update_values))
            conn.commit()
            update_cursor.close()

            flash("✅ Your application has been updated successfully!", "success")
            return redirect(url_for('views.application_status_page'))

        # --- Handle GET request (Show the form) ---
        all_programs = _get_program_list()
        return render_template('edit_application.html', 
                               application=application, 
                               programs=all_programs, 
                               student_logged_in=True)

    except Exception as e:
        flash(f"An error occurred: {e}", "danger")
        traceback.print_exc()
        return redirect(url_for('views.application_status_page'))
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()

# --- Manage About Page Content ---
# --- auth.py ---

@auth.route('/admin/manage-about-content', methods=['GET', 'POST'])
def admin_manage_about_content():
    if session.get('user_role') not in ['admin', 'registrar', 'secretary']: 
        return redirect(url_for('auth.admin'))
    
    conn = None; cursor = None
    try:
        conn = get_db_connection()
        if request.method == 'POST':
            cursor = conn.cursor()

            # 1. Update Text Fields
            text_fields = [
                'community_title', 'community_text', 'president_name', 'president_role', 'president_desc',
                'about_identity_title', 'about_slogan', 'about_vision_text', 'about_mission_text', 'about_core_values',
                # Hero Fields
                'about_hero_title', 'about_hero_subtitle', 'about_hero_type' 
            ]
            
            for key in text_fields:
                value = request.form.get(key, '')
                cursor.execute("""
                    INSERT INTO page_content (content_key, content_value) VALUES (%s, %s)
                    ON DUPLICATE KEY UPDATE content_value = %s
                """, (key, value, value))

            # 2. Handle Media Upload (Image or Video)
            file = request.files.get('about_hero_bg_file')
            if file and file.filename != '':
                # Use save_file_to_uploads (your generic helper) to allow .mp4
                new_filename, error = save_file_to_uploads(file, prefix='hero')
                if not error:
                    cursor.execute("""
                        INSERT INTO page_content (content_key, content_value) VALUES (%s, %s)
                        ON DUPLICATE KEY UPDATE content_value = %s
                    """, ('about_hero_bg_file', new_filename, new_filename))

            # Handle other images (president, community, etc)
            other_images = [('community_image', 'community'), ('president_image', 'president'), ('about_identity_image', 'identity')]
            for form_key, prefix in other_images:
                f = request.files.get(form_key)
                if f and f.filename != '':
                    fname, err = save_image_to_uploads(f, prefix=prefix)
                    if not err:
                        cursor.execute("INSERT INTO page_content (content_key, content_value) VALUES (%s, %s) ON DUPLICATE KEY UPDATE content_value = %s", (form_key, fname, fname))

            conn.commit()
            flash("✅ About page updated successfully!", "success")
            return redirect(url_for('auth.admin_manage_about_content'))

        # GET Logic
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT content_key, content_value FROM page_content")
        content = {row['content_key']: row['content_value'] for row in cursor.fetchall()}
        _, stats = _get_all_applications_and_stats()
        return render_template('admin_manage_about_content.html', content=content, stats=stats, active_page='manage_about_content')

    except Exception as e:
        flash(f"Error: {e}", "danger")
        return redirect(url_for('auth.admin_manage_content'))
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

# --- Manage Contact & General Content ---

@auth.route('/admin/manage-contact-content', methods=['GET', 'POST'])
def admin_manage_contact_content():
    if session.get('user_role') not in ['admin', 'registrar', 'secretary']: 
        return redirect(url_for('auth.admin'))
    
    conn = None; cursor = None
    try:
        conn = get_db_connection()
        
        if request.method == 'POST':
            cursor = conn.cursor()

            # 1. Handle Page Content Text Fields
            text_fields = [
                'contact_email', 'contact_phone', 
                'contact_address', 'contact_hours', 'contact_map_url',
                'footer_title', 'footer_subtitle', 'footer_connect_text', 'footer_facebook_link',
                'consent_title', 'consent_text', 'oath_title', 'oath_text',
                'sidebar_info_title', 'sidebar_info_1', 'sidebar_info_2', 'sidebar_info_3',
                'sidebar_contact_title', 'sidebar_phone_label', 'sidebar_email_label', 'sidebar_email_value', 
                'sidebar_hours_label', 'sidebar_hours_value',
                'requirements_header_text',
                'status_msg_pending', 'status_msg_review', 'status_msg_approved', 'status_msg_scheduled', 
                'status_msg_rejected', 'status_msg_passed', 'status_msg_failed', 'status_msg_enrolling', 
                'status_msg_enrolled', 'status_msg_dropped', 'status_msg_not_enrolled'
            ]
            
            for key in text_fields:
                value = request.form.get(key, '')
                cursor.execute("""
                    INSERT INTO page_content (content_key, content_value) VALUES (%s, %s)
                    ON DUPLICATE KEY UPDATE content_value = %s
                """, (key, value, value))

            # 2. Handle Toggles (Email Triggers & Finance)
            # Checkboxes return 'true' if checked, nothing if unchecked. We map this to 'true'/'false' strings.
            toggle_fields = ['email_trigger_student', 'email_trigger_admin', 'email_trigger_notes', 'email_trigger_announcements', 'issuance_of_receipt']
            for tf in toggle_fields:
                val = 'true' if tf in request.form else 'false'
                cursor.execute("""
                    INSERT INTO system_settings (setting_key, setting_value) VALUES (%s, %s)
                    ON DUPLICATE KEY UPDATE setting_value = %s
                """, (tf, val, val))

            # 3. Handle Footer Logo Upload
            file = request.files.get('footer_logo_image')
            if file and file.filename:
                new_filename, error = save_image_to_uploads(file, prefix='logo')
                if not error:
                    cursor.execute("""
                        INSERT INTO page_content (content_key, content_value) VALUES ('footer_logo_image', %s)
                        ON DUPLICATE KEY UPDATE content_value = %s
                    """, (new_filename, new_filename))

            conn.commit()
            flash("Website settings updated successfully.", "success")
            return redirect(url_for('auth.admin_manage_contact_content'))

        # --- GET Request ---
        cursor = conn.cursor(dictionary=True)
        
        # 1. Fetch Page Content
        cursor.execute("SELECT * FROM page_content")
        rows = cursor.fetchall()
        content = {row['content_key']: row['content_value'] for row in rows}
        
        # 2. Fetch System Settings (Admissions Toggles, Email Triggers, Finance)
        cursor.execute("SELECT setting_key, setting_value FROM system_settings")
        settings_rows = cursor.fetchall()
        for row in settings_rows:
            content[row['setting_key']] = row['setting_value']

        # Set defaults if keys don't exist yet
        content.setdefault('admission_open_new', 'true')
        content.setdefault('admission_open_old', 'true')
        content.setdefault('email_trigger_student', 'true')
        content.setdefault('email_trigger_admin', 'true')
        content.setdefault('email_trigger_notes', 'true')
        content.setdefault('email_trigger_announcements', 'true')
        content.setdefault('issuance_of_receipt', 'true')

        # 3. Fetch Social Media Links & Requirements & Uploads (Existing logic)
        cursor.execute("SELECT * FROM social_media_links")
        social_links = cursor.fetchall()

        cursor.execute("SELECT * FROM status_requirements ORDER BY id ASC")
        requirements = defaultdict(list)
        for row in cursor.fetchall():
            requirements[row['status_key']].append(row)

        cursor.execute("SELECT program_id, title FROM programs ORDER BY title")
        programs = cursor.fetchall()

        cursor.execute("""
            SELECT u.*, p.title as program_title 
            FROM status_uploads u 
            LEFT JOIN programs p ON u.program_id COLLATE utf8mb4_0900_ai_ci = p.program_id COLLATE utf8mb4_0900_ai_ci
            ORDER BY u.uploaded_at DESC
        """)
        uploads_raw = cursor.fetchall()
        uploads = defaultdict(list)
        for row in uploads_raw:
            uploads[row['status_key']].append(row)

        _, stats = _get_all_applications_and_stats()
        
        return render_template('admin_manage_contact_content.html', 
                               content=content, 
                               stats=stats, 
                               social_links=social_links,
                               requirements=requirements,
                               programs=programs,
                               uploads=uploads,
                               active_page='manage_contact_content')

    except Exception as e:
        flash(f"Error: {e}", "danger")
        traceback.print_exc()
        return redirect(url_for('auth.admin_manage_content'))
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

# --- OPEN FILE: auth.py ---

@auth.route('/admin/toggle-registration', methods=['POST'])
def admin_toggle_registration():
    if session.get('user_role') not in ['admin', 'registrar', 'secretary']: 
        return redirect(url_for('auth.admin'))
    
    target_setting = request.form.get('setting_key')
    
    # UPDATED: Added 'enable_account_creation' to valid keys
    if target_setting not in ['admission_open_new', 'admission_open_old', 'enable_account_creation']:
        flash("Invalid setting key.", "danger")
        return redirect(url_for('auth.admin_manage_contact_content'))

    conn = None; cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("SELECT setting_value FROM system_settings WHERE setting_key = %s", (target_setting,))
        row = cursor.fetchone()
        
        current_status = row['setting_value'] == 'true' if row else True
        new_status = 'false' if current_status else 'true'
        
        cursor.execute("""
            INSERT INTO system_settings (setting_key, setting_value) VALUES (%s, %s)
            ON DUPLICATE KEY UPDATE setting_value = %s
        """, (target_setting, new_status, new_status))
        
        conn.commit()
        
        # Determine readable name for flash message
        names = {
            'admission_open_new': "New Student Admission",
            'admission_open_old': "Old Student Admission",
            'enable_account_creation': "Global Account Creation"
        }
        
        status_text = "OPEN/VISIBLE" if new_status == 'true' else "CLOSED/HIDDEN"
        flash(f"{names.get(target_setting)} is now {status_text}.", "success" if new_status == 'true' else "warning")
        
    except Exception as e:
        print(f"Error toggling registration: {e}")
        flash("Error updating status.", "danger")
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

    return redirect(url_for('auth.admin_manage_contact_content'))
        
# --- Hero Slider Management ---
@auth.route('/admin/manage-slider')
def admin_manage_slider():
    if session.get('user_role') not in ['admin', 'registrar', 'secretary']: return redirect(url_for('auth.admin'))
    _, stats = _get_all_applications_and_stats()
    slides = []
    conn = None; cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM hero_slides ORDER BY sort_order ASC, id DESC")
        slides = cursor.fetchall()
    except Exception as e:
        flash(f"Error fetching slides: {e}", "danger")
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()
    return render_template('admin_manage_slider.html', slides=slides, stats=stats, active_page='manage_slider')

@auth.route('/admin/slider/form', defaults={'slide_id': None}, methods=['GET', 'POST'])
@auth.route('/admin/slider/form/<int:slide_id>', methods=['GET', 'POST'])
def admin_slider_form(slide_id):
    if session.get('user_role') not in ['admin', 'registrar', 'secretary']: return redirect(url_for('auth.admin'))
    conn = None; cursor = None
    try:
        conn = get_db_connection()
        if request.method == 'POST':
            intro_text = request.form.get('intro_text')
            title_text = request.form.get('title_text')
            highlight_text = request.form.get('highlight_text')
            button_text = request.form.get('button_text')
            button_link = request.form.get('button_link')
            sort_order = request.form.get('sort_order', 0)
            is_active = 'is_active' in request.form
            image_file = request.files.get('image')

            cursor = conn.cursor(dictionary=True)
            new_image_filename = None
            if image_file and image_file.filename:
                new_image_filename, error = save_image_to_uploads(image_file, prefix='slide')
                if error:
                    flash(f"Image upload error: {error}", "danger")
                    return redirect(request.url)

            if slide_id: # Update
                if new_image_filename:
                    cursor.execute("SELECT image_filename FROM hero_slides WHERE id = %s", (slide_id,))
                    old_file = cursor.fetchone()
                    if old_file and old_file['image_filename']:
                        delete_image_from_uploads(old_file['image_filename'])
                
                img_sql = ", image_filename = %s" if new_image_filename else ""
                params = [intro_text, title_text, highlight_text, button_text, button_link, sort_order, is_active]
                if new_image_filename: params.append(new_image_filename)
                params.append(slide_id)
                
                cursor.execute(f"""
                    UPDATE hero_slides 
                    SET intro_text=%s, title_text=%s, highlight_text=%s, button_text=%s, button_link=%s, sort_order=%s, is_active=%s{img_sql} 
                    WHERE id=%s
                """, tuple(params))
            else: # Insert
                cursor.execute("""
                    INSERT INTO hero_slides (intro_text, title_text, highlight_text, button_text, button_link, sort_order, is_active, image_filename) 
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """, (intro_text, title_text, highlight_text, button_text, button_link, sort_order, is_active, new_image_filename))
            
            conn.commit()
            flash("Slide saved successfully!", "success")
            return redirect(url_for('auth.admin_manage_slider'))

        # GET Request
        slide = None
        if slide_id:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT * FROM hero_slides WHERE id = %s", (slide_id,))
            slide = cursor.fetchone()
            if not slide: flash("Slide not found.", "danger"); return redirect(url_for('auth.admin_manage_slider'))
        
        _, stats = _get_all_applications_and_stats()
        return render_template('admin_slider_form.html', slide=slide, stats=stats, active_page='manage_slider')

    except Exception as e:
        flash(f"An error occurred: {e}", "danger")
        traceback.print_exc()
        return redirect(url_for('auth.admin_manage_slider'))
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()

@auth.route('/admin/slider/delete/<int:slide_id>', methods=['POST'])
def admin_delete_slide(slide_id):
    if session.get('user_role') not in ['admin', 'registrar', 'secretary']: return redirect(url_for('auth.admin'))
    conn = None; cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT image_filename FROM hero_slides WHERE id = %s", (slide_id,))
        slide = cursor.fetchone()
        if slide and slide['image_filename']:
            delete_image_from_uploads(slide['image_filename'])
        
        cursor.execute("DELETE FROM hero_slides WHERE id = %s", (slide_id,))
        conn.commit()
        flash("Slide deleted.", "success")
    except Exception as e:
        flash(f"Error deleting slide: {e}", "danger")
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()
    return redirect(url_for('auth.admin_manage_slider'))

# --- Admission Steps Management ---
@auth.route('/admin/manage-admission')
def admin_manage_admission():
    if session.get('user_role') not in ['admin', 'registrar', 'secretary']: return redirect(url_for('auth.admin'))
    _, stats = _get_all_applications_and_stats()
    steps = []
    conn = None; cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM admission_steps ORDER BY step_number ASC")
        steps = cursor.fetchall()
    finally:
        if cursor: cursor.close()
        if conn: conn.close()
    return render_template('admin_manage_admission.html', steps=steps, stats=stats, active_page='manage_admission')

@auth.route('/admin/admission/form', defaults={'step_id': None}, methods=['GET', 'POST'])
@auth.route('/admin/admission/form/<int:step_id>', methods=['GET', 'POST'])
def admin_admission_form(step_id):
    if session.get('user_role') not in ['admin', 'registrar', 'secretary']: return redirect(url_for('auth.admin'))
    conn = None; cursor = None
    try:
        conn = get_db_connection()
        if request.method == 'POST':
            step_number = request.form.get('step_number')
            title = request.form.get('title')
            description = request.form.get('description')
            
            cursor = conn.cursor()
            if step_id:
                cursor.execute("UPDATE admission_steps SET step_number=%s, title=%s, description=%s WHERE id=%s", (step_number, title, description, step_id))
            else:
                cursor.execute("INSERT INTO admission_steps (step_number, title, description) VALUES (%s, %s, %s)", (step_number, title, description))
            conn.commit()
            flash("Admission step saved.", "success")
            return redirect(url_for('auth.admin_manage_admission'))

        step = None
        if step_id:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT * FROM admission_steps WHERE id=%s", (step_id,))
            step = cursor.fetchone()
        
        _, stats = _get_all_applications_and_stats()
        return render_template('admin_admission_form.html', step=step, stats=stats, active_page='manage_admission')
    except Exception as e:
        flash(f"Error: {e}", "danger")
        return redirect(url_for('auth.admin_manage_admission'))
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

@auth.route('/admin/admission/delete/<int:step_id>', methods=['POST'])
def admin_delete_admission(step_id):
    if session.get('user_role') not in ['admin', 'registrar', 'secretary']: return redirect(url_for('auth.admin'))
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM admission_steps WHERE id=%s", (step_id,))
    conn.commit()
    cursor.close()
    conn.close()
    flash("Step deleted.", "success")
    return redirect(url_for('auth.admin_manage_admission'))

# --- FAQ Management ---
@auth.route('/admin/manage-faqs')
def admin_manage_faqs():
    if session.get('user_role') not in ['admin', 'registrar', 'secretary']: return redirect(url_for('auth.admin'))
    _, stats = _get_all_applications_and_stats()
    faqs = []
    conn = None; cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM faqs ORDER BY sort_order ASC")
        faqs = cursor.fetchall()
    finally:
        if cursor: cursor.close()
        if conn: conn.close()
    return render_template('admin_manage_faqs.html', faqs=faqs, stats=stats, active_page='manage_faqs')


@auth.route('/admin/social-media/add', methods=['POST'])
def admin_add_social_media():
    if session.get('user_role') not in ['admin', 'registrar']: 
        return redirect(url_for('auth.admin'))
    
    platform = request.form.get('platform')
    url = request.form.get('url')
    
    if not platform or not url:
        flash("Platform and URL are required.", "danger")
        return redirect(url_for('auth.admin_manage_contact_content'))

    # Map Platform to RemixIcon classes (used in your index/base footers)
    icons = {
        'Facebook': 'ri-facebook-box-fill',
        'Twitter': 'ri-twitter-x-fill',
        'Instagram': 'ri-instagram-line',
        'YouTube': 'ri-youtube-fill',
        'LinkedIn': 'ri-linkedin-box-fill',
        'TikTok': 'ri-tiktok-fill',
        'Website': 'ri-global-line'
    }
    icon_class = icons.get(platform, 'ri-link')

    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("INSERT INTO social_media_links (platform_name, url, icon_class) VALUES (%s, %s, %s)", 
                       (platform, url, icon_class))
        conn.commit()
        flash(f"{platform} link added successfully.", "success")
    except Exception as e:
        flash(f"Error adding link: {e}", "danger")
    finally:
        if cursor: cursor.close()
        if conn: conn.close()
        
    return redirect(url_for('auth.admin_manage_contact_content'))

@auth.route('/admin/social-media/delete/<int:link_id>', methods=['POST'])
def admin_delete_social_media(link_id):
    if session.get('user_role') not in ['admin', 'registrar']: return redirect(url_for('auth.admin'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM social_media_links WHERE id = %s", (link_id,))
        conn.commit()
        flash("Link deleted.", "success")
    except Exception as e:
        flash(f"Error deleting link: {e}", "danger")
    finally:
        cursor.close()
        conn.close()
        
    return redirect(url_for('auth.admin_manage_contact_content'))

@auth.route('/admin/faq/form', defaults={'faq_id': None}, methods=['GET', 'POST'])
@auth.route('/admin/faq/form/<int:faq_id>', methods=['GET', 'POST'])
def admin_faq_form(faq_id):
    if session.get('user_role') not in ['admin', 'registrar', 'secretary']: return redirect(url_for('auth.admin'))
    conn = None; cursor = None
    try:
        conn = get_db_connection()
        if request.method == 'POST':
            question = request.form.get('question')
            answer = request.form.get('answer')
            sort_order = request.form.get('sort_order', 0)
            
            cursor = conn.cursor()
            if faq_id:
                cursor.execute("UPDATE faqs SET question=%s, answer=%s, sort_order=%s WHERE id=%s", (question, answer, sort_order, faq_id))
            else:
                cursor.execute("INSERT INTO faqs (question, answer, sort_order) VALUES (%s, %s, %s)", (question, answer, sort_order))
            conn.commit()
            flash("FAQ saved.", "success")
            return redirect(url_for('auth.admin_manage_faqs'))

        faq = None
        if faq_id:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT * FROM faqs WHERE id=%s", (faq_id,))
            faq = cursor.fetchone()
        
        _, stats = _get_all_applications_and_stats()
        return render_template('admin_faq_form.html', faq=faq, stats=stats, active_page='manage_faqs')
    except Exception as e:
        flash(f"Error: {e}", "danger")
        return redirect(url_for('auth.admin_manage_faqs'))
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

@auth.route('/admin/faq/delete/<int:faq_id>', methods=['POST'])
def admin_delete_faq(faq_id):
    if session.get('user_role') not in ['admin', 'registrar', 'secretary']: return redirect(url_for('auth.admin'))
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM faqs WHERE id=%s", (faq_id,))
    conn.commit()
    cursor.close()
    conn.close()
    flash("FAQ deleted.", "success")
    return redirect(url_for('auth.admin_manage_faqs'))

# NEW: Endpoint for instant photo update
@auth.route('/edit-application/<int:applicant_id>/update-photo', methods=['POST'])
def update_application_photo(applicant_id):
    if 'student_id' not in session:
        return jsonify({"success": False, "message": "Authentication required."}), 401

    student_user_id = session['student_id']
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({"success": False, "message": "Database connection error."}), 500
        
        cursor = conn.cursor(dictionary=True)
        # Security check: Ensure the user owns this application
        cursor.execute("SELECT student_user_id FROM applicants WHERE applicant_id = %s", (applicant_id,))
        app = cursor.fetchone()

        if not app or app['student_user_id'] != student_user_id:
            return jsonify({"success": False, "message": "Unauthorized access."}), 403
        
        photo_file = request.files.get('photo')
        if not photo_file or not photo_file.filename:
            return jsonify({"success": False, "message": "No photo file provided."}), 400

        photo_data, _, photo_mimetype, error = process_uploaded_file(photo_file, "2x2 Photo", max_size_mb=5)
        if error:
            return jsonify({"success": False, "message": error}), 400

        # Update the database with the new photo
        update_cursor = conn.cursor()
        update_cursor.execute("""
            UPDATE applicants 
            SET photo = %s, photo_mimetype = %s, last_updated_at = NOW() 
            WHERE applicant_id = %s
        """, (photo_data, photo_mimetype, applicant_id))
        conn.commit()
        update_cursor.close()

        return jsonify({"success": True, "message": "Photo updated successfully."})

    except Exception as e:
        if conn: conn.rollback()
        print(f"Error updating photo for applicant {applicant_id}: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "message": "A server error occurred."}), 500
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()


@auth.route('/admin/program/<program_id>/subject/add', methods=['GET', 'POST'])
@auth.route('/admin/subject/<int:subject_id>/edit', methods=['GET', 'POST'])
def admin_subject_form(program_id=None, subject_id=None):
    if session.get('user_role') not in ['admin', 'registrar']: return redirect(url_for('auth.admin'))
    
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        if not conn:
            flash("Database connection failed.", "danger")
            return redirect(url_for('auth.admin_manage_programs'))

        if request.method == 'POST':
            # On POST, we get the program_id from a hidden form field
            target_program_id = request.form.get('program_id')
            subject_code = request.form.get('subject_code')
            subject_title = request.form.get('subject_title')
            units = request.form.get('units')
            year_level = request.form.get('year_level')
            semester = request.form.get('semester')
            prerequisite_id = request.form.get('prerequisite_subject_id')
            if not prerequisite_id or prerequisite_id == 'None':
                prerequisite_id = None

            cursor = conn.cursor()
            if subject_id: # Update
                cursor.execute("""
                    UPDATE subjects 
                    SET subject_code=%s, subject_title=%s, units=%s, year_level=%s, semester=%s, prerequisite_subject_id=%s, program_id=%s
                    WHERE id=%s
                """, (subject_code, subject_title, units, year_level, semester, prerequisite_id, target_program_id, subject_id))
            else: # Insert
                cursor.execute("""
                    INSERT INTO subjects (program_id, subject_code, subject_title, units, year_level, semester, prerequisite_subject_id) 
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (target_program_id, subject_code, subject_title, units, year_level, semester, prerequisite_id))
            
            conn.commit()
            flash(f"Subject '{subject_code}' saved successfully!", "success")
            return redirect(url_for('auth.admin_manage_subjects', program_id=target_program_id))

        # GET Request
        subject = None
        program_title = ""
        potential_prerequisites = []
        target_program_id_for_form = program_id # Use program_id from URL if adding

        cursor = conn.cursor(dictionary=True)
        if subject_id:
            cursor.execute("SELECT * FROM subjects WHERE id = %s", (subject_id,))
            subject = cursor.fetchone()
            if not subject:
                flash("Subject not found.", "danger")
                return redirect(url_for('auth.admin_manage_programs'))
            # When editing, the program_id comes from the subject data itself
            target_program_id_for_form = subject['program_id']

        # Get program title for the form header
        cursor.execute("SELECT title FROM programs WHERE program_id = %s", (target_program_id_for_form,))
        program = cursor.fetchone()
        if program:
            program_title = program['title']
        else:
             flash(f"Program with ID '{target_program_id_for_form}' not found.", "danger")
             return redirect(url_for('auth.admin_manage_programs'))

        # Fetch potential prerequisites from the same program
        cursor.execute("SELECT id, subject_code, subject_title FROM subjects WHERE program_id = %s ORDER BY subject_code", (target_program_id_for_form,))
        potential_prerequisites = cursor.fetchall()

        _, stats = _get_all_applications_and_stats()
        return render_template('admin_subject_form.html', 
                               subject=subject, 
                               stats=stats,
                               program_id=target_program_id_for_form,
                               program_title=program_title,
                               potential_prerequisites=potential_prerequisites,
                               active_page='manage_programs')

    except Exception as e:
        flash(f"An error occurred: {e}", "danger")
        traceback.print_exc()
        if conn: conn.rollback()
        return redirect(url_for('auth.admin_manage_programs'))
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()

@auth.route('/admin/subject/delete/<int:subject_id>', methods=['POST'])
def admin_delete_subject(subject_id):
    if session.get('user_role') not in ['admin', 'registrar']: return redirect(url_for('auth.admin'))
    conn = None; cursor = None
    program_id_to_redirect = None
    try:
        conn = get_db_connection()
        if not conn:
            flash("Database connection failed.", "danger")
            return redirect(url_for('auth.admin_manage_programs'))
        
        cursor = conn.cursor(dictionary=True)
        # First, get the program_id so we know where to redirect back to
        cursor.execute("SELECT program_id FROM subjects WHERE id = %s", (subject_id,))
        subject = cursor.fetchone()
        if subject:
            program_id_to_redirect = subject['program_id']

        # Now, delete the subject
        cursor.execute("DELETE FROM subjects WHERE id = %s", (subject_id,))
        conn.commit()
        flash("Subject deleted successfully.", "success")
    except Exception as e:
        flash(f"Error deleting subject: {e}", "danger")
        traceback.print_exc()
        if conn: conn.rollback()
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()
    
    if program_id_to_redirect:
        return redirect(url_for('auth.admin_manage_subjects', program_id=program_id_to_redirect))
    return redirect(url_for('auth.admin_manage_programs'))


# --- News ---
@auth.route('/admin/manage-news')
def admin_manage_news():
    if session.get('user_role') not in ['admin', 'registrar','secretary']: return redirect(url_for('auth.admin'))
    _, stats = _get_all_applications_and_stats()
    articles = []
    conn = None; cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM news_articles ORDER BY publish_date DESC, id DESC")
        articles = cursor.fetchall()
    except Exception as e:
        flash(f"Error fetching news articles: {e}", "danger")
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()
    return render_template('admin_manage_news.html', articles=articles, stats=stats, active_page='manage_news')

@auth.route('/admin/news/form', defaults={'article_id': None}, methods=['GET', 'POST'])
@auth.route('/admin/news/form/<int:article_id>', methods=['GET', 'POST'])
def admin_news_form(article_id):
    if session.get('user_role') not in ['admin', 'registrar', 'secretary']: return redirect(url_for('auth.admin'))
    conn = None; cursor = None
    try:
        conn = get_db_connection()
        if request.method == 'POST':
            title = request.form.get('title')
            publish_date_str = request.form.get('publish_date')
            publish_date = datetime.datetime.strptime(publish_date_str, '%Y-%m-%d').date() if publish_date_str else None
            description = request.form.get('description')
            read_more_link = request.form.get('read_more_link')
            image_file = request.files.get('image')

            cursor = conn.cursor(dictionary=True)
            new_image_filename = None
            if image_file and image_file.filename:
                new_image_filename, error = save_image_to_uploads(image_file, prefix='news')
                if error:
                    flash(f"Image upload error: {error}", "danger")
                    return redirect(request.url)

            if article_id: # Update
                if new_image_filename:
                    cursor.execute("SELECT image_filename FROM news_articles WHERE id = %s", (article_id,))
                    old_file = cursor.fetchone()
                    if old_file and old_file['image_filename']:
                        delete_image_from_uploads(old_file['image_filename'])
                
                img_sql = ", image_filename = %s" if new_image_filename else ""
                params = [title, publish_date, description, read_more_link]
                if new_image_filename: params.append(new_image_filename)
                params.append(article_id)
                cursor.execute(f"UPDATE news_articles SET title=%s, publish_date=%s, description=%s, read_more_link=%s{img_sql} WHERE id=%s", tuple(params))
            else: # Insert
                cursor.execute("INSERT INTO news_articles (title, publish_date, description, read_more_link, image_filename) VALUES (%s, %s, %s, %s, %s)",
                               (title, publish_date, description, read_more_link, new_image_filename))
            conn.commit()
            flash("News article saved!", "success")
            return redirect(url_for('auth.admin_manage_news'))

        # GET Request
        article = None
        if article_id:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT * FROM news_articles WHERE id = %s", (article_id,))
            article = cursor.fetchone()
            if not article: flash("Article not found.", "danger"); return redirect(url_for('auth.admin_manage_news'))
        
        _, stats = _get_all_applications_and_stats()
        return render_template('admin_news_form.html', article=article, stats=stats, active_page='manage_news')
    except Exception as e:
        flash(f"An error occurred: {e}", "danger")
        traceback.print_exc()
        return redirect(url_for('auth.admin_manage_news'))
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()

@auth.route('/admin/news/delete/<int:article_id>', methods=['POST'])
def admin_delete_news(article_id):
    if session.get('user_role') not in ['admin', 'registrar', 'secretary']: return redirect(url_for('auth.admin'))
    conn = None; cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT image_filename FROM news_articles WHERE id = %s", (article_id,))
        article = cursor.fetchone()
        if article and article['image_filename']:
            delete_image_from_uploads(article['image_filename'])
        cursor.execute("DELETE FROM news_articles WHERE id = %s", (article_id,))
        conn.commit()
        flash("News article deleted.", "success")
    except Exception as e:
        flash(f"Error deleting article: {e}", "danger")
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()
    return redirect(url_for('auth.admin_manage_news'))

# --- Faculty ---
@auth.route('/admin/manage-faculty')
def admin_manage_faculty():
    if session.get('user_role') not in ['admin', 'registrar', 'secretary']: return redirect(url_for('auth.admin'))
    _, stats = _get_all_applications_and_stats()
    faculty = []
    conn = None; cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM faculty ORDER BY category, id")
        faculty = cursor.fetchall()
    except Exception as e:
        flash(f"Error fetching faculty: {e}", "danger")
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()
    return render_template('admin_manage_faculty.html', faculty=faculty, stats=stats, active_page='manage_faculty')

@auth.route('/admin/faculty/form', defaults={'member_id': None}, methods=['GET', 'POST'])
@auth.route('/admin/faculty/form/<int:member_id>', methods=['GET', 'POST'])
def admin_faculty_form(member_id):
    if session.get('user_role') not in ['admin', 'registrar', 'secretary']: return redirect(url_for('auth.admin'))
    conn = None; cursor = None
    try:
        conn = get_db_connection()
        if request.method == 'POST':
            name = request.form.get('name')
            role = request.form.get('role')
            category = request.form.get('category')
            image_file = request.files.get('image')

            cursor = conn.cursor(dictionary=True)
            new_image_filename = None
            if image_file and image_file.filename:
                new_image_filename, error = save_image_to_uploads(image_file, prefix='faculty')
                if error:
                    flash(f"Image upload error: {error}", "danger")
                    return redirect(request.url)

            if member_id: # Update
                if new_image_filename:
                    cursor.execute("SELECT image_filename FROM faculty WHERE id = %s", (member_id,))
                    old_file = cursor.fetchone()
                    if old_file and old_file['image_filename']:
                        delete_image_from_uploads(old_file['image_filename'])
                
                img_sql = ", image_filename = %s" if new_image_filename else ""
                params = [name, role, category]
                if new_image_filename: params.append(new_image_filename)
                params.append(member_id)
                cursor.execute(f"UPDATE faculty SET name=%s, role=%s, category=%s{img_sql} WHERE id=%s", tuple(params))
            else: # Insert
                cursor.execute("INSERT INTO faculty (name, role, category, image_filename) VALUES (%s, %s, %s, %s)",
                               (name, role, category, new_image_filename))
            conn.commit()
            flash("Faculty/Staff member saved!", "success")
            return redirect(url_for('auth.admin_manage_faculty'))

        # GET Request
        member = None
        if member_id:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT * FROM faculty WHERE id = %s", (member_id,))
            member = cursor.fetchone()
            if not member: flash("Member not found.", "danger"); return redirect(url_for('auth.admin_manage_faculty'))
        
        _, stats = _get_all_applications_and_stats()
        return render_template('admin_faculty_form.html', member=member, stats=stats, active_page='manage_faculty')
    except Exception as e:
        flash(f"An error occurred: {e}", "danger")
        traceback.print_exc()
        return redirect(url_for('auth.admin_manage_faculty'))
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()

@auth.route('/admin/faculty/delete/<int:member_id>', methods=['POST'])
def admin_delete_faculty(member_id):
    if session.get('user_role') not in ['admin', 'registrar', 'secretary']: return redirect(url_for('auth.admin'))
    conn = None; cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT image_filename FROM faculty WHERE id = %s", (member_id,))
        member = cursor.fetchone()
        if member and member['image_filename']:
            delete_image_from_uploads(member['image_filename'])
        cursor.execute("DELETE FROM faculty WHERE id = %s", (member_id,))
        conn.commit()
        flash("Faculty/Staff member deleted.", "success")
    except Exception as e:
        flash(f"Error deleting member: {e}", "danger")
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()
    return redirect(url_for('auth.admin_manage_faculty'))

# --- Announcements ---
@auth.route('/admin/manage-announcements')
def admin_manage_announcements():
    if session.get('user_role') not in ['admin', 'registrar', 'secretary']: return redirect(url_for('auth.admin'))
    _, stats = _get_all_applications_and_stats()
    announcements = []
    conn = None; cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM announcements ORDER BY created_at DESC")
        announcements = cursor.fetchall()
    except Exception as e:
        flash(f"Error fetching announcements: {e}", "danger")
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()
    return render_template('admin_manage_announcements.html', announcements=announcements, stats=stats, active_page='manage_announcements')

@auth.route('/admin/announcement/form', defaults={'announcement_id': None}, methods=['GET', 'POST'])
@auth.route('/admin/announcement/form/<int:announcement_id>', methods=['GET', 'POST'])
def admin_announcement_form(announcement_id):
    if session.get('user_role') not in ['admin', 'registrar', 'secretary']: return redirect(url_for('auth.admin'))
    conn = None; cursor = None
    try:
        conn = get_db_connection()
        if not conn:
            flash("Database connection failed.", "danger")
            return redirect(url_for('auth.admin_manage_announcements'))

        if request.method == 'POST':
            title = request.form.get('title')
            content = request.form.get('content')
            is_active = 'is_active' in request.form
            image_file = request.files.get('image')

            cursor = conn.cursor(dictionary=True)
            new_image_filename = None
            if image_file and image_file.filename:
                new_image_filename, error = save_image_to_uploads(image_file, prefix='announcement')
                if error:
                    flash(f"Image upload error: {error}", "danger")
                    return redirect(request.url)

            if announcement_id: # Update
                if new_image_filename:
                    cursor.execute("SELECT image_filename FROM announcements WHERE id = %s", (announcement_id,))
                    old_file = cursor.fetchone()
                    if old_file and old_file['image_filename']:
                        delete_image_from_uploads(old_file['image_filename'])
                
                img_sql = ", image_filename = %s" if new_image_filename else ""
                params = [title, content, is_active]
                if new_image_filename: params.append(new_image_filename)
                params.append(announcement_id)
                
                update_cursor = conn.cursor()
                update_cursor.execute(f"UPDATE announcements SET title=%s, content=%s, is_active=%s, updated_at=NOW(), created_at=NOW(){img_sql} WHERE id=%s", tuple(params))
                update_cursor.close()

            else: # Insert
                insert_cursor = conn.cursor()
                insert_cursor.execute("INSERT INTO announcements (title, content, is_active, image_filename) VALUES (%s, %s, %s, %s)",
                               (title, content, is_active, new_image_filename))
                insert_cursor.close()
            
            conn.commit()

            # --- NEW: Email Blast Logic ---
            # Check if Announcement is ACTIVE and if Email Blast is ENABLED
            if is_active and _is_email_trigger_enabled('email_trigger_announcements'):
                # Fetch ALL distinct emails from applicants
                cursor = conn.cursor(dictionary=True)
                cursor.execute("SELECT DISTINCT email_address FROM applicants WHERE email_address IS NOT NULL AND email_address != ''")
                email_rows = cursor.fetchall()
                recipients = [row['email_address'] for row in email_rows]
                cursor.close()
                
                if recipients:
                    sender_name = current_app.config.get('MAIL_SENDER_NAME', 'PGPC')
                    sender_email = current_app.config.get('MAIL_USERNAME')
                    
                    email_subject = f"New Announcement: {title}"
                    email_body = f"""
                    <h2>{title}</h2>
                    <p>{content}</p>
                    <hr>
                    <p>Visit our website for more details.</p>
                    """
                    
                    # Use BCC for mass email privacy
                    msg = Message(subject=email_subject, sender=(sender_name, sender_email), bcc=recipients, html=email_body)
                    
                    try:
                        current_app.extensions.get('mail').send(msg)
                        flash("Announcement saved and email blast sent to all students.", "success")
                        return redirect(url_for('auth.admin_manage_announcements'))
                    except Exception as e:
                        print(f"Email Blast Error: {e}")
                        flash("Announcement saved, but failed to send email blast.", "warning")
            # ------------------------------

            flash("Announcement saved successfully!", "success")
            return redirect(url_for('auth.admin_manage_announcements'))

        # GET Request
        announcement = None
        if announcement_id:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT * FROM announcements WHERE id = %s", (announcement_id,))
            announcement = cursor.fetchone()
            if not announcement:
                flash("Announcement not found.", "danger")
                return redirect(url_for('auth.admin_manage_announcements'))
        
        _, stats = _get_all_applications_and_stats()
        return render_template('admin_announcement_form.html', announcement=announcement, stats=stats, active_page='manage_announcements')

    except Exception as e:
        flash(f"An error occurred: {e}", "danger")
        traceback.print_exc()
        if conn: conn.rollback()
        return redirect(url_for('auth.admin_manage_announcements'))
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()

@auth.route('/admin/announcement/delete/<int:announcement_id>', methods=['POST'])
def admin_delete_announcement(announcement_id):
    if session.get('user_role') not in ['admin', 'registrar', 'secretary']: return redirect(url_for('auth.admin'))
    conn = None; cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        # First, delete the associated image file from the server
        cursor.execute("SELECT image_filename FROM announcements WHERE id = %s", (announcement_id,))
        announcement = cursor.fetchone()
        if announcement and announcement['image_filename']:
            delete_image_from_uploads(announcement['image_filename'])
        
        # Then, delete the record from the database
        cursor.execute("DELETE FROM announcements WHERE id = %s", (announcement_id,))
        conn.commit()
        flash("Announcement deleted.", "success")
    except Exception as e:
        flash(f"Error deleting announcement: {e}", "danger")
        if conn: conn.rollback()
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()
    return redirect(url_for('auth.admin_manage_announcements'))

@auth.route('/admin/announcement/toggle/<int:announcement_id>', methods=['POST'])
def admin_toggle_announcement(announcement_id):
    if session.get('user_role') not in ['admin', 'registrar', 'secretary']:
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    
    conn = None; cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        # Toggle the is_active field
        cursor.execute("UPDATE announcements SET is_active = NOT is_active, updated_at=NOW() WHERE id = %s", (announcement_id,))
        conn.commit()
        if cursor.rowcount > 0:
            cursor.execute("SELECT is_active FROM announcements WHERE id = %s", (announcement_id,))
            new_status = cursor.fetchone()[0]
            return jsonify({"success": True, "message": "Status updated", "is_active": bool(new_status)})
        else:
            return jsonify({"success": False, "message": "Announcement not found"}), 404
    except Exception as e:
        print(f"Error toggling announcement {announcement_id}: {e}")
        return jsonify({"success": False, "message": "Server error"}), 500
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()


# --- NEW ROUTE FOR REAL-TIME EMAIL CHECK ---
@auth.route('/admin/check-email', methods=['POST'])
def admin_check_email_exists():
    if session.get('user_role') not in ['admin', 'registrar']:
        return jsonify({"error": "Unauthorized"}), 401
    
    email = request.json.get('email')
    if not email:
        return jsonify({"exists": False}) # No email to check

    conn = None; cursor = None
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({"error": "Database connection failed"}), 500
        
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT applicant_id FROM applicants 
            WHERE email_address = %s 
            AND application_status IN ('Pending', 'In Review', 'Approved', 'Scheduled', 'Passed', 'Enrolling')
        """, (email,))
        
        if cursor.fetchone():
            return jsonify({"exists": True})
        else:
            return jsonify({"exists": False})

    except Exception as e:
        print(f"Error checking email existence: {e}")
        return jsonify({"error": "Server error during email check"}), 500
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()

@auth.route('/admin/create_application_by_admin', methods=['POST'])
def admin_add_application_by_admin_action():
    if session.get('user_role') not in ['admin', 'registrar']: return jsonify({"success": False, "message": "Unauthorized"}), 401
    if request.method != 'POST': return jsonify({"success": False, "message": "Invalid request method."}), 405

    conn = None
    MAX_PHOTO_SIZE_MB, MAX_DOC_SIZE_MB = 5, 5

    try:
        active_term = _get_active_term()
        if not active_term:
            return jsonify({"success": False, "message": "Cannot add application: Admissions are currently closed (no active school year)."}), 400

        conn = get_db_connection()
        if not conn: return jsonify({"success": False, "message": "Database connection error."}), 500
        
        conn.start_transaction()
        cursor = conn.cursor(dictionary=True)

        student_type = request.form.get('student_type', 'new')
        old_student_id = request.form.get('old_student_id')

        if student_type == 'old':
            if not old_student_id or not old_student_id.strip():
                return jsonify({"success": False, "message": "Old Student ID is required for existing students."}), 400
            
            # Server-side check for duplicate Old Student ID
            cursor.execute("SELECT id FROM student_users WHERE old_student_id = %s", (old_student_id,))
            if cursor.fetchone():
                return jsonify({"success": False, "message": f"Old Student ID '{old_student_id}' is already registered to an account."}), 409

        field_list = [
            'program_choice', 'last_name', 'first_name', 'middle_name', 'date_of_birth', 'place_of_birth',
            'sex', 'civil_status', 'religion', 'citizenship', 'mobile_number', 'email_address',
            'permanent_address_street_barangay', 'permanent_address_city_municipality',
            'permanent_address_province', 'permanent_address_postal_code',
            'cultural_minority_group', 'physical_disability',
            'date_of_application', 'academic_year', 'average_family_income',
            'father_name', 'father_occupation', 'father_company_address', 'father_contact_number',
            'mother_maiden_name', 'mother_occupation', 'mother_company_address', 'mother_contact_number',
            'guardian_name', 'guardian_occupation', 'guardian_company_address', 'guardian_contact_number',
            'senior_high_school', 'senior_high_school_address', 'senior_high_school_track_strand',
            'senior_high_school_year_from', 'senior_high_school_year_to',
            'tertiary_school', 'tertiary_school_address', 'tertiary_course',
            'tertiary_year_from', 'tertiary_year_to', 'agreements', 'final_submission_date'
        ]
        # MODIFIED: Guardian is required, Father/Mother are not.
        req_text_fields = [f for f in field_list if f not in [
            'middle_name', 
            'father_name', 'father_occupation', 'father_company_address', 'father_contact_number',
            'mother_maiden_name', 'mother_occupation', 'mother_company_address', 'mother_contact_number',
            'tertiary_school', 'tertiary_school_address', 'tertiary_course', 
            'tertiary_year_from', 'tertiary_year_to'
        ]]
        req_date_fields = ['date_of_birth', 'date_of_application', 'final_submission_date']
        
        form_data = {}
        for field in field_list:
            val = request.form.get(field)
            if field in req_text_fields and (not val or not val.strip()): return jsonify({"success": False, "message": f"{field.replace('_',' ').title()} is required."}), 400
            if field in req_date_fields and not val: return jsonify({"success": False, "message": f"{field.replace('_',' ').title()} date is required."}), 400
            if field == 'agreements' and val != 'Yes': return jsonify({"success": False, "message": "Admin must confirm agreement."}), 400
            form_data[field] = val.strip() if val and isinstance(val, str) else val
        
        # MODIFIED: Check for duplicate full name
        first_name_check = form_data.get('first_name', '').strip().lower()
        last_name_check = form_data.get('last_name', '').strip().lower()
        middle_name_check = form_data.get('middle_name', '').strip().lower()

        cursor.execute("""
            SELECT applicant_id FROM applicants 
            WHERE LOWER(TRIM(first_name)) = %s 
              AND LOWER(TRIM(last_name)) = %s 
              AND LOWER(TRIM(COALESCE(middle_name, ''))) = %s
        """, (first_name_check, last_name_check, middle_name_check))

        if cursor.fetchone():
            conn.rollback() 
            return jsonify({"success": False, "message": "An applicant with this full name already exists in the database."}), 409
        
        applicant_email = form_data.get('email_address')
        if not applicant_email: return jsonify({"success": False, "message": "Email is required."}), 400

        cursor.execute("""
            SELECT applicant_id FROM applicants 
            WHERE email_address = %s 
            AND application_status IN ('Pending', 'In Review', 'Approved', 'Scheduled', 'Passed', 'Enrolling')
        """, (applicant_email,))
        if cursor.fetchone():
            return jsonify({"success": False, "message": f"An active application already exists for the email '{applicant_email}'."}), 409
        
        student_user_id, temp_pass, acc_msg = None, None, ""
        cursor.execute("SELECT id FROM student_users WHERE email = %s", (applicant_email,))
        existing_user = cursor.fetchone()
        
        if existing_user:
            student_user_id = existing_user['id']
            cursor.execute("SELECT applicant_id FROM applicants WHERE student_user_id = %s AND application_status IN ('Pending', 'In Review', 'Approved', 'Scheduled', 'Passed', 'Enrolling')", (student_user_id,))
            if cursor.fetchone():
                return jsonify({"success": False, "message": f"This email is already linked to an account with an active application."}), 409
            acc_msg = f"Linked to existing student account (ID: {student_user_id})."
        else:
            temp_pass = secrets.token_urlsafe(10)
            hashed_pass = generate_password_hash(temp_pass, method='pbkdf2:sha256')
            try:
                cursor.execute("INSERT INTO student_users (email, password, is_verified, created_at, updated_at, student_type, old_student_id) VALUES (%s, %s, TRUE, NOW(), NOW(), %s, %s)", 
                                         (applicant_email, hashed_pass, student_type, old_student_id if student_type == 'old' else None))
                student_user_id = cursor.lastrowid
                if not student_user_id: 
                    raise Exception("Failed to get last row ID for new student user.")
                acc_msg = f"New student account created (ID: {student_user_id})."
            except mysql.connector.Error as user_err:
                raise Exception(f"DB error creating student account: {user_err}") from user_err

        # MODIFIED: Generate 'A' ID for new students immediately
        new_control_number = None
        if student_type == 'new':
             new_control_number = _generate_admission_id(cursor) # Generates A00001
             if not new_control_number:
                raise Exception("Failed to generate admission ID.")

        files_data = {}
        file_fields = {
            'photo': ('2x2 Photo', MAX_PHOTO_SIZE_MB), 
            'shs_diploma_file_input': ('SHS Diploma', MAX_DOC_SIZE_MB),
            'shs_card_file_input': ('SHS Card', MAX_DOC_SIZE_MB),
            'birth_certificate_file_input': ('Birth Certificate', MAX_DOC_SIZE_MB)
        }
        for form_key, (desc, max_size) in file_fields.items():
            file_store = request.files.get(form_key)
            if file_store and file_store.filename:
                data, fname, mtype, err = process_uploaded_file(file_store, desc, max_size)
                if err: return jsonify({"success": False, "message": f"{desc} Error: {err}"}), 400
                files_data[form_key] = {'data': data, 'filename': fname, 'mimetype': mtype}

        db_cols = [
            'student_user_id', 'program_choice', 'last_name', 'first_name', 'middle_name', 'date_of_birth', 'place_of_birth',
            'sex', 'civil_status', 'religion', 'citizenship', 'mobile_number', 'email_address',
            'permanent_address_street_barangay', 'permanent_address_city_municipality',
            'permanent_address_province', 'permanent_address_postal_code',
            'cultural_minority_group', 'physical_disability', 'control_number', 'permit_control_no',
            'date_of_application', 'academic_year', 'average_family_income',
            'father_name', 'father_occupation', 'father_company_address', 'father_contact_number',
            'mother_maiden_name', 'mother_occupation', 'mother_company_address', 'mother_contact_number',
            'guardian_name', 'guardian_occupation', 'guardian_company_address', 'guardian_contact_number',
            'senior_high_school', 'senior_high_school_address', 'senior_high_school_track_strand',
            'senior_high_school_year_from', 'senior_high_school_year_to',
            'tertiary_school', 'tertiary_school_address', 'tertiary_course',
            'tertiary_year_from', 'tertiary_year_to', 'agreements', 'final_submission_date', 'old_student_id',
            'photo', 'photo_mimetype', 'shs_diploma_file', 'shs_diploma_filename', 'shs_diploma_mimetype',
            'shs_card_file', 'shs_card_filename', 'shs_card_mimetype',
            'birth_certificate_file', 'birth_certificate_filename', 'birth_certificate_mimetype',
            'application_status', 'submitted_at', 'last_updated_at', 'exam_status'
        ]
        
        final_insert_vals = []
        now_dt = datetime.datetime.now()
        application_status_to_set = 'Enrolling' if student_type == 'old' else 'Pending'

        val_map = {
            'student_user_id': student_user_id, 'control_number': new_control_number, 'permit_control_no': None,
            'old_student_id': old_student_id if student_type == 'old' else None, 'academic_year': active_term['year_name'],
            'photo': files_data.get('photo', {}).get('data'), 'photo_mimetype': files_data.get('photo', {}).get('mimetype'),
            'shs_diploma_file': files_data.get('shs_diploma_file_input', {}).get('data'),
            'shs_diploma_filename': files_data.get('shs_diploma_file_input', {}).get('filename'),
            'shs_diploma_mimetype': files_data.get('shs_diploma_file_input', {}).get('mimetype'),
            'shs_card_file': files_data.get('shs_card_file_input', {}).get('data'),
            'shs_card_filename': files_data.get('shs_card_file_input', {}).get('filename'),
            'shs_card_mimetype': files_data.get('shs_card_file_input', {}).get('mimetype'),
            'birth_certificate_file': files_data.get('birth_certificate_file_input', {}).get('data'),
            'birth_certificate_filename': files_data.get('birth_certificate_file_input', {}).get('filename'),
            'birth_certificate_mimetype': files_data.get('birth_certificate_file_input', {}).get('mimetype'),
            'application_status': application_status_to_set, 'submitted_at': now_dt, 'last_updated_at': now_dt,
            'exam_status': None
        }
        
        for col in db_cols:
            final_insert_vals.append(val_map.get(col, form_data.get(col)))

        if len(final_insert_vals) != len(db_cols):
             raise Exception(f"Internal Error: Column count mismatch ({len(final_insert_vals)} vs {len(db_cols)}).")

        query = f"INSERT INTO applicants (`{ '`, `'.join(db_cols) }`) VALUES ({ ', '.join(['%s']*len(db_cols)) })"
        cursor.execute(query, tuple(final_insert_vals))
        new_app_id = cursor.lastrowid
        
        conn.commit()
        
        email_notif_msg = ""
        if temp_pass and student_user_id:
            full_name = f"{form_data.get('first_name','')} {form_data.get('last_name','_')}".strip()
            email_sent = send_admin_created_account_email(applicant_email, full_name, temp_pass)
            email_notif_msg = " Account credentials email sent." if email_sent else " Failed to send credentials email."
        
        # MODIFIED: Use new_control_number for display
        display_id = new_control_number or old_student_id or f"App #{new_app_id}"
        final_message = f"Application {display_id} added. {acc_msg}{email_notif_msg}"
        if not temp_pass and existing_user:
            final_message += " No email was sent because the student is an existing user."

        return jsonify({"success": True, "message": final_message})

        

    except Exception as e:
        if conn: conn.rollback()
        print(f"Unexpected Error (Admin Add App): {e}"); traceback.print_exc()
        return jsonify({"success": False, "message": f"Unexpected error: {e}"}), 500
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()

@auth.route('/admin', methods=['GET'])
def admin():
    return render_template("admin.html")

@auth.route('/logout', methods=['GET'])
def logout():
    session.clear()
    flash("✅ You have been logged out.", "success")
    return redirect(url_for('views.home'))

# --- IN auth.py ---

@auth.route('/admin_login', methods=['GET', 'POST'])
def admin_login():
    # --- Handle GET requests ---
    if request.method == 'GET':
        return redirect(url_for('auth.admin'))

    username = request.form.get('username')
    password = request.form.get('password')

    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # 1. Check in system_users table
        cursor.execute("SELECT * FROM system_users WHERE username = %s", (username,))
        user = cursor.fetchone()

        # 2. FALLBACK / INITIAL SETUP (For first-time deployment)
        if not user and username == 'admin' and password == 'adminpgpc@2025':
             hashed = generate_password_hash('adminpgpc@2025', method='pbkdf2:sha256')
             cursor.execute("INSERT INTO system_users (username, password, role, full_name) VALUES (%s, %s, %s, %s)", 
                           ('admin', hashed, 'admin', 'System Administrator'))
             conn.commit()

             session['user_role'] = 'admin'
             session['user_name'] = 'System Administrator'
             session['user_id'] = cursor.lastrowid # Capture ID of new user
             session['can_edit_student'] = True # Admin always has edit rights
             
             flash("⚠️ Initial Admin account created successfully. You are logged in.", "success")
             return redirect(url_for('auth.admin_dashboard'))

        # 3. STANDARD LOGIN
        if user and check_password_hash(user['password'], password):
            session['user_role'] = user['role']
            session['user_id'] = user['id']
            session['user_name'] = user['full_name']
            
            # Store permission in session (Default True for Admin, others check DB)
            if user['role'] == 'admin':
                session['can_edit_student'] = True
            else:
                # 1 = True, 0 = False in MySQL
                session['can_edit_student'] = bool(user.get('can_edit', 1)) 
    
            flash(f"✅ Welcome back, {user['full_name']}!", "success")
            
            # --- REDIRECT LOGIC BASED ON ROLE ---
            
            # Cashier & Accounting -> Cashier Dashboard
            if user['role'] in ['cashier', 'accounting']:
                return redirect(url_for('auth.cashier_dashboard'))
            
            # Guidance & OSA -> Enrolled Students List
            elif user['role'] in ['guidance', 'osa']:
                return redirect(url_for('auth.admin_enrolled_applications'))
            
            # Secretary -> Content Management (No access to student data)
            elif user['role'] == 'secretary':
                return redirect(url_for('auth.admin_manage_content'))
            
            # President -> President Dashboard
            elif user['role'] == 'president':
                return redirect(url_for('auth.president_dashboard'))
                
            # Admin & Registrar -> Main Admin Dashboard
            else:
                return redirect(url_for('auth.admin_dashboard'))
        else:
            flash('Invalid credentials', "danger")
            return redirect(url_for('auth.admin'))
            
    except Exception as e:
        flash(f"Login error: {e}", "danger")
        print(f"Login Error: {e}") 
        return redirect(url_for('auth.admin'))
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

@auth.route('/admin/manage-users')
def admin_manage_users():
    if session.get('user_role') != 'admin':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('auth.admin_dashboard'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # --- FIX IS HERE: Added ", can_edit" to the SELECT statement ---
    cursor.execute("""
        SELECT id, username, role, full_name, created_at, can_edit 
        FROM system_users 
        ORDER BY role, username
    """)
    
    users = cursor.fetchall()
    cursor.close()
    conn.close()
    
    _, stats = _get_all_applications_and_stats()
    
    return render_template('admin_manage_users.html', users=users, stats=stats, active_page='manage_users')


# 1. UPDATE: admin_add_user (Enforce Single Admin Rule)
@auth.route('/admin/users/add', methods=['POST'])
def admin_add_user():
    if session.get('user_role') != 'admin': return redirect(url_for('auth.admin'))
    
    username = request.form.get('username')
    password = request.form.get('password')
    full_name = request.form.get('full_name')
    role = request.form.get('role')

    if not all([username, password, full_name, role]):
        flash("All fields are required.", "danger")
        return redirect(url_for('auth.admin_manage_users'))

    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # --- ENFORCE SINGLE ADMIN RULE ---
        if role == 'admin':
            cursor.execute("SELECT id FROM system_users WHERE role = 'admin'")
            existing_admin = cursor.fetchone()
            if existing_admin:
                flash("Operation Failed: There can only be one Administrator account.", "danger")
                return redirect(url_for('auth.admin_manage_users'))
        # ---------------------------------

        hashed_password = generate_password_hash(password, method='pbkdf2:sha256')
        cursor.execute("INSERT INTO system_users (username, password, full_name, role) VALUES (%s, %s, %s, %s)", 
                       (username, hashed_password, full_name, role))
        conn.commit()
        flash(f"User {username} ({role}) added successfully.", "success")

    except mysql.connector.Error as err:
        if err.errno == 1062:
            flash("Username already exists.", "danger")
        else:
            flash(f"Database error: {err}", "danger")
    finally:
        if cursor: cursor.close()
        if conn: conn.close()
        
    return redirect(url_for('auth.admin_manage_users'))

# 2. UPDATE: admin_delete_user (Protect Admin Account)
@auth.route('/admin/users/delete/<int:user_id>', methods=['POST'])
def admin_delete_user(user_id):
    if session.get('user_role') != 'admin': return redirect(url_for('auth.admin'))
    
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # Check user role before deleting
        cursor.execute("SELECT role FROM system_users WHERE id = %s", (user_id,))
        target_user = cursor.fetchone()

        if target_user and target_user['role'] == 'admin':
            flash("Critical Action Blocked: The main Administrator account cannot be deleted.", "danger")
            return redirect(url_for('auth.admin_manage_users'))

        # Proceed with delete
        cursor.execute("DELETE FROM system_users WHERE id = %s", (user_id,))
        conn.commit()
        flash("User deleted successfully.", "success")

    except Exception as e:
        flash(f"Error deleting user: {e}", "danger")
    finally:
        if cursor: cursor.close()
        if conn: conn.close()
    
    return redirect(url_for('auth.admin_manage_users'))

@auth.route('/admin/users/edit', methods=['POST'])
def admin_edit_user():
    if session.get('user_role') != 'admin': return redirect(url_for('auth.admin'))
    
    user_id = request.form.get('user_id')
    full_name = request.form.get('full_name')
    username = request.form.get('username')
    role = request.form.get('role')

    if not all([user_id, full_name, username, role]):
        flash("Missing required fields.", "danger")
        return redirect(url_for('auth.admin_manage_users'))

    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # Prevent editing the main Admin's role/username to something else (Self-lockout protection)
        cursor.execute("SELECT role FROM system_users WHERE id = %s", (user_id,))
        current_user_data = cursor.fetchone()
        
        if current_user_data and current_user_data['role'] == 'admin':
            # If editing the admin, force role to stay admin to prevent accidental demotion
            role = 'admin' 
            
        # Check for duplicate username (exclude current user)
        cursor.execute("SELECT id FROM system_users WHERE username = %s AND id != %s", (username, user_id))
        if cursor.fetchone():
            flash(f"Username '{username}' is already taken.", "danger")
            return redirect(url_for('auth.admin_manage_users'))

        # Execute Update
        cursor.execute("""
            UPDATE system_users 
            SET full_name = %s, username = %s, role = %s 
            WHERE id = %s
        """, (full_name, username, role, user_id))
        
        conn.commit()
        flash(f"User '{username}' updated successfully.", "success")

    except Exception as e:
        flash(f"Error updating user: {e}", "danger")
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

    return redirect(url_for('auth.admin_manage_users'))

# 3. NEW ROUTE: Change User Password (Admin Action)
@auth.route('/admin/users/change-password', methods=['POST'])
def admin_change_user_password():
    if session.get('user_role') != 'admin': return redirect(url_for('auth.admin'))

    user_id = request.form.get('user_id')
    new_password = request.form.get('new_password')

    if not user_id or not new_password:
        flash("Missing data for password reset.", "danger")
        return redirect(url_for('auth.admin_manage_users'))
    
    if len(new_password) < 6:
        flash("Password must be at least 6 characters.", "warning")
        return redirect(url_for('auth.admin_manage_users'))

    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        hashed_password = generate_password_hash(new_password, method='pbkdf2:sha256')
        
        cursor.execute("UPDATE system_users SET password = %s WHERE id = %s", (hashed_password, user_id))
        conn.commit()
        
        flash("User password updated successfully.", "success")

    except Exception as e:
        flash(f"Error updating password: {e}", "danger")
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

    return redirect(url_for('auth.admin_manage_users'))


@auth.route('/admin/application/<int:applicant_id>/print', methods=['GET'])
def admin_print_application_form(applicant_id):
    if session.get('user_role') not in ['admin', 'registrar', 'guidance', 'osa']:
        flash("⚠️ Please log in to access this page.", "warning")
        return redirect(url_for('auth.admin'))
    conn = None; cursor = None
    try:
        conn = get_db_connection()
        if not conn: flash("Database connection error.", "danger"); return redirect(url_for('auth.admin_dashboard'))
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT a.*, su.email as student_account_email FROM applicants a LEFT JOIN student_users su ON a.student_user_id = su.id WHERE a.applicant_id = %s", (applicant_id,))
        app_data = cursor.fetchone()
        if not app_data: flash("Application not found.", "danger"); return redirect(url_for('auth.admin_dashboard'))

        application = app_data.copy()
        string_fields = [
            'program_choice', 'last_name', 'first_name', 'middle_name', 'place_of_birth', 'sex', 'civil_status', 
            'religion', 'citizenship', 'mobile_number', 'email_address', 'permanent_address_street_barangay', 
            'permanent_address_city_municipality', 'permanent_address_province', 'permanent_address_postal_code',
            'cultural_minority_group', 'physical_disability', 'average_family_income', 'father_name', 
            'father_occupation', 'father_company_address', 'father_contact_number', 'mother_maiden_name', 
            'mother_occupation', 'mother_company_address', 'mother_contact_number', 'guardian_name', 
            'guardian_occupation', 'guardian_company_address', 'guardian_contact_number', 'senior_high_school', 
            'senior_high_school_address', 'senior_high_school_track_strand', 'senior_high_school_year_from', 
            'senior_high_school_year_to', 'tertiary_school', 'tertiary_school_address', 'tertiary_course', 
            'tertiary_year_from', 'tertiary_year_to', 'agreements', 'academic_year', 'permit_control_no', 
            'permit_exam_time', 'permit_testing_room', 'control_number', 'application_status', 'exam_status', 'admin_notes'
        ]
        for key in string_fields:
            if key in application and isinstance(application[key], bytes):
                try: application[key] = application[key].decode('utf-8')
                except UnicodeDecodeError: application[key] = f"Undecodable ({len(application[key])} bytes)"
            elif key in application and application[key] is not None and not isinstance(application[key], (str, int, float, datetime.date, datetime.datetime)):
                 application[key] = str(application[key])

        if application.get('photo') and isinstance(application['photo'], bytes):
            fmt = "jpeg"
            if application['photo'].startswith(b'\x89PNG\r\n\x1a\n'): fmt = "png"
            application['photo_base64'] = f"data:image/{fmt};base64,{base64.b64encode(application['photo']).decode('utf-8')}"
        else: application['photo_base64'] = None

        date_fields_display = ['date_of_birth', 'date_of_application', 'final_submission_date', 'permit_exam_date']
        datetime_fields_display = ['submitted_at', 'decision_date', 'last_updated_at']
        for field in date_fields_display + datetime_fields_display:
            val = application.get(field)
            if isinstance(val, (datetime.datetime, datetime.date)):
                application[field + '_formatted'] = val.strftime('%B %d, %Y' if field in date_fields_display else '%Y-%m-%d %I:%M %p')
            elif val: application[field + '_formatted'] = str(val)
            else: application[field + '_formatted'] = 'N/A'
        
        all_expected_for_template = string_fields + date_fields_display + datetime_fields_display
        for key in all_expected_for_template:
            application.setdefault(key, '')
            application.setdefault(key + '_formatted', 'N/A' if application[key] == '' else str(application[key]))

        return render_template('printable_application_form.html', application=application)
    except Exception as e:
        flash(f"Error printing application: {e}", "danger"); traceback.print_exc()
        return redirect(url_for('auth.admin_dashboard'))
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()



@auth.route('/admin/inventory/<int:applicant_id>/print', methods=['GET'])
def admin_print_inventory_form(applicant_id):
    if session.get('user_role') not in ['admin', 'registrar', 'guidance', 'osa']:
        flash("⚠️ Please log in to access this page.", "warning")
        return redirect(url_for('auth.admin'))
    
    conn = None; cursor = None
    try:
        conn = get_db_connection()
        if not conn: 
            flash("Database connection error.", "danger")
            return redirect(url_for('auth.admin_dashboard'))
            
        cursor = conn.cursor(dictionary=True)
        # Fetch all data, including inventory fields
        cursor.execute("SELECT * FROM applicants WHERE applicant_id = %s", (applicant_id,))
        application_data = cursor.fetchone()
        
        if not application_data: 
            flash("Application not found.", "danger")
            return redirect(url_for('auth.admin_dashboard'))

        # The new template will handle displaying the inventory fields
        return render_template('printable_inventory_form.html', application=application_data)
        
    except Exception as e:
        flash(f"Error generating inventory form: {e}", "danger")
        traceback.print_exc()
        return redirect(url_for('auth.admin_dashboard'))
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()

@auth.route('/admin/application/view/<int:applicant_id>')
def admin_view_application_page(applicant_id):
    if session.get('user_role') not in ['admin', 'registrar', 'guidance', 'osa']:
        flash("⚠️ Please log in to access this page.", "warning")
        return redirect(url_for('auth.admin'))
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        if not conn:
            flash("Database connection error.", "danger")
            return redirect(url_for('auth.admin_dashboard'))
        
        cursor = conn.cursor(dictionary=True)
        
        # Modified query to fetch section_name via JOIN
        cursor.execute("""
            SELECT 
                a.*, 
                su.email as student_account_email,
                s.section_name
            FROM applicants a 
            LEFT JOIN student_users su ON a.student_user_id = su.id 
            LEFT JOIN sections s ON a.section_id = s.id
            WHERE a.applicant_id = %s
        """, (applicant_id,))
        app_data = cursor.fetchone()
        
        if not app_data:
            flash("Application not found.", "danger")
            return redirect(url_for('auth.admin_dashboard'))

        # Process data for template
        application = app_data.copy()

        # --- Check for Academic Status (Regular/Irregular) ---
        student_academic_status = None
        # Check for academic status if student is enrolled, or has been enrolled before (has grades).
        if application.get('student_user_id'):
            cursor.execute("SELECT 1 FROM student_grades WHERE student_user_id = %s LIMIT 1", (application.get('student_user_id'),))
            has_grades = cursor.fetchone()
            if application.get('application_status') == 'Enrolled' or has_grades:
                student_academic_status, _ = _get_student_status_and_failed_subjects(cursor, application['student_user_id'])
        
        # Define all columns that contain binary file data to prevent them from being decoded as text.
        file_columns = {
            'photo', 'shs_diploma_file', 'shs_card_file', 'birth_certificate_file',
            'enrollment_good_moral_file', 'enrollment_photos_2x2_file', 'enrollment_entrance_fee_proof_file',
            'enrollment_voters_id_file', 'enrollment_cbs_file', 'enrollment_brgy_cert_file',
            'enrollment_psa_birth_file', 
            'pwd_id_file'  
        }

        for key, value in application.items():
            # Only attempt to decode byte strings that are NOT file data columns.
            if isinstance(value, bytes) and key not in file_columns:
                try:
                    application[key] = value.decode('utf-8')
                except UnicodeDecodeError:
                    application[key] = "Binary Data"
        
        if application.get('photo') and isinstance(application['photo'], bytes):
            application['photo_base64'] = f"data:image/jpeg;base64,{base64.b64encode(application['photo']).decode('utf-8')}"
        else:
            application['photo_base64'] = None
        
        # Format dates for display
        date_fields_to_format = ['date_of_birth', 'date_of_application', 'final_submission_date', 'permit_exam_date']
        datetime_fields_to_format = ['submitted_at', 'decision_date', 'last_updated_at']

        for field in date_fields_to_format:
            if application.get(field) and isinstance(application.get(field), (datetime.date, datetime.datetime)):
                application[field + '_formatted'] = application[field].strftime('%B %d, %Y')
        
        for field in datetime_fields_to_format:
            if application.get(field) and isinstance(application.get(field), (datetime.date, datetime.datetime)):
                 application[field + '_formatted'] = application[field].strftime('%Y-%m-%d %I:%M %p')

        # Pass stats to the base template for sidebar rendering
        _, stats = _get_all_applications_and_stats()

        # NEW: Fetch list of programs for the edit modal
        programs = _get_program_list()

        return render_template('admin_view_application_layout.html', 
                               application=application, 
                               stats=stats, 
                               programs=programs, # Pass programs list here
                               student_academic_status=student_academic_status, 
                               active_page='all_applications')
    except Exception as e:
        flash(f"Error viewing application: {e}", "danger")
        traceback.print_exc()
        return redirect(url_for('auth.admin_dashboard'))
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()

# ----------------- STUDENT AUTHENTICATION ROUTES -----------------

# NEW: Endpoint for checking if an old_student_id exists
@auth.route('/check-student-id', methods=['POST'])
def check_student_id_exists():
    student_id = request.json.get('student_id')
    if not student_id:
        return jsonify({"exists": False})

    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({"error": "Database connection failed"}), 500
        
        cursor = conn.cursor(dictionary=True)
        
        # MODIFIED: Only return true if the ID exists AND the account is verified
        cursor.execute("SELECT id FROM student_users WHERE old_student_id = %s AND is_verified = 1", (student_id,))
        
        if cursor.fetchone():
            return jsonify({"exists": True})
        else:
            return jsonify({"exists": False})

    except Exception as e:
        print(f"Error checking student ID existence: {e}")
        return jsonify({"error": "Server error during student ID check"}), 500
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()

@auth.route('/create-student-account', methods=['GET', 'POST'])
def create_student_account_page(): 
    if 'student_id' in session: return redirect(url_for('views.application_status_page'))
    
    # 1. Check if Account Creation is Enabled (Global Switch)
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT setting_value FROM system_settings WHERE setting_key = 'enable_account_creation'")
    acct_setting = cursor.fetchone()
    # Default to True if not set
    is_account_creation_open = acct_setting['setting_value'] == 'true' if acct_setting else True
    
    # Also check specific admission toggles for logic inside POST
    cursor.execute("SELECT setting_value FROM system_settings WHERE setting_key = 'admission_open_new'")
    new_open_row = cursor.fetchone()
    is_new_open = new_open_row['setting_value'] == 'true' if new_open_row else True
    
    cursor.close()
    conn.close()

    # If Global switch is OFF, block POST actions
    if not is_account_creation_open:
        if request.method == 'POST':
             flash("Account creation is currently disabled by the administrator.", "danger")
             return redirect(url_for('views.home'))
        # If GET, the template will render the "Closed" message instead of the form

    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        student_type = request.form.get('student_type')
        old_student_id = request.form.get('old_student_id')

        # --- LOGIC: Block specific types if closed ---
        if not is_new_open and student_type == 'new':
            flash("New student admission is currently closed. Only old/existing students can register at this time.", "warning")
            return redirect(url_for('auth.create_student_account_page'))
        # ---------------------------------------------

        if not all([email, password, confirm_password, student_type]):
            flash('Please fill out all fields.', 'warning')
            return redirect(url_for('auth.create_student_account_page'))
        
        conn = None; cursor = None
        try:
            conn = get_db_connection()
            if not conn: flash("Database error.", "danger"); return redirect(url_for('auth.create_student_account_page'))
            cursor = conn.cursor(dictionary=True)

            # --- OLD STUDENT ID LOGIC ---
            if student_type == 'old':
                if not old_student_id or not old_student_id.strip():
                    flash('Please provide your Old Student ID.', 'danger')
                    return redirect(url_for('auth.create_student_account_page'))
                
                # Check if ID is taken
                cursor.execute("SELECT id, is_verified, email FROM student_users WHERE old_student_id = %s", (old_student_id,))
                id_holder = cursor.fetchone()

                if id_holder:
                    if id_holder['is_verified']:
                        # Case 1: ID is already permanently registered. BLOCK.
                        flash('This Old Student ID is already registered and verified. Please log in.', 'danger')
                        return redirect(url_for('auth.create_student_account_page'))
                    elif id_holder['email'] != email:
                        # Case 2: ID is held by an UNVERIFIED account with a DIFFERENT email. (Stale record)
                        cursor.execute("DELETE FROM student_users WHERE id = %s", (id_holder['id'],))
                        conn.commit()
            
            if password != confirm_password: flash('Passwords do not match.', 'danger'); return redirect(url_for('auth.create_student_account_page'))
            if len(password) < 8: flash('Password must be at least 8 characters long.', 'danger'); return redirect(url_for('auth.create_student_account_page'))
            
            hashed_password = generate_password_hash(password, method='pbkdf2:sha256')
            otp_code = f"{secrets.randbelow(1000000):06d}" 
            otp_expiry = datetime.datetime.now() + timedelta(minutes=10)

            cursor.execute("SELECT email, is_verified FROM student_users WHERE email = %s", (email,))
            existing_user = cursor.fetchone()

            if existing_user:
                if existing_user['is_verified']:
                    flash('Email already registered and verified. Please log in.', 'danger')
                    return redirect(url_for('auth.student_login_page'))
                else:
                    # Account exists but not verified. Update details.
                    cursor.execute("""
                        UPDATE student_users 
                        SET password = %s, otp_code = %s, otp_expiry = %s, student_type = %s, old_student_id = %s, updated_at = NOW()
                        WHERE email = %s
                    """, (hashed_password, otp_code, otp_expiry, student_type, old_student_id if student_type == 'old' else None, email))
                    conn.commit()
                    email_sent = send_otp_email(email, otp_code)
                    if email_sent:
                        flash('Account existed but was not verified. A new OTP has been sent. Please verify.', 'info')
                    else:
                        flash('Failed to send OTP email.', 'danger')
                    session['pending_verification_email'] = email 
                    return redirect(url_for('auth.verify_otp_page'))
            
            # Create new account
            cursor.execute("""
                INSERT INTO student_users 
                (email, password, created_at, updated_at, otp_code, otp_expiry, is_verified, student_type, old_student_id) 
                VALUES (%s, %s, NOW(), NOW(), %s, %s, %s, %s, %s)
            """, (email, hashed_password, otp_code, otp_expiry, False, student_type, old_student_id if student_type == 'old' else None))
            conn.commit()
            
            email_sent = send_otp_email(email, otp_code)
            if email_sent:
                flash('✅ Account created! Please check your email for an OTP.', 'success')
            else:
                flash('Account created, but failed to send OTP.', 'warning')
            
            session['pending_verification_email'] = email 
            return redirect(url_for('auth.verify_otp_page'))

        except mysql.connector.Error as db_err:
            if db_err.errno == 1062: 
                 flash('Email or Student ID already registered. Please try logging in.', 'danger')
            else:
                flash(f'Database error creating account: {db_err}', 'danger')
            traceback.print_exc()
            return redirect(url_for('auth.create_student_account_page'))
        except Exception as e:
            flash(f'Error creating account: {e}', 'danger'); traceback.print_exc()
            return redirect(url_for('auth.create_student_account_page'))
        finally:
            if cursor: cursor.close()
            if conn and conn.is_connected(): conn.close()
            
    return render_template('create_account.html')

@auth.route('/verify-otp', methods=['GET', 'POST'])
def verify_otp_page():
    email_to_verify = session.get('pending_verification_email')
    if not email_to_verify: 
        email_to_verify = request.form.get('email_for_verification') 
        if not email_to_verify and request.args.get('email'): 
            email_to_verify = request.args.get('email')
            session['pending_verification_email'] = email_to_verify 
        elif not email_to_verify :
            flash("No email found for verification. Please start by creating an account or trying to log in.", "warning")
            return redirect(url_for('auth.create_student_account_page'))

    if request.method == 'POST':
        otp_entered = request.form.get('otp_code')
        email_from_form = request.form.get('email_for_verification')
        if email_from_form and email_from_form != email_to_verify: 
            email_to_verify = email_from_form
            session['pending_verification_email'] = email_to_verify

        if not otp_entered:
            flash("Please enter the OTP.", "warning")
            return render_template('verify_otp.html', email=email_to_verify)

        conn = None; cursor = None
        try:
            conn = get_db_connection()
            if not conn:
                flash("Database connection error.", "danger")
                return render_template('verify_otp.html', email=email_to_verify)
            
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT id, otp_code, otp_expiry, is_verified FROM student_users WHERE email = %s", (email_to_verify,))
            user = cursor.fetchone()

            if not user:
                flash("User not found. Please try creating an account again.", "danger")
                session.pop('pending_verification_email', None) 
                return redirect(url_for('auth.create_student_account_page'))
            
            if user['is_verified']:
                flash("Account already verified. You can log in.", "info")
                session.pop('pending_verification_email', None) 
                return redirect(url_for('auth.student_login_page'))

            db_otp_expiry = user['otp_expiry']
            if isinstance(db_otp_expiry, str): 
                try: db_otp_expiry = datetime.datetime.fromisoformat(db_otp_expiry)
                except ValueError:
                    flash("OTP expiry data error. Contact support.", "danger")
                    return render_template('verify_otp.html', email=email_to_verify)
            
            if not user['otp_code'] or not db_otp_expiry :
                flash("OTP not set or issue with expiry for this account. Try resending OTP.", "danger")
                return render_template('verify_otp.html', email=email_to_verify, show_resend=True)

            if db_otp_expiry < datetime.datetime.now():
                flash("OTP has expired. Please request a new one.", "danger")
                return render_template('verify_otp.html', email=email_to_verify, show_resend=True)

            if user['otp_code'] == otp_entered:
                cursor.execute("""
                    UPDATE student_users 
                    SET is_verified = TRUE, otp_code = NULL, otp_expiry = NULL, updated_at = NOW()
                    WHERE id = %s
                """, (user['id'],))
                conn.commit()
                flash("✅ Account verified successfully! You can now log in.", "success")
                session.pop('pending_verification_email', None) 
                return redirect(url_for('auth.student_login_page'))
            else:
                flash("Invalid OTP. Please try again.", "danger")
                return render_template('verify_otp.html', email=email_to_verify)

        except Exception as e:
            flash(f"An error occurred during OTP verification: {e}", "danger")
            traceback.print_exc()
            return render_template('verify_otp.html', email=email_to_verify)
        finally:
            if cursor: cursor.close()
            if conn and conn.is_connected(): conn.close()

    return render_template('verify_otp.html', email=email_to_verify)

@auth.route('/resend-otp', methods=['POST'])
def resend_otp_action():
    email_to_resend = request.form.get('email') 
    
    if not email_to_resend:
        flash("Email not provided for OTP resend.", "warning")
        email_to_resend = session.get('pending_verification_email')
        if not email_to_resend:
             return redirect(url_for('auth.create_student_account_page'))

    conn = None; cursor = None
    try:
        conn = get_db_connection()
        if not conn:
            flash("Database connection error.", "danger")
            return redirect(url_for('auth.verify_otp_page', email=email_to_resend))

        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id, is_verified FROM student_users WHERE email = %s", (email_to_resend,))
        user = cursor.fetchone()

        if not user:
            flash(f"No account found for email {email_to_resend}. Please create an account.", "danger")
            session.pop('pending_verification_email', None)
            return redirect(url_for('auth.create_student_account_page'))

        if user['is_verified']:
            flash("Account is already verified. You can log in.", "info")
            session.pop('pending_verification_email', None)
            return redirect(url_for('auth.student_login_page'))

        new_otp_code = f"{secrets.randbelow(1000000):06d}"
        new_otp_expiry = datetime.datetime.now() + timedelta(minutes=10)

        cursor.execute("""
            UPDATE student_users 
            SET otp_code = %s, otp_expiry = %s, updated_at = NOW()
            WHERE id = %s
        """, (new_otp_code, new_otp_expiry, user['id']))
        conn.commit()

        email_sent = send_otp_email(email_to_resend, new_otp_code)
        if email_sent:
            flash("A new OTP has been sent to your email address.", "success")
        else:
            flash("Failed to send new OTP. Please try again later or contact support.", "danger")
        
        session['pending_verification_email'] = email_to_resend 
        return redirect(url_for('auth.verify_otp_page')) 

    except Exception as e:
        flash(f"Error resending OTP: {e}", "danger")
        traceback.print_exc()
        if email_to_resend:
            session['pending_verification_email'] = email_to_resend
        return redirect(url_for('auth.verify_otp_page'))
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()


@auth.route('/student-login', methods=['GET', 'POST'])
def student_login_page(): 
    # If already logged in, let the views controller decide where to go
    if 'student_id' in session: return redirect(url_for('views.application_status_page'))

    if request.method == 'POST':
        email, password = request.form.get('email'), request.form.get('password')
        if not email or not password: flash('Please fill out all fields.', 'warning'); return redirect(url_for('auth.student_login_page'))
        
        conn = None; cursor = None
        try:
            conn = get_db_connection()
            if not conn: flash("Database error.", "danger"); return redirect(url_for('auth.student_login_page'))
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT id, email, password, is_verified FROM student_users WHERE email = %s", (email,))
            user = cursor.fetchone()

            if user:
                if not user['is_verified']:
                    flash('Your account is not verified. Please check your email for an OTP.', 'warning')
                    session['pending_verification_email'] = email 
                    return redirect(url_for('auth.verify_otp_page')) 

                if check_password_hash(user['password'], password):
                    session['student_id'] = user['id']
                    session['student_email'] = user['email']
                    session.pop('pending_verification_email', None) 
                    
                    flash('✅ Login successful!', 'success')

                    # --- MODIFIED REDIRECT LOGIC ---
                    # Check Application Status to determine destination
                    cursor.execute("""
                        SELECT applicant_id, application_status 
                        FROM applicants 
                        WHERE student_user_id = %s 
                        ORDER BY submitted_at DESC LIMIT 1
                    """, (user['id'],))
                    app = cursor.fetchone()

                    if not app:
                        # Case 1: No application yet -> Go to New Student Form
                        return redirect(url_for('views.new_student'))
                    
                    elif app['application_status'] == 'Passed':
                        # Case 2: Passed (Exam passed OR Old Student initial registration) -> Go to Enrollment Form
                        return redirect(url_for('views.enrollment_form_page', applicant_id=app['applicant_id']))
                    
                    else:
                        # Case 3: Standard Status (Pending, Enrolled, etc.) -> Go to Dashboard
                        return redirect(url_for('views.application_status_page'))
                    # -------------------------------

                else:
                    flash('Invalid email or password.', 'danger')
                    return redirect(url_for('auth.student_login_page'))
            else:
                flash('Invalid email or password.', 'danger')
                return redirect(url_for('auth.student_login_page'))
        except Exception as e:
            flash(f'Login error: {e}', 'danger'); traceback.print_exc()
            return redirect(url_for('auth.student_login_page'))
        finally:
            if cursor: cursor.close()
            if conn and conn.is_connected(): conn.close()
    return render_template('student_login.html')

# ----------------- STUDENT FORGOT/RESET PASSWORD ROUTES -----------------
@auth.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password_request_page(): 
    if request.method == 'POST':
        email = request.form.get('email')
        if not email: flash("Please enter your email.", "warning"); return redirect(url_for('auth.forgot_password_request_page'))
        conn = None; cursor = None
        try:
            conn = get_db_connection()
            if not conn: flash("Database error.", "danger"); return redirect(url_for('auth.forgot_password_request_page'))
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT id FROM student_users WHERE email = %s", (email,))
            user = cursor.fetchone()
            if user:
                token = secrets.token_urlsafe(32)
                expiry = datetime.datetime.now() + timedelta(hours=1)
                cursor.execute("UPDATE student_users SET reset_token = %s, reset_token_expiry = %s WHERE id = %s", (token, expiry, user['id']))
                conn.commit()
                email_sent, reset_url = send_password_reset_email(email, token)
                if email_sent: flash(f"Password reset link sent to {email}.", "success")
                else: flash("Error sending email. Try again later.", "danger"); print(f"Failed reset email to {email}. Link: {reset_url}")
            else:
                flash(f"Email '{email}' not registered.", "danger")
            return redirect(url_for('auth.forgot_password_request_page'))
        except Exception as e:
            flash(f'Error processing request: {e}', 'danger'); traceback.print_exc()
            return redirect(url_for('auth.forgot_password_request_page'))
        finally:
            if cursor: cursor.close()
            if conn and conn.is_connected(): conn.close()
    return render_template('forgot_password_request.html')

@auth.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password_page(token): 
    conn = None; cursor = None
    try:
        conn = get_db_connection()
        if not conn: flash("Database error.", "danger"); return redirect(url_for('auth.forgot_password_request_page'))
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id, reset_token_expiry FROM student_users WHERE reset_token = %s", (token,))
        user = cursor.fetchone()

        if not user: flash("Invalid token.", "danger"); return redirect(url_for('auth.forgot_password_request_page'))
        
        expiry_time = user.get('reset_token_expiry')
        if isinstance(expiry_time, str): 
            try: expiry_time = datetime.datetime.fromisoformat(expiry_time) 
            except ValueError: 
                flash("Token expiry format error.", "danger"); return redirect(url_for('auth.forgot_password_request_page'))

        if not expiry_time or expiry_time < datetime.datetime.now():
            flash("Token expired.", "danger")
            cursor.execute("UPDATE student_users SET reset_token = NULL, reset_token_expiry = NULL WHERE id = %s", (user['id'],))
            conn.commit()
            return redirect(url_for('auth.forgot_password_request_page'))

        if request.method == 'POST':
            password, confirm_pass = request.form.get('password'), request.form.get('confirm_password')
            if not all([password, confirm_pass]): flash("Fill all fields.", "warning"); return render_template('reset_password_form.html', token=token)
            if password != confirm_pass: flash("Passwords don't match.", "danger"); return render_template('reset_password_form.html', token=token)
            if len(password) < 8: flash('Password too short.', 'danger'); return render_template('reset_password_form.html', token=token)
            
            hashed_pass = generate_password_hash(password, method='pbkdf2:sha256')
            cursor.execute("UPDATE student_users SET password = %s, reset_token = NULL, reset_token_expiry = NULL, updated_at = NOW() WHERE id = %s", (hashed_pass, user['id']))
            conn.commit()
            flash('✅ Password reset! Please log in.', 'success')
            return redirect(url_for('auth.student_login_page'))
        
        return render_template('reset_password_form.html', token=token)

    except Exception as e:
        flash(f'Error resetting password: {e}', 'danger'); traceback.print_exc()
        return redirect(url_for('auth.forgot_password_request_page'))
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()

# ----------------- NEW ROUTE: STUDENT CHANGE PASSWORD (LOGGED IN) -----------------
@auth.route('/change-password', methods=['GET', 'POST'])
def change_password_page():
    if 'student_id' not in session:
        flash("⚠️ Please log in to change your password.", "warning")
        return redirect(url_for('auth.student_login_page'))

    student_id = session['student_id']
    conn = None
    cursor = None
    
    # --- FIX START: Fetch Student Info for Sidebar ---
    student_app = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM applicants WHERE student_user_id = %s ORDER BY submitted_at DESC LIMIT 1", (student_id,))
        student_app = cursor.fetchone()
    except Exception as e:
        print(f"Error fetching student info for sidebar: {e}")
    # --- FIX END ---

    if request.method == 'POST':
        current_password = request.form.get('current_password')
        new_password = request.form.get('new_password')
        confirm_new_password = request.form.get('confirm_new_password')

        if not all([current_password, new_password, confirm_new_password]):
            flash("Please fill out all fields.", "danger")
            return redirect(url_for('auth.change_password_page'))

        if new_password != confirm_new_password:
            flash("New passwords do not match.", "danger")
            return redirect(url_for('auth.change_password_page'))

        if len(new_password) < 8:
            flash("New password must be at least 8 characters long.", "danger")
            return redirect(url_for('auth.change_password_page'))

        try:
            # Re-using existing connection
            cursor.execute("SELECT password FROM student_users WHERE id = %s", (student_id,))
            user = cursor.fetchone()

            if not user or not check_password_hash(user['password'], current_password):
                flash("Incorrect current password.", "danger")
                return redirect(url_for('auth.change_password_page'))

            hashed_new_password = generate_password_hash(new_password, method='pbkdf2:sha256')
            cursor.execute("UPDATE student_users SET password = %s, updated_at = NOW() WHERE id = %s", (hashed_new_password, student_id))
            conn.commit()
            
            flash("✅ Your password has been changed successfully.", "success")
            return redirect(url_for('views.application_status_page'))

        except Exception as e:
            flash(f"An error occurred: {e}", "danger")
            traceback.print_exc()
            return redirect(url_for('auth.change_password_page'))
        finally:
            if cursor: cursor.close()
            if conn: conn.close()
    
    # GET request: Render template with application data
    return render_template('change_password.html', 
                           student_logged_in=True, 
                           application=student_app)


# ----------------- APPLICATION FORM SUBMISSION (NEW STUDENT) -----------------
@auth.route('/submit_application', methods=['POST'])
def submit_application():
    if 'student_id' not in session: flash("⚠️ Log in to submit.", "warning"); return redirect(url_for('auth.student_login_page'))
    if request.method != 'POST': return redirect(url_for('views.new_student'))
    
    active_term = _get_active_term()
    if not active_term:
        flash("⚠️ Admissions are currently closed. Cannot submit application.", "warning")
        return redirect(url_for('views.home'))

    conn = None; cursor = None
    student_user_id = session['student_id']
    MAX_PHOTO_SIZE_MB, MAX_DOC_SIZE_MB = 5, 5

    try:
        conn = get_db_connection()
        if not conn:
            flash("Database error.", "danger")
            return redirect(url_for('views.new_student'))

        conn.start_transaction()
        cursor = conn.cursor(dictionary=True)

        # 1. Process Standard Files
        files_to_upload = {
            'photo': ('2x2 Photo', MAX_PHOTO_SIZE_MB, True), 
            'shs_diploma_file_input': ('SHS Diploma', MAX_DOC_SIZE_MB, True),
            'shs_card_file_input': ('SHS Card', MAX_DOC_SIZE_MB, True),
            'birth_certificate_file_input': ('Birth Certificate', MAX_DOC_SIZE_MB, True)
        }
        processed_files = {}
        for key, (desc, max_size, is_required) in files_to_upload.items():
            file_storage = request.files.get(key)
            data, fname, mtype, err = process_uploaded_file(file_storage, desc, max_size)
            if err: flash(f"⚠️ {err}", "danger"); return redirect(url_for('views.new_student'))
            if is_required and not data: flash(f"⚠️ {desc} is required.", "danger"); return redirect(url_for('views.new_student'))
            if data: processed_files[key] = {'data': data, 'filename': fname, 'mimetype': mtype}

        # 2. NEW: Handle Conditional PWD File
        physical_disability_status = request.form.get('physical_disability')
        if physical_disability_status == 'Yes':
            pwd_file_storage = request.files.get('pwd_id_file_input')
            if pwd_file_storage and pwd_file_storage.filename:
                p_data, p_fname, p_mtype, p_err = process_uploaded_file(pwd_file_storage, "PWD/Medical Certificate", MAX_DOC_SIZE_MB)
                if p_err:
                    flash(f"⚠️ {p_err}", "danger")
                    return redirect(url_for('views.new_student'))
                processed_files['pwd_id_file_input'] = {'data': p_data, 'filename': p_fname, 'mimetype': p_mtype}
            else:
                flash("⚠️ PWD ID or Medical Certificate is required when selecting 'Yes'.", "danger")
                return redirect(url_for('views.new_student'))

        # 3. Process Text Fields
        field_list = [
            'program_choice', 'last_name', 'first_name', 'middle_name', 'date_of_birth', 'place_of_birth',
            'sex', 'civil_status', 'religion', 'citizenship', 'mobile_number', 'email_address',
            'permanent_address_street_barangay', 'permanent_address_city_municipality',
            'permanent_address_province', 'permanent_address_postal_code',
            'cultural_minority_group', 'physical_disability',
            'date_of_application', 'academic_year', 'average_family_income',
            'father_name', 'father_occupation', 'father_company_address', 'father_contact_number',
            'mother_maiden_name', 'mother_occupation', 'mother_company_address', 'mother_contact_number',
            'guardian_name', 'guardian_occupation', 'guardian_company_address', 'guardian_contact_number',
            'senior_high_school', 'senior_high_school_address', 'senior_high_school_track_strand',
            'senior_high_school_year_from', 'senior_high_school_year_to',
            'tertiary_school', 'tertiary_school_address', 'tertiary_course',
            'tertiary_year_from', 'tertiary_year_to', 'agreements', 'final_submission_date'
        ]
        # Guardian is required, Father/Mother are not.
        req_text_fields = [f for f in field_list if f not in [
            'middle_name', 
            'father_name', 'father_occupation', 'father_company_address', 'father_contact_number',
            'mother_maiden_name', 'mother_occupation', 'mother_company_address', 'mother_contact_number',
            'tertiary_school', 'tertiary_school_address', 'tertiary_course', 
            'tertiary_year_from', 'tertiary_year_to'
        ]]
        req_date_fields = ['date_of_birth', 'date_of_application', 'final_submission_date']
        
        form_data = {}
        for field in field_list:
            val = request.form.get(field)
            if field in req_text_fields and (not val or not val.strip()): flash(f"⚠️ {field.replace('_',' ').title()} is required.", "danger"); return redirect(url_for('views.new_student'))
            if field in req_date_fields and not val: flash(f"⚠️ {field.replace('_',' ').title()} date required.", "danger"); return redirect(url_for('views.new_student'))
            if field == 'agreements' and val != 'Yes': flash("⚠️ Agree to terms.", "danger"); return redirect(url_for('views.new_student'))
            form_data[field] = val.strip() if val and isinstance(val, str) else val

        
        # Check for duplicate full name
        first_name = form_data.get('first_name', '').strip().lower()
        last_name = form_data.get('last_name', '').strip().lower()
        middle_name = form_data.get('middle_name', '').strip().lower()
        
        cursor.execute("""
            SELECT applicant_id FROM applicants 
            WHERE LOWER(TRIM(first_name)) = %s 
              AND LOWER(TRIM(last_name)) = %s 
              AND LOWER(TRIM(COALESCE(middle_name, ''))) = %s
        """, (first_name, last_name, middle_name))
        
        if cursor.fetchone():
            conn.rollback()
            flash('An applicant with this full name already exists. Please contact admissions if you believe this is an error.', 'danger')
            return redirect(url_for('views.new_student'))
        
        cursor.execute("SELECT applicant_id FROM applicants WHERE student_user_id = %s AND application_status IN ('Pending', 'In Review', 'Approved', 'Passed', 'Scheduled', 'Enrolling')", (student_user_id,))
        if cursor.fetchone(): 
            conn.rollback()
            flash("⚠️ Active/Approved/Passed/Scheduled application exists.", "warning"); 
            return redirect(url_for('views.application_status_page'))

        student_type = request.form.get('student_type', 'new')
        old_student_id = request.form.get('old_student_id')
        
        # Generate Admission ID (A-series) for new students
        new_control_number = None
        if student_type == 'new':
            new_control_number = _generate_admission_id(cursor)
            if not new_control_number:
                raise Exception("Failed to generate admission ID.")
        
        application_status = 'Pending'
        if student_type == 'old':
            cursor.execute("""
                SELECT 1 FROM applicants 
                WHERE student_user_id = %s AND inventory_gender IS NOT NULL AND inventory_gender != '' 
                LIMIT 1
            """, (student_user_id,))
            has_inventory = cursor.fetchone()
            
            if has_inventory:
                application_status = 'Enrolling'
            else:
                application_status = 'Passed'
        
        db_cols_app_insert = [
            'student_user_id', 'program_choice', 'last_name', 'first_name', 'middle_name', 'date_of_birth', 'place_of_birth', 'sex', 'civil_status', 
            'religion', 'citizenship', 'mobile_number', 'email_address', 'permanent_address_street_barangay', 
            'permanent_address_city_municipality', 'permanent_address_province', 'permanent_address_postal_code', 
            'cultural_minority_group', 'physical_disability', 'control_number', 'permit_control_no', 'old_student_id',
            'date_of_application', 'academic_year', 
            'average_family_income', 'father_name', 'father_occupation', 'father_company_address', 'father_contact_number', 
            'mother_maiden_name', 'mother_occupation', 'mother_company_address', 'mother_contact_number', 'guardian_name', 
            'guardian_occupation', 'guardian_company_address', 'guardian_contact_number', 'senior_high_school', 
            'senior_high_school_address', 'senior_high_school_track_strand', 'senior_high_school_year_from', 
            'senior_high_school_year_to', 'tertiary_school', 'tertiary_school_address', 'tertiary_course', 
            'tertiary_year_from', 'tertiary_year_to', 'agreements', 'final_submission_date', 'photo', 'photo_mimetype',
            'shs_diploma_file', 'shs_diploma_filename', 'shs_diploma_mimetype', 'shs_card_file', 'shs_card_filename', 
            'shs_card_mimetype', 'birth_certificate_file', 'birth_certificate_filename', 'birth_certificate_mimetype', 
            'application_status', 'submitted_at', 'last_updated_at', 'exam_status', 'enrollment_student_type',
            'pwd_id_file', 'pwd_id_filename', 'pwd_id_mimetype'
        ]
        now_dt = datetime.datetime.now()
        
        val_map = {
            'student_user_id': student_user_id, 'control_number': new_control_number, 'permit_control_no': None,
            'old_student_id': old_student_id, 'academic_year': active_term['year_name'],
            'photo': processed_files.get('photo', {}).get('data'),
            'photo_mimetype': processed_files.get('photo', {}).get('mimetype'),
            'shs_diploma_file': processed_files.get('shs_diploma_file_input', {}).get('data'),
            'shs_diploma_filename': processed_files.get('shs_diploma_file_input', {}).get('filename'),
            'shs_diploma_mimetype': processed_files.get('shs_diploma_file_input', {}).get('mimetype'),
            'shs_card_file': processed_files.get('shs_card_file_input', {}).get('data'),
            'shs_card_filename': processed_files.get('shs_card_file_input', {}).get('filename'),
            'shs_card_mimetype': processed_files.get('shs_card_file_input', {}).get('mimetype'),
            'birth_certificate_file': processed_files.get('birth_certificate_file_input', {}).get('data'),
            'birth_certificate_filename': processed_files.get('birth_certificate_file_input', {}).get('filename'),
            'birth_certificate_mimetype': processed_files.get('birth_certificate_file_input', {}).get('mimetype'),
            'pwd_id_file': processed_files.get('pwd_id_file_input', {}).get('data'),
            'pwd_id_filename': processed_files.get('pwd_id_file_input', {}).get('filename'),
            'pwd_id_mimetype': processed_files.get('pwd_id_file_input', {}).get('mimetype'),
            'application_status': application_status, 'submitted_at': now_dt, 'last_updated_at': now_dt, 
            'exam_status': None, 'enrollment_student_type': 'Existing' if student_type == 'old' else 'New'
        }

        app_insert_vals = [val_map.get(col, form_data.get(col)) for col in db_cols_app_insert]
        
        if len(app_insert_vals) != len(db_cols_app_insert):
            raise Exception(f"Internal error: Mismatch {len(app_insert_vals)} vs {len(db_cols_app_insert)}.")

        insert_cursor = conn.cursor()
        query = f"INSERT INTO applicants (`{ '`, `'.join(db_cols_app_insert) }`) VALUES ({ ', '.join(['%s']*len(db_cols_app_insert)) })"
        insert_cursor.execute(query, tuple(app_insert_vals))
        new_applicant_id = insert_cursor.lastrowid
        insert_cursor.close()
        
        conn.commit()

        # --- OPTIONAL: Notify Admin of New Application (Respects Toggle) ---
        if _is_email_trigger_enabled('email_trigger_admin'):
            admin_email = current_app.config.get('ADMIN_EMAIL')
            if admin_email:
                applicant_name = f"{form_data.get('first_name')} {form_data.get('last_name')}"
                email_subject = f"New Student Application: {applicant_name}"
                email_body = f"""
                <h3>New Application Received</h3>
                <p><strong>Applicant:</strong> {applicant_name}</p>
                <p><strong>Program:</strong> {form_data.get('program_choice')}</p>
                <p><strong>Type:</strong> {student_type.title()}</p>
                <p>Please log in to the admin panel to review.</p>
                """
                _send_email(email_subject, [admin_email], email_body)
        # -------------------------------------------------------------------
        
        if application_status == 'Passed':
            flash("✅ Application submitted! Please complete the enrollment & inventory form to proceed.", "success")
            return redirect(url_for('auth.enrollment_form_page', applicant_id=new_applicant_id))
        else: 
            flash("✅ Application submitted! Check status now.", "success")
            return redirect(url_for('views.application_status_page'))

    except Exception as e:
        if conn: conn.rollback()
        flash(f"⚠️ Error submitting: {e}", "danger"); traceback.print_exc()
        return redirect(url_for('views.new_student'))
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()


# ----------------- STUDENT ENROLLMENT ROUTES -----------------
@auth.route('/enrollment-form/<int:applicant_id>', methods=['GET'])
def enrollment_form_page(applicant_id):
    if 'student_id' not in session:
        flash("⚠️ Please log in to access the enrollment form.", "warning")
        return redirect(url_for('auth.student_login_page'))

    student_user_id = session['student_id']
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        if not conn:
            flash("Database connection error.", "danger")
            return redirect(url_for('views.application_status_page'))
            
        cursor = conn.cursor(dictionary=True)
        # MODIFIED: Select original_enrollment_status to check for re-enrollment
        cursor.execute("""
            SELECT 
                applicant_id, first_name, last_name, middle_name, email_address, program_choice, 
                academic_year, application_status, date_of_birth, place_of_birth, sex, civil_status,
                religion, mobile_number, permanent_address_street_barangay, permanent_address_city_municipality,
                permanent_address_province, senior_high_school, senior_high_school_year_from,
                senior_high_school_year_to, father_name, father_occupation, father_contact_number,
                mother_maiden_name, mother_occupation, mother_contact_number,
                guardian_name, guardian_contact_number,
                shs_diploma_file, shs_card_file, birth_certificate_file, photo,
                original_enrollment_status, student_user_id, old_student_id, control_number
            FROM applicants 
            WHERE applicant_id = %s AND student_user_id = %s
        """, (applicant_id, student_user_id))
        application = cursor.fetchone()
        
        if not application:
            flash("Application not found or you do not have permission.", "danger")
            return redirect(url_for('views.application_status_page'))
        
        # MODIFIED LOGIC: Check if it's a re-enrollment or a first-time enrollment
        is_re_enrollment = application.get('original_enrollment_status') == 'Enrolled'

        if application['application_status'] not in ['Passed', 'Enrolling']:
            flash(f"This application is not ready for enrollment (Status: {application['application_status']}).", "warning")
            return redirect(url_for('views.application_status_page'))
            

        if is_re_enrollment:
            # Fetch programs for the dropdown
            all_programs = _get_program_list()
            
            active_term = _get_active_term()
            return render_template('re_enrollment_form.html', 
                                   application=application, 
                                   programs=all_programs, # PASS PROGRAMS HERE
                                   active_school_year=active_term['year_name'] if active_term else application.get('academic_year'),
                                   student_logged_in=True)

        # Original logic for first-time enrollment
        docs_submitted = {
            'diploma': bool(application.get('shs_diploma_file')),
            'card': bool(application.get('shs_card_file')),  # Corresponds to Form 138
            'birth_cert': bool(application.get('birth_certificate_file')),
            'photo': bool(application.get('photo'))
        }

        return render_template('enrollment_form.html', application=application, docs_submitted=docs_submitted, student_logged_in=True)

    except Exception as e:
        flash(f"An error occurred: {e}", "danger")
        traceback.print_exc()
        return redirect(url_for('views.application_status_page'))
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()

# Helper to get Program ID from Title (needed because applicants table stores Title)
def _get_program_id_by_title(cursor, title):
    cursor.execute("SELECT program_id FROM programs WHERE title = %s", (title,))
    res = cursor.fetchone()
    if res:
        if isinstance(res, dict): return res['program_id']
        return res[0]
    return None

# Helper to auto-assign a section based on capacity
def _assign_section_automatically(cursor, applicant_id, program_id, year_level):
    """
    Finds available sections for the program and year level.
    Assigns the student to the FIRST section with space (First-Come, First-Served).
    Returns the assigned section_id or None if all sections are full.
    """
    # 1. Get all active sections for this program and year, ordered by letter (A, B, C)
    cursor.execute("""
        SELECT id, section_name, max_students 
        FROM sections 
        WHERE program_id = %s AND year_level = %s AND is_active = TRUE 
        ORDER BY section_letter ASC
    """, (program_id, year_level))
    sections = cursor.fetchall()

    if not sections:
        print(f"No sections found for {program_id} {year_level}")
        return None

    for section in sections:
        # Handle both dictionary and tuple cursors for safety
        if isinstance(section, dict):
            sec_id = section['id']
            sec_name = section['section_name']
            max_students = section['max_students']
        else:
            sec_id = section[0]
            sec_name = section[1]
            max_students = section[2]

        # 2. Check current count for this section
        cursor.execute("SELECT COUNT(*) as count FROM applicants WHERE section_id = %s", (sec_id,))
        res = cursor.fetchone()
        
        current_count = 0
        if res:
            if isinstance(res, dict):
                current_count = res['count']
            else:
                current_count = res[0]

        # 3. If space exists, assign immediately (First-Come, First-Served Logic)
        if current_count < max_students:
            # Execute Update
            cursor.execute("""
                UPDATE applicants 
                SET section_id = %s, is_section_permanent = TRUE 
                WHERE applicant_id = %s
            """, (sec_id, applicant_id))
            
            print(f"Auto-assigned applicant {applicant_id} to section {sec_name} ({current_count + 1}/{max_students})")
            return sec_id
    
    print(f"All sections full for {program_id} {year_level}")
    return None

@auth.route('/admin/manage-sections', methods=['GET', 'POST'])
def admin_manage_sections():
    if session.get('user_role') not in ['admin', 'registrar']: 
        return redirect(url_for('auth.admin'))

    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # --- HANDLE POST REQUESTS ---
        if request.method == 'POST':
            action = request.form.get('action')
            
            # 1. GENERATE SECTIONS
            if action == 'generate':
                program_id = request.form.get('program_id')
                year_level = request.form.get('year_level')
                num_sections = int(request.form.get('num_sections'))
                max_students = int(request.form.get('max_students'))
                
                import string
                letters = string.ascii_uppercase[:num_sections]
                for letter in letters:
                    year_digit = year_level[0] 
                    section_name = f"{program_id} {year_digit}{letter}"
                    cursor.execute("""
                        INSERT INTO sections (program_id, year_level, section_name, section_letter, max_students)
                        VALUES (%s, %s, %s, %s, %s)
                        ON DUPLICATE KEY UPDATE max_students = %s, is_active = TRUE
                    """, (program_id, year_level, section_name, letter, max_students, max_students))
                conn.commit()
                flash(f"Generated {num_sections} sections.", "success")

            # 2. EDIT SECTION
            elif action == 'edit_section':
                section_id = request.form.get('section_id')
                max_students = request.form.get('max_students_edit')
                is_active = True if request.form.get('is_active_edit') else False
                cursor.execute("UPDATE sections SET max_students = %s, is_active = %s WHERE id = %s", (max_students, is_active, section_id))
                conn.commit()
                flash("Section updated.", "success")

            # 3. DELETE SECTION
            elif action == 'delete_section':
                section_id = request.form.get('section_id')
                cursor.execute("SELECT COUNT(*) as count FROM applicants WHERE section_id = %s", (section_id,))
                if cursor.fetchone()['count'] > 0:
                    flash("Cannot delete section with assigned students.", "danger")
                else:
                    cursor.execute("DELETE FROM sections WHERE id = %s", (section_id,))
                    conn.commit()
                    flash("Section deleted.", "success")

            # 4. EQUALIZE / DISTRIBUTE STUDENTS
            elif action == 'distribute_year_level':
                program_title = request.form.get('program_title')
                year_str = request.form.get('year_level')
                cursor.execute("SELECT program_id FROM programs WHERE title = %s", (program_title,))
                prog_row = cursor.fetchone()
                if prog_row:
                    program_id = prog_row['program_id']
                    cursor.execute("SELECT id, max_students FROM sections WHERE program_id = %s AND year_level = %s AND is_active = TRUE ORDER BY section_letter", (program_id, year_str))
                    sections = cursor.fetchall()
                    cursor.execute("SELECT applicant_id FROM applicants WHERE program_choice = %s AND enrollment_year_level = %s AND application_status IN ('Enrolling', 'Enrolled') ORDER BY last_name", (program_title, year_str))
                    all_students = cursor.fetchall()
                    if sections and all_students:
                        base_count = len(all_students) // len(sections)
                        remainder = len(all_students) % len(sections)
                        idx = 0
                        for i, sec in enumerate(sections):
                            target = base_count + (1 if i < remainder else 0)
                            for _ in range(target):
                                if idx < len(all_students):
                                    cursor.execute("UPDATE applicants SET section_id = %s WHERE applicant_id = %s", (sec['id'], all_students[idx]['applicant_id']))
                                    idx += 1
                    conn.commit()
                    flash("Students distributed equally.", "success")
            
            return redirect(url_for('auth.admin_manage_sections'))

        # --- HANDLE GET REQUEST (THE FIX IS HERE) ---
        
        # 1. Fetch Page Content & System Settings (This fixes the 'content' undefined error)
        content = {}
        cursor.execute("SELECT content_key, content_value FROM page_content")
        for row in cursor.fetchall():
            content[row['content_key']] = row['content_value']
            
        cursor.execute("SELECT setting_key, setting_value FROM system_settings")
        for row in cursor.fetchall():
            content[row['setting_key']] = row['setting_value']

        # 2. Fetch Programs for Dropdown
        cursor.execute("SELECT program_id, title FROM programs ORDER BY title")
        programs = cursor.fetchall()
        
        # 3. Fetch All Sections with Student Counts
        cursor.execute("""
            SELECT s.*, p.title as program_title,
            (SELECT COUNT(*) FROM applicants a WHERE a.section_id = s.id) as current_students
            FROM sections s
            JOIN programs p ON s.program_id = p.program_id
            ORDER BY p.title, s.year_level, s.section_letter
        """)
        all_sections = cursor.fetchall()
        
        # 4. Group Data for UI
        sections_by_program_year = defaultdict(lambda: defaultdict(list))
        for sec in all_sections:
            sections_by_program_year[sec['program_title']][sec['year_level']].append(sec)

        # 5. Get Sidebar Stats
        _, stats = _get_all_applications_and_stats() 

        return render_template('admin_manage_sections.html', 
                               programs=programs, 
                               sections_by_program_year=sections_by_program_year,
                               content=content, # <--- Passing the dictionary to the template
                               stats=stats, 
                               active_page='manage_sections')

    except Exception as e:
        print(f"Error in manage sections: {e}")
        traceback.print_exc()
        flash(f"Error: {e}", "danger")
        return redirect(url_for('auth.admin_dashboard'))
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

@auth.route('/api/get-sections')
def api_get_sections():
    program_title = request.args.get('program')
    year_level = request.args.get('year')
    
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    cursor.execute("SELECT program_id FROM programs WHERE title = %s", (program_title,))
    res = cursor.fetchone()
    
    sections = []
    if res:
        prog_id = res['program_id']
        # Modified query to include current count
        cursor.execute("""
            SELECT s.id, s.section_name, s.max_students,
            (SELECT COUNT(*) FROM applicants a WHERE a.section_id = s.id) as current_count
            FROM sections s
            WHERE s.program_id = %s AND s.year_level = %s AND s.is_active = TRUE 
            ORDER BY s.section_name
        """, (prog_id, year_level))
        sections = cursor.fetchall()
    
    cursor.close()
    conn.close()
    return jsonify(sections)

# 2. Update the submit_enrollment function
@auth.route('/submit-enrollment/<int:applicant_id>', methods=['POST'])
def submit_enrollment(applicant_id):
    if 'student_id' not in session:
        flash("⚠️ Please log in to submit enrollment.", "warning")
        return redirect(url_for('auth.student_login_page'))

    student_user_id = session['student_id']
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Fetch Application Data
        cursor.execute("SELECT * FROM applicants WHERE applicant_id = %s AND student_user_id = %s", (applicant_id, student_user_id))
        application = cursor.fetchone()
        
        if not application:
            flash("Application not found or unauthorized.", "danger")
            return redirect(url_for('views.application_status_page'))

        # Check User Type (New vs Old)
        cursor.execute("SELECT student_type FROM student_users WHERE id = %s", (student_user_id,))
        user_data = cursor.fetchone()
        student_type_val = user_data['student_type'] if user_data else 'new'
        
        # Get Basic Form Data
        year_level = request.form.get('year_level')
        semester = request.form.get('semester')
        program_choice = request.form.get('program_choice')
        manual_section_id = request.form.get('section_id') 

        update_clauses = []
        update_values = []

        # --- 1. HANDLE DOCUMENTS (Existing Logic) ---
        enrollment_files_map = {
            'enrollment_shs_card_file_input': 'enrollment_shs_card',
            'enrollment_shs_diploma_file_input': 'enrollment_shs_diploma',
            'enrollment_good_moral_file_input': 'enrollment_good_moral',
            'enrollment_psa_birth_file_input': 'enrollment_psa_birth',
            'enrollment_photos_2x2_file_input': 'enrollment_photos_2x2',
            'enrollment_entrance_fee_proof_file_input': 'enrollment_entrance_fee_proof',
            'enrollment_voters_id_file_input': 'enrollment_voters_id',
            'enrollment_cbs_file_input': 'enrollment_cbs',
            'enrollment_brgy_cert_file_input': 'enrollment_brgy_cert'
        }
        
        for form_name, db_prefix in enrollment_files_map.items():
            file_storage = request.files.get(form_name)
            if file_storage and file_storage.filename:
                file_data, filename, mimetype, error = process_uploaded_file(file_storage, "Document") 
                if not error:
                    update_clauses.extend([f"{db_prefix}_file = %s", f"{db_prefix}_filename = %s", f"{db_prefix}_mimetype = %s"])
                    update_values.extend([file_data, filename, mimetype])

        # --- 2. HANDLE INVENTORY FIELDS (NEW CODE) ---
        # Map DB Column Name : HTML Form Input Name
        inventory_map = {
            'inventory_gender': 'gender',
            'inventory_age': 'age',
            'inventory_religion': 'religion',
            'inventory_complete_address': 'complete_address',
            'inventory_mobile_number': 'mobile_number',
            'inventory_facebook_account': 'facebook_account',
            'inventory_interest_hobbies': 'interest_hobbies',
            'inventory_health_condition': 'health_condition',
            
            # Education
            'inventory_pre_elementary_school': 'pre_elementary_school',
            'inventory_pre_elementary_dates': 'pre_elementary_dates',
            'inventory_pre_elementary_awards': 'pre_elementary_awards',
            'inventory_elementary_school': 'elementary_school',
            'inventory_elementary_dates': 'elementary_dates',
            'inventory_elementary_awards': 'elementary_awards',
            'inventory_secondary_school': 'secondary_school',
            'inventory_secondary_dates': 'secondary_dates',
            'inventory_secondary_awards': 'secondary_awards',
            'inventory_vocational_school': 'vocational_school',
            'inventory_vocational_dates': 'vocational_dates',
            'inventory_vocational_awards': 'vocational_awards',

            # Family (Father)
            'inventory_father_name': 'father_name',
            'inventory_father_age': 'father_age',
            'inventory_father_status': 'father_status',
            'inventory_father_education': 'father_education',
            'inventory_father_occupation': 'father_occupation',
            'inventory_father_contact': 'father_contact',

            # Family (Mother)
            'inventory_mother_name': 'mother_name',
            'inventory_mother_age': 'mother_age',
            'inventory_mother_status': 'mother_status',
            'inventory_mother_education': 'mother_education',
            'inventory_mother_occupation': 'mother_occupation',
            'inventory_mother_contact': 'mother_contact',

            # Family (General)
            'inventory_parents_marital_status': 'parents_marital_status',
            'inventory_number_of_children': 'number_of_children',
            'inventory_number_of_brothers': 'number_of_brothers',
            'inventory_number_of_sisters': 'number_of_sisters',
            'inventory_guardian_name': 'guardian_name',
            'inventory_guardian_relationship': 'guardian_relationship',
            'inventory_guardian_contact': 'guardian_contact',
            'inventory_family_income': 'family_income',
            'inventory_favorite_colors': 'favorite_colors',
            'inventory_favorite_sports': 'favorite_sports',
            'inventory_favorite_foods': 'favorite_foods',
            'inventory_family_description': 'family_description',

            # Emergency
            'inventory_emergency_contact_name': 'emergency_contact_name',
            'inventory_emergency_contact_number': 'emergency_contact_number',
            'inventory_emergency_contact_relationship': 'emergency_contact_relationship'
        }

        for db_col, form_name in inventory_map.items():
            val = request.form.get(form_name)
            # We update even if empty string to allow clearing or setting data
            update_clauses.append(f"{db_col} = %s")
            update_values.append(val)

        # --- 3. SECTION LOGIC (Existing) ---
        section_id_to_update = None
        is_permanent = False
        
        if manual_section_id and student_type_val == 'old':
            section_id_to_update = manual_section_id
            is_permanent = True 
        elif student_type_val == 'new' and year_level == '1st Year':
            cursor.execute("SELECT program_id FROM programs WHERE title = %s", (program_choice,))
            prog_res = cursor.fetchone()
            if prog_res:
                program_code = prog_res['program_id']
                section_id_to_update = _assign_section_automatically(cursor, applicant_id, program_code, year_level)
                if section_id_to_update:
                    is_permanent = True

        # --- 4. CORE FIELDS (Existing) ---
        update_clauses.extend([
            "enrollment_student_type = %s",
            "enrollment_year_level = %s",
            "enrollment_semester = %s",
            "program_choice = %s", 
            "application_status = 'Enrolling'", # STATUS SET TO ENROLLING ON SUBMIT
            "last_updated_at = %s"
        ])
        
        db_student_type_str = 'Existing' if student_type_val == 'old' else 'New'
        update_values.extend([db_student_type_str, year_level, semester, program_choice, datetime.datetime.now()])

        if section_id_to_update:
            update_clauses.append("section_id = %s")
            update_values.append(section_id_to_update)
            update_clauses.append("is_section_permanent = %s")
            update_values.append(is_permanent)

        # --- 5. EXECUTE UPDATE ---
        query = f"UPDATE applicants SET {', '.join(update_clauses)} WHERE applicant_id = %s"
        update_values.append(applicant_id)
        
        update_cursor = conn.cursor()
        update_cursor.execute(query, tuple(update_values))
        conn.commit()
        update_cursor.close()

        flash("✅ Enrollment and Inventory submitted successfully!", "success")
        return redirect(url_for('views.application_status_page'))

    except Exception as e:
        if conn: conn.rollback()
        print(f"Error submitting enrollment: {e}") # Debug print
        traceback.print_exc()
        flash(f"Error: {e}", "danger")
        return redirect(url_for('views.application_status_page'))
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

@auth.route('/submit-re-enrollment/<int:applicant_id>', methods=['POST'])
def submit_re_enrollment(applicant_id):
    if 'student_id' not in session:
        return redirect(url_for('auth.student_login_page'))

    student_user_id = session['student_id']
    conn = None; cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("SELECT * FROM applicants WHERE applicant_id = %s AND student_user_id = %s", (applicant_id, student_user_id))
        app_data = cursor.fetchone()
        active_term = _get_active_term()

        if not app_data or not active_term:
            flash("System error: Enrollment is currently unavailable.", "danger")
            return redirect(url_for('views.application_status_page'))

        # --- SHIFTING & PROGRESSION LOGIC ---
        submitted_program = request.form.get('program_choice')
        is_first_sem = active_term['semester'] == '1st Semester'
        
        final_program = app_data['program_choice']
        final_year = ""
        final_section_id = None

        # Scenario A: Shifting Program (Allowed only in 1st Semester)
        if is_first_sem and submitted_program != app_data['program_choice']:
            final_program = submitted_program
            final_year = "1st Year" # Shifters reset to 1st Year
            final_section_id = None # Shifters need new section assignment
        
        # Scenario B: Standard Progression
        else:
            final_program = app_data['program_choice']
            final_year, _, _ = _calculate_next_term(app_data['enrollment_year_level'], app_data['enrollment_semester'])
            
            # --- SECTION PROGRESSION (e.g., 1A -> 2A) ---
            if app_data['section_id']:
                cursor.execute("SELECT section_letter, program_id FROM sections WHERE id = %s", (app_data['section_id'],))
                sec_info = cursor.fetchone()
                if sec_info:
                    # Find the section in the NEXT year level with the same letter
                    cursor.execute("""
                        SELECT id FROM sections 
                        WHERE program_id = %s AND year_level = %s AND section_letter = %s AND is_active = TRUE
                    """, (sec_info['program_id'], final_year, sec_info['section_letter']))
                    match = cursor.fetchone()
                    if match: final_section_id = match['id']

        # 1. Update Database (New Level and Status: Enrolling)
        cursor.execute("""
            UPDATE applicants SET
                program_choice = %s,
                enrollment_year_level = %s,
                enrollment_semester = %s,
                section_id = %s,
                academic_year = %s,
                application_status = 'Enrolling',
                last_updated_at = NOW()
            WHERE applicant_id = %s
        """, (final_program, final_year, active_term['semester'], final_section_id, active_term['year_name'], applicant_id))
        
        # 2. IMMEDIATE CUMULATIVE FEE INJECTION
        # Re-fetch data to get the newly updated year/sem for accurate matching
        cursor.execute("SELECT * FROM applicants WHERE applicant_id = %s", (applicant_id,))
        updated_student = cursor.fetchone()
        fees_added = _auto_assign_fees(cursor, applicant_id, updated_student)

        conn.commit()
        flash(f"✅ Submission successful. {fees_added} new fees added to your ledger.", "success")
        return redirect(url_for('views.application_status_page'))

    except Exception as e:
        if conn: conn.rollback()
        flash(f"Error: {e}", "danger")
        return redirect(url_for('views.application_status_page'))
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

@auth.route('/student/change-section', methods=['POST'])
def student_change_section():
    if 'student_id' not in session:
        flash("Please log in.", "warning")
        return redirect(url_for('auth.student_login_page'))

    student_user_id = session['student_id']
    new_section_id = request.form.get('new_section_id')

    if not new_section_id:
        flash("Please select a valid section.", "warning")
        return redirect(url_for('views.application_status_page'))

    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # 1. Get Current Student Details
        cursor.execute("""
            SELECT applicant_id, program_choice, enrollment_year_level, section_id, application_status
            FROM applicants 
            WHERE student_user_id = %s
            ORDER BY submitted_at DESC LIMIT 1
        """, (student_user_id,))
        student = cursor.fetchone()

        if not student or student['application_status'] != 'Enrolled':
            flash("You must be enrolled to change sections.", "danger")
            return redirect(url_for('views.application_status_page'))

        # 2. Get Target Section Details
        cursor.execute("SELECT * FROM sections WHERE id = %s", (new_section_id,))
        target_section = cursor.fetchone()

        if not target_section:
            flash("Section not found.", "danger")
            return redirect(url_for('views.application_status_page'))

        # 3. Security Check: Does Section match Student's Program & Year?
        # We need the program_id for the student's program title
        cursor.execute("SELECT program_id FROM programs WHERE title = %s", (student['program_choice'],))
        prog_data = cursor.fetchone()
        
        if not prog_data or target_section['program_id'] != prog_data['program_id'] or target_section['year_level'] != student['enrollment_year_level']:
            flash("Invalid section selection for your course/year.", "danger")
            return redirect(url_for('views.application_status_page'))

        # 4. Capacity Check
        cursor.execute("SELECT COUNT(*) as count FROM applicants WHERE section_id = %s", (new_section_id,))
        current_count = cursor.fetchone()['count']

        if current_count >= target_section['max_students']:
            flash(f"Section {target_section['section_name']} is full.", "danger")
            return redirect(url_for('views.application_status_page'))

        # 5. Update
        cursor = conn.cursor() # Switch to non-dictionary cursor for update if needed, or stick to dict
        cursor.execute("""
            UPDATE applicants 
            SET section_id = %s, is_section_permanent = TRUE 
            WHERE applicant_id = %s
        """, (new_section_id, student['applicant_id']))
        
        conn.commit()
        flash(f"Successfully transferred to section {target_section['section_name']}.", "success")

    except Exception as e:
        print(f"Error changing section: {e}")
        flash("An error occurred while changing sections.", "danger")
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

    return redirect(url_for('views.application_status_page'))

# ----------------- ADMIN APPLICATION MANAGEMENT API ROUTES -----------------
@auth.route('/admin/application/<int:applicant_id>/status', methods=['POST'])
def admin_update_application_status(applicant_id):
    if session.get('user_role') not in ['admin', 'registrar']: return jsonify({"success": False, "message": "Unauthorized"}), 401
    new_status = request.form.get('status')
    valid_statuses = ['Pending', 'In Review', 'Approved', 'Scheduled', 'Rejected', 'Passed', 'Failed', 'Enrolling', 'Dropped', 'Enrolled', 'Not Enrolled']
    if not new_status or new_status not in valid_statuses: return jsonify({"success": False, "message": "Invalid status"}), 400

    conn = None; cursor = None 
    try:
        conn = get_db_connection()
        if not conn: return jsonify({"success": False, "message": "DB error"}), 500
        
        cursor = conn.cursor(dictionary=True)
        # MODIFIED: Added control_number and academic_year to the query
        cursor.execute("SELECT a.first_name, a.last_name, a.program_choice, a.exam_status, a.permit_control_no, su.email as student_account_email, a.email_address as application_form_email, a.old_student_id, a.control_number, a.academic_year FROM applicants a LEFT JOIN student_users su ON a.student_user_id = su.id WHERE a.applicant_id = %s", (applicant_id,))
        app_info = cursor.fetchone()
        if not app_info: return jsonify({"success": False, "message": "App not found"}), 404

        email_to_notify = app_info.get('student_account_email') or app_info.get('application_form_email')
        applicant_name = f"{app_info.get('first_name','')} {app_info.get('last_name','')}".strip()
        now_dt = datetime.datetime.now()
        
        generated_permit_control_no = app_info.get('permit_control_no') 
        new_permit_control_no_generated_flag = False

        sql_update_parts = ["application_status = %s", "last_updated_at = %s"]
        params_update = [new_status, now_dt]

        # When main status is set to Passed or Failed, sync the exam_status as well.
        if new_status == 'Passed':
            sql_update_parts.append("exam_status = %s")
            params_update.append('Passed')
        elif new_status == 'Failed':
            sql_update_parts.append("exam_status = %s")
            params_update.append('Failed')

        if new_status in ['Approved', 'Rejected', 'Passed', 'Failed', 'Dropped', 'Enrolled', 'Not Enrolled']:
            sql_update_parts.append("decision_date = %s")
            params_update.append(now_dt)
        else:
            sql_update_parts.append("decision_date = NULL") 

        if new_status == 'Approved' and not generated_permit_control_no:
            cursor.execute("SELECT MAX(CAST(permit_control_no AS UNSIGNED)) as max_pcn FROM applicants WHERE permit_control_no REGEXP '^[0-9]+$'")
            max_pcn_row = cursor.fetchone()
            next_pcn_int = 1
            if max_pcn_row and max_pcn_row['max_pcn'] is not None:
                next_pcn_int = int(max_pcn_row['max_pcn']) + 1
            generated_permit_control_no = f"{next_pcn_int:04d}" 
            new_permit_control_no_generated_flag = True
            sql_update_parts.append("permit_control_no = %s")
            params_update.append(generated_permit_control_no)
        
        if new_status in ['Pending', 'In Review', 'Rejected']:
            sql_update_parts.extend(["permit_exam_date = NULL", "permit_exam_time = NULL", "permit_testing_room = NULL"])

        params_update.append(applicant_id)
        final_sql = f"UPDATE applicants SET {', '.join(sql_update_parts)} WHERE applicant_id = %s"
        
        update_cursor = conn.cursor() 
        update_cursor.execute(final_sql, tuple(params_update))
        conn.commit()
        update_cursor.close()

        email_sent = False
        permit_details_for_email = None
        
        if new_status in ['Approved', 'Scheduled', 'Rejected', 'Passed', 'Failed', 'Dropped', 'Enrolled', 'Not Enrolled'] and email_to_notify:
            if new_status == 'Scheduled' or (new_status == 'Approved' and new_permit_control_no_generated_flag):
                details_cursor = conn.cursor(dictionary=True)
                details_cursor.execute("SELECT permit_exam_date, permit_exam_time, permit_testing_room, permit_control_no FROM applicants WHERE applicant_id = %s", (applicant_id,))
                permit_data = details_cursor.fetchone()
                details_cursor.close()
                permit_details_for_email = permit_data 
            
            cursor.execute("SELECT exam_status FROM applicants WHERE applicant_id = %s", (applicant_id,))
            latest_app_info = cursor.fetchone()
            current_exam_status = latest_app_info['exam_status'] if latest_app_info else app_info.get('exam_status')

            # MODIFIED: Pass control_number and academic_year to the email function
            email_sent = send_application_status_email(
                email_to_notify, 
                applicant_name, 
                new_status, 
                applicant_id, 
                app_info.get('program_choice'), 
                current_exam_status,
                permit_details=permit_details_for_email,
                old_student_id=app_info.get('old_student_id'),
                control_number=app_info.get('control_number'),
                academic_year=app_info.get('academic_year')
            )
        
        # MODIFIED: Updated ID formatting logic
        old_student_id = app_info.get('old_student_id')
        control_number = app_info.get('control_number')
        display_id = old_student_id or control_number or f"App #{applicant_id}"
        msg = f"Action processed for {display_id}: Status set to {new_status}."

        if new_permit_control_no_generated_flag:
             msg += f" Permit Control No. {generated_permit_control_no} assigned."
        
        if new_status in ['Approved', 'Scheduled', 'Rejected', 'Passed', 'Failed', 'Dropped', 'Enrolled', 'Not Enrolled']:
             msg += " Notification email sent." if email_sent else " Notification email failed to send."

        effective_pcn = generated_permit_control_no if new_permit_control_no_generated_flag else app_info.get('permit_control_no')
        if permit_details_for_email and permit_details_for_email.get('permit_control_no'):
            effective_pcn = permit_details_for_email.get('permit_control_no')

        return jsonify({"success": True, "message": msg, "new_status": new_status, "applicant_id": applicant_id, "permit_control_no": effective_pcn })

    except mysql.connector.Error as db_err: 
        print(f"Database error updating status for {applicant_id}: {db_err}"); traceback.print_exc()
        return jsonify({"success": False, "message": f"Database error: {db_err.msg}"}), 500
    except Exception as e:
        print(f"Error updating status for {applicant_id}: {e}"); traceback.print_exc()
        return jsonify({"success": False, "message": "Server error"}), 500
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()


@auth.route('/admin/application/<int:applicant_id>/exam-status', methods=['POST'])
def admin_update_exam_status(applicant_id):
    if session.get('user_role') not in ['admin', 'registrar']: return jsonify({"success": False, "message": "Unauthorized"}), 401
    new_exam_status = request.form.get('exam_status')
    valid_exam_statuses = ['Passed', 'Failed', 'Not Taken', None, "null", ""]
    if new_exam_status not in valid_exam_statuses: return jsonify({"success": False, "message": "Invalid exam status."}), 400
    if new_exam_status in ["null", ""]: new_exam_status = None

    conn = None; cursor = None; app_info_cursor = None
    try:
        conn = get_db_connection()
        if not conn: return jsonify({"success": False, "message": "DB error"}), 500

        app_info_cursor = conn.cursor(dictionary=True)
        app_info_cursor.execute("SELECT a.first_name, a.last_name, a.program_choice, a.application_status, su.email as student_account_email, a.email_address as application_form_email, a.old_student_id, a.control_number, a.academic_year FROM applicants a LEFT JOIN student_users su ON a.student_user_id = su.id WHERE a.applicant_id = %s", (applicant_id,))
        app_info = app_info_cursor.fetchone()
        app_info = app_info_cursor.fetchone()
        if not app_info: return jsonify({"success": False, "message": "App not found."}), 404
        
        cursor = conn.cursor()
        cursor.execute("UPDATE applicants SET exam_status = %s, last_updated_at = NOW() WHERE applicant_id = %s", (new_exam_status, applicant_id))
        conn.commit()

        if cursor.rowcount > 0:
            email_to_notify = app_info.get('student_account_email') or app_info.get('application_form_email')
            applicant_name = f"{app_info.get('first_name','')} {app_info.get('last_name','')}".strip()
            email_sent = False
            if new_exam_status in ['Passed', 'Failed'] and email_to_notify and app_info.get('application_status') not in ['Approved', 'Scheduled', 'Rejected', 'Passed', 'Failed']:
                email_sent = send_application_status_email(email_to_notify, applicant_name, app_info.get('application_status'), applicant_id, app_info.get('program_choice'), new_exam_status, old_student_id=app_info.get('old_student_id'), control_number=app_info.get('control_number'), academic_year=app_info.get('academic_year'))
            
            old_student_id = app_info.get('old_student_id')
            control_number = app_info.get('control_number')
            display_id = old_student_id or control_number or f"App #{applicant_id}"
            msg = f"Exam status for {display_id} to '{new_exam_status or 'Not Set'}'."

            if new_exam_status in ['Passed', 'Failed'] and email_to_notify and app_info.get('application_status') not in ['Approved', 'Scheduled', 'Rejected', 'Passed', 'Failed']:
                 msg += " Email sent." if email_sent else " Email failed."
            return jsonify({"success": True, "message": msg, "new_exam_status": new_exam_status, "applicant_id": applicant_id})
        else:
            cursor.execute("SELECT exam_status FROM applicants WHERE applicant_id = %s", (applicant_id,))
            db_status_row = cursor.fetchone()
            current_db_exam_status = db_status_row[0] if db_status_row else None 
            if current_db_exam_status == new_exam_status: 
                 return jsonify({"success": True, "message": f"Exam status already '{new_exam_status or 'Not Set'}'." , "new_exam_status": new_exam_status, "applicant_id": applicant_id})
            return jsonify({"success": False, "message": "App not found or status not updated."}), 404 
    except Exception as e:
        print(f"Error updating exam status for {applicant_id}: {e}"); traceback.print_exc()
        return jsonify({"success": False, "message": "Server error"}), 500
    finally:
        if app_info_cursor: app_info_cursor.close()
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()

@auth.route('/admin/application/<int:applicant_id>/delete', methods=['POST'])
def admin_delete_application(applicant_id):
    if session.get('user_role') not in ['admin', 'registrar']: return jsonify({"success": False, "message": "Unauthorized"}), 401
    conn = None; cursor = None
    try:
        conn = get_db_connection()
        if not conn: return jsonify({"success": False, "message": "DB error"}), 500
        cursor = conn.cursor()
        cursor.execute("DELETE FROM applicants WHERE applicant_id = %s", (applicant_id,))
        conn.commit()
        if cursor.rowcount > 0: return jsonify({"success": True, "message": f"Application #{applicant_id} deleted."})
        return jsonify({"success": False, "message": "App not found"}), 404
    except Exception as e:
        print(f"Error deleting {applicant_id}: {e}"); traceback.print_exc()
        return jsonify({"success": False, "message": "Server error deleting."}), 500
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()

@auth.route('/admin/applications/bulk-update-status', methods=['POST'])
def admin_bulk_update_status():
    if session.get('user_role') not in ['admin', 'registrar']:
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    
    data = request.get_json()
    ids = data.get('ids')
    action = data.get('action')
    action_type = data.get('type', 'status') # 'status' or 'exam_status'

    if not ids or not isinstance(ids, list) or not action:
        return jsonify({"success": False, "message": "Invalid request. Missing IDs or action."}), 400

    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({"success": False, "message": "Database connection error."}), 500
        
        cursor = conn.cursor(dictionary=True)
        
        id_placeholders = ', '.join(['%s'] * len(ids))

        # --- Fetch applicant data for email notifications BEFORE updating ---
        applicants_to_notify = []
        email_triggering_actions = ['Approved', 'Scheduled', 'Rejected', 'Passed', 'Failed', 'Enrolled', 'Not Enrolled']
        if action in email_triggering_actions:
            query_fetch = f"""
                SELECT a.applicant_id, a.first_name, a.last_name, a.program_choice, a.academic_year,
                       su.email as student_account_email, a.email_address as application_form_email, a.old_student_id, a.control_number
                FROM applicants a 
                LEFT JOIN student_users su ON a.student_user_id = su.id
                WHERE a.applicant_id IN ({id_placeholders})
            """
            cursor.execute(query_fetch, tuple(ids))
            applicants_to_notify = cursor.fetchall()


        # --- Perform the database update ---
        if action_type == 'status':
            valid_statuses = ['Pending', 'In Review', 'Approved', 'Rejected', 'Passed', 'Failed', 'Enrolling', 'Dropped', 'Enrolled', 'Not Enrolled']
            if action not in valid_statuses:
                return jsonify({"success": False, "message": f"Invalid status action: {action}"}), 400
            
            sql_parts = ["application_status = %s", "last_updated_at = NOW()"]
            params = [action]
            
            if action in ['Approved', 'Rejected', 'Passed', 'Failed', 'Dropped', 'Enrolled', 'Not Enrolled']:
                sql_parts.append("decision_date = NOW()")
            else:
                sql_parts.append("decision_date = NULL")

            if action == 'Passed':
                sql_parts.append("exam_status = 'Passed'")
            elif action == 'Failed':
                sql_parts.append("exam_status = 'Failed'")
            
            if action in ['Pending', 'In Review', 'Rejected']:
                sql_parts.extend(["permit_exam_date = NULL", "permit_exam_time = NULL", "permit_testing_room = NULL"])

            sql_query = f"UPDATE applicants SET {', '.join(sql_parts)} WHERE applicant_id IN ({id_placeholders})"
            params.extend(ids)
            cursor.execute(sql_query, tuple(params))

        elif action_type == 'exam_status':
            valid_exam_statuses = ['Passed', 'Failed', 'Not Taken']
            if action not in valid_exam_statuses:
                 return jsonify({"success": False, "message": f"Invalid exam status action: {action}"}), 400

            sql_query = f"UPDATE applicants SET exam_status = %s, last_updated_at = NOW() WHERE applicant_id IN ({id_placeholders})"
            cursor.execute(sql_query, (action, *ids))
            
        else:
            return jsonify({"success": False, "message": "Invalid action type."}), 400

        updated_rows = cursor.rowcount
        conn.commit()
        
        # --- Send emails after successful commit ---
        emails_sent_count = 0
        emails_failed_count = 0
        if applicants_to_notify:
            for app_data in applicants_to_notify:
                email_to_notify = app_data.get('student_account_email') or app_data.get('application_form_email')
                applicant_name = f"{app_data.get('first_name', '')} {app_data.get('last_name', '')}".strip()
                
                if email_to_notify:
                    email_sent = send_application_status_email(
                        applicant_email=email_to_notify,
                        applicant_name=applicant_name,
                        new_status=action,
                        application_id=app_data['applicant_id'],
                        program_choice=app_data.get('program_choice'),
                        old_student_id=app_data.get('old_student_id'),
                        control_number=app_data.get('control_number'),
                        academic_year=app_data.get('academic_year')
                    )
                    if email_sent:
                        emails_sent_count += 1
                    else:
                        emails_failed_count += 1

        # --- Construct final message ---
        final_message = f"{updated_rows} application(s) updated."
        if emails_sent_count > 0 or emails_failed_count > 0:
            final_message += f" {emails_sent_count} email(s) sent"
            if emails_failed_count > 0:
                final_message += f", {emails_failed_count} failed."
            else:
                final_message += "."

        return jsonify({"success": True, "message": final_message})

    except Exception as e:
        if conn: conn.rollback()
        print(f"Error during bulk update: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "message": "A server error occurred during the bulk update."}), 500
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()



@auth.route('/admin/applications/bulk-permit-update', methods=['POST'])
def admin_bulk_permit_update():
    if session.get('user_role') not in ['admin', 'registrar']:
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    
    data = request.get_json()
    ids = data.get('ids')
    details = data.get('details')

    if not ids or not isinstance(ids, list) or not details:
        return jsonify({"success": False, "message": "Invalid request. Missing IDs or details."}), 400

    exam_date = details.get('date')
    exam_time = details.get('time')
    exam_room = details.get('room')

    exam_date_obj = None
    if exam_date:
        try:
            exam_date_obj = datetime.datetime.strptime(exam_date, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({"success": False, "message": "Invalid date format. Use YYYY-MM-DD."}), 400

    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({"success": False, "message": "Database connection error."}), 500
        
        cursor = conn.cursor(dictionary=True)

        id_placeholders = ', '.join(['%s'] * len(ids))
        
        query_fetch = f"""
            SELECT a.applicant_id, a.first_name, a.last_name, a.program_choice, a.academic_year,
                   su.email as student_account_email, a.email_address as application_form_email, a.old_student_id, a.control_number
            FROM applicants a 
            LEFT JOIN student_users su ON a.student_user_id = su.id
            WHERE a.applicant_id IN ({id_placeholders})
        """
        cursor.execute(query_fetch, tuple(ids))
        applicants_to_notify = cursor.fetchall()

        update_query = f"""
            UPDATE applicants 
            SET 
                permit_exam_date = %s, 
                permit_exam_time = %s, 
                permit_testing_room = %s,
                application_status = 'Scheduled',
                exam_status = NULL,
                last_updated_at = NOW()
            WHERE applicant_id IN ({id_placeholders})
        """
        params = [exam_date_obj, exam_time, exam_room] + ids
        cursor.execute(update_query, tuple(params))
        updated_rows = cursor.rowcount
        conn.commit()

        emails_sent_count = 0
        emails_failed_count = 0
        if applicants_to_notify:
            for app_data in applicants_to_notify:
                email_to_notify = app_data.get('student_account_email') or app_data.get('application_form_email')
                applicant_name = f"{app_data.get('first_name', '')} {app_data.get('last_name', '')}".strip()
                
                if email_to_notify:
                    permit_details_for_email = {
                        'permit_exam_date': exam_date_obj,
                        'permit_exam_time': exam_time,
                        'permit_testing_room': exam_room
                    }
                    email_sent = send_application_status_email(
                        applicant_email=email_to_notify,
                        applicant_name=applicant_name,
                        new_status='Scheduled',
                        application_id=app_data['applicant_id'],
                        program_choice=app_data.get('program_choice'),
                        permit_details=permit_details_for_email,
                        old_student_id=app_data.get('old_student_id'),
                        control_number=app_data.get('control_number'),
                        academic_year=app_data.get('academic_year')
                    )
                    if email_sent: emails_sent_count += 1
                    else: emails_failed_count += 1
        
        final_message = f"{updated_rows} application(s) scheduled successfully."
        if emails_sent_count > 0 or emails_failed_count > 0:
            final_message += f" {emails_sent_count} email(s) sent"
            if emails_failed_count > 0:
                final_message += f", {emails_failed_count} failed."
            else:
                final_message += "."

        return jsonify({"success": True, "message": final_message})

    except Exception as e:
        if conn: conn.rollback()
        print(f"Error during bulk permit update: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "message": "A server error occurred during the bulk update."}), 500
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()

@auth.route('/admin/application/<int:applicant_id>/details', methods=['GET'])
def admin_get_application_details(applicant_id):
    if session.get('user_role') not in ['admin', 'president']: return jsonify({"success": False, "message": "Unauthorized"}), 401
    conn = None; cursor = None
    try:
        conn = get_db_connection()
        if not conn: return jsonify({"success": False, "message": "DB error"}), 500
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT a.*, su.email as student_account_email FROM applicants a LEFT JOIN student_users su ON a.student_user_id = su.id WHERE a.applicant_id = %s", (applicant_id,))
        app_data = cursor.fetchone()
        if app_data:
            for key, value in app_data.items():
                if isinstance(value, (datetime.datetime, datetime.date)): app_data[key] = value.isoformat()
                elif key == 'photo' and isinstance(value, bytes):
                    fmt = "jpeg"
                    if value.startswith(b'\x89PNG\r\n\x1a\n'): fmt = "png"
                    app_data[key] = f"data:image/{fmt};base64,{base64.b64encode(value).decode('utf-8')}" if value else None
                elif key in ['shs_diploma_file', 'shs_card_file', 'birth_certificate_file']:
                    app_data[key] = bool(value)
                elif isinstance(value, bytes): 
                    try: app_data[key] = value.decode('utf-8')
                    except UnicodeDecodeError: app_data[key] = f"Binary ({len(value)} bytes)"
            return jsonify({"success": True, "data": app_data})
        return jsonify({"success": False, "message": "App not found"}), 404
    except Exception as e:
        print(f"Error fetching details for {applicant_id}: {e}"); traceback.print_exc()
        return jsonify({"success": False, "message": "Server error"}), 500
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()

@auth.route('/applicant-photo/<int:applicant_id>')
def get_applicant_photo(applicant_id):
    # Updated to include registrar, guidance, and osa
    is_admin = session.get('user_role') in ['admin', 'registrar', 'guidance', 'osa']
    student_user_id = session.get('student_id')
    if not is_admin and not student_user_id:
        return "Unauthorized", 401

    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        if not conn:
            return send_file(os.path.join(current_app.static_folder, 'images', 'default_avatar.png'), mimetype='image/png')
        
        cursor = conn.cursor(dictionary=True)

        if not is_admin:
            cursor.execute("SELECT student_user_id FROM applicants WHERE applicant_id = %s", (applicant_id,))
            app_owner = cursor.fetchone()
            if not app_owner or app_owner['student_user_id'] != student_user_id:
                return "Forbidden", 403

        cursor.execute("SELECT photo, photo_mimetype FROM applicants WHERE applicant_id = %s", (applicant_id,))
        result = cursor.fetchone()

        if result and result.get('photo'):
            photo_data = result['photo']
            mimetype = result.get('photo_mimetype', 'image/jpeg') 
            return send_file(io.BytesIO(photo_data), mimetype=mimetype)
        else:
            return send_file(os.path.join(current_app.static_folder, 'images', 'default_avatar.png'), mimetype='image/png')
    except Exception as e:
        print(f"Error serving photo for applicant {applicant_id}: {e}")
        traceback.print_exc()
        return send_file(os.path.join(current_app.static_folder, 'images', 'default_avatar.png'), mimetype='image/png')
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()
@auth.route('/applicant-document/<int:applicant_id>/<doc_type>')
def get_applicant_document(applicant_id, doc_type):
    # Updated to include registrar, guidance, and osa
    is_admin = session.get('user_role') in ['admin', 'registrar', 'guidance', 'osa']
    student_user_id = session.get('student_id')
    if not is_admin and not student_user_id:
        return "Unauthorized", 401

    # Check if a download is being forced
    force_download = request.args.get('action') == 'download'

    # CORRECTED: This map now points to the correct, prefixed DB columns from your schema.
    doc_map = {
        # Initial Application Documents
        'diploma': ('shs_diploma_file', 'shs_diploma_mimetype', 'shs_diploma_filename'),
        'card': ('shs_card_file', 'shs_card_mimetype', 'shs_card_filename'),
        'birth_cert': ('birth_certificate_file', 'birth_certificate_mimetype', 'birth_certificate_filename'),
        # Enrollment Documents (now mapping to their actual prefixed DB columns)
        'enrollment_card': ('enrollment_shs_card_file', 'enrollment_shs_card_mimetype', 'enrollment_shs_card_filename'),
        'enrollment_diploma': ('enrollment_shs_diploma_file', 'enrollment_shs_diploma_mimetype', 'enrollment_shs_diploma_filename'),
        'enrollment_good_moral': ('enrollment_good_moral_file', 'enrollment_good_moral_mimetype', 'enrollment_good_moral_filename'),
        'enrollment_psa': ('enrollment_psa_birth_file', 'enrollment_psa_birth_mimetype', 'enrollment_psa_birth_filename'),
        'enrollment_photos': ('enrollment_photos_2x2_file', 'enrollment_photos_2x2_mimetype', 'enrollment_photos_2x2_filename'),
        'enrollment_fee': ('enrollment_entrance_fee_proof_file', 'enrollment_entrance_fee_proof_mimetype', 'enrollment_entrance_fee_proof_filename'),
        'enrollment_voters': ('enrollment_voters_id_file', 'enrollment_voters_id_mimetype', 'enrollment_voters_id_filename'),
        'enrollment_cbs': ('enrollment_cbs_file', 'enrollment_cbs_mimetype', 'enrollment_cbs_filename'),
        'enrollment_brgy': ('enrollment_brgy_cert_file', 'enrollment_brgy_cert_mimetype', 'enrollment_brgy_cert_filename'),
        'pwd_id': ('pwd_id_file', 'pwd_id_mimetype', 'pwd_id_filename')
    }

    if doc_type not in doc_map:
        return "Invalid document type", 404

    file_col, mime_col, fname_col = doc_map[doc_type]

    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        if not conn:
            return "Database connection error", 500
        
        cursor = conn.cursor(dictionary=True)

        if not is_admin:
            cursor.execute("SELECT student_user_id FROM applicants WHERE applicant_id = %s", (applicant_id,))
            app_owner = cursor.fetchone()
            if not app_owner or app_owner['student_user_id'] != student_user_id:
                return "Forbidden", 403

        cursor.execute(f"SELECT {file_col}, {mime_col}, {fname_col} FROM applicants WHERE applicant_id = %s", (applicant_id,))
        result = cursor.fetchone()

        if result and result.get(file_col):
            doc_data = result[file_col]
            mimetype = result.get(mime_col, 'application/octet-stream')
            filename = result.get(fname_col, f"{doc_type}_{applicant_id}.bin")

            return send_file(
                io.BytesIO(doc_data),
                mimetype=mimetype,
                as_attachment=force_download,
                download_name=filename
            )
        else:
            return "Document not found", 404
    except Exception as e:
        print(f"Error serving document '{doc_type}' for applicant {applicant_id}: {e}")
        traceback.print_exc()
        return "Server error", 500
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()

@auth.route('/my-grades')
def view_student_grades():
    if 'student_id' not in session:
        flash("⚠️ Please log in to view your grades.", "warning")
        return redirect(url_for('auth.student_login_page'))

    student_user_id = session['student_id']
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        if not conn:
            flash("Database connection error.", "danger")
            return redirect(url_for('views.application_status_page'))
        
        cursor = conn.cursor(dictionary=True)
        
        # MODIFIED: Select * so that applicant_id and other sidebar data are available
        cursor.execute("""
            SELECT *
            FROM applicants 
            WHERE student_user_id = %s
            ORDER BY submitted_at DESC
            LIMIT 1
        """, (student_user_id,))
        student_info = cursor.fetchone()

        if not student_info:
            flash("No application found for your account. Cannot display grades.", "info")
            return redirect(url_for('views.application_status_page'))
        
        # --- NEW LOGIC: Query for all existing grades and group them by term ---
        cursor.execute("""
            SELECT 
                sg.academic_year, 
                sg.semester, 
                s.subject_code, 
                s.subject_title, 
                sg.grade, 
                sg.remarks
            FROM student_grades sg
            JOIN subjects s ON sg.subject_id = s.id
            WHERE sg.student_user_id = %s
            ORDER BY sg.academic_year DESC, sg.semester ASC, s.subject_code ASC
        """, (student_user_id,))
        
        all_grades = cursor.fetchall()
        
        # Group grades by term using a dictionary
        grades_by_term = defaultdict(list)
        for grade in all_grades:
            term_key = f"{grade['academic_year']} - {grade['semester']}"
            grades_by_term[term_key].append(grade)

        return render_template('student_grades.html',
                               student=student_info,
                               application=student_info, # <--- FIX: Added this line
                               grades_by_term=grades_by_term,
                               student_logged_in=True)

    except Exception as e:
        flash(f"An error occurred while fetching grades: {e}", "danger")
        traceback.print_exc()
        return redirect(url_for('views.application_status_page'))
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()